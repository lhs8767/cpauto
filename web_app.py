from __future__ import annotations

import cgi
import base64
import html
import io
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from features.auth.pages import admin_page, login_page
from features.auth.service import AuthManager
from features.growth_incentive.service import calculate_year, load_config, save_config

BASE_DIR = Path(__file__).resolve().parent
SCRIPTS_DIR = BASE_DIR / "scripts"
RUNS_DIR = BASE_DIR / "web_runs"
MASTER_DIR = BASE_DIR / "input" / "master"
DATA_DIR = BASE_DIR / "data"
MONTHLY_SALES_PATH = DATA_DIR / "월매출_납품상품.xlsx"
SALES_LEDGER_PATH = DATA_DIR / "_월매출_내부자료.xlsx"
YEAR_MANUAL_PATH = DATA_DIR / "연도총매출_수기입력.json"
SALES_CONFIRM_PATH = DATA_DIR / "계산서_발행확인.json"
GROWTH_INCENTIVE_PATH = DATA_DIR / "성장장려금_기초자료.json"
TESTER_FILES_DIR = DATA_DIR / "체험단_자료"
TESTER_FILES_META_PATH = DATA_DIR / "체험단_파일목록.json"


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_env_file(BASE_DIR / ".env")
AUTH = AuthManager(DATA_DIR)


def clean_supabase_url(value: str) -> str:
    url = value.strip().rstrip("/")
    if url and not url.startswith(("http://", "https://")):
        url = f"https://{url}"
    return url


SUPABASE_URL = clean_supabase_url(os.environ.get("SUPABASE_URL", ""))
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()


def supabase_request(method: str, path: str, body: object | None = None) -> object | None:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        return None
    payload = json.dumps(body).encode("utf-8") if body is not None else None
    request = urllib.request.Request(f"{SUPABASE_URL}{path}", data=payload, method=method)
    request.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    request.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    request.add_header("Content-Type", "application/json")
    request.add_header("Prefer", "resolution=merge-duplicates,return=representation")
    try:
        with urllib.request.urlopen(request, timeout=15) as response:
            text = response.read().decode("utf-8")
            return json.loads(text) if text else None
    except Exception as exc:
        print(f"[supabase-files] {method} {path} failed: {exc}", flush=True)
        return None


def restore_file_from_supabase(file_key: str, target_path: Path) -> None:
    if target_path.exists():
        return
    rows = supabase_request(
        "GET",
        f"/rest/v1/app_files?select=content_base64&file_key=eq.{urllib.parse.quote(file_key)}",
    )
    if not rows:
        return
    content = rows[0].get("content_base64") if isinstance(rows, list) else None
    if not content:
        return
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(base64.b64decode(content))


def backup_file_to_supabase(file_key: str, source_path: Path) -> None:
    if not source_path.exists():
        return
    body = {
        "file_key": file_key,
        "content_base64": base64.b64encode(source_path.read_bytes()).decode("ascii"),
    }
    supabase_request("POST", "/rest/v1/app_files?on_conflict=file_key", body)


def restore_sales_files_from_supabase() -> None:
    restore_file_from_supabase("sales_ledger", SALES_LEDGER_PATH)
    restore_file_from_supabase("monthly_sales", MONTHLY_SALES_PATH)
    restore_file_from_supabase("sales_confirmations", SALES_CONFIRM_PATH)


def backup_sales_files_to_supabase() -> None:
    backup_file_to_supabase("sales_ledger", SALES_LEDGER_PATH)
    backup_file_to_supabase("monthly_sales", MONTHLY_SALES_PATH)
    backup_file_to_supabase("sales_confirmations", SALES_CONFIRM_PATH)


def load_sales_confirmations() -> dict[str, object]:
    restore_file_from_supabase("sales_confirmations", SALES_CONFIRM_PATH)
    if not SALES_CONFIRM_PATH.exists():
        return {"months": {}}
    try:
        data = json.loads(SALES_CONFIRM_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {"months": {}}
    except (OSError, json.JSONDecodeError):
        return {"months": {}}


def save_sales_confirmations(data: dict[str, object]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    SALES_CONFIRM_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    backup_file_to_supabase("sales_confirmations", SALES_CONFIRM_PATH)


def sales_confirmation(month: str) -> dict[str, object]:
    months = load_sales_confirmations().get("months", {})
    value = months.get(month, {}) if isinstance(months, dict) else {}
    return value if isinstance(value, dict) else {}


def append_sales_audit(month: str, event: dict[str, object]) -> None:
    data = load_sales_confirmations()
    months = data.setdefault("months", {})
    record = months.setdefault(month, {})
    history = record.setdefault("history", [])
    history.append(event)
    save_sales_confirmations(data)


def restore_master_file_from_supabase() -> None:
    if get_local_master_path() is not None:
        return
    restore_file_from_supabase("master_file", MASTER_DIR / "기초자료.xlsx")


def backup_master_file_to_supabase(source_path: Path) -> None:
    backup_file_to_supabase("master_file", source_path)

sys.path.insert(0, str(SCRIPTS_DIR))
from po_automation import create_processed_po, read_master, read_po_lines, write_summary_workbook  # noqa: E402


HTML_PAGE = """<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>보니애가구 쿠팡 PO 변환</title>
  <style>
    :root {
      --bg: #f3f6fb;
      --panel: #ffffff;
      --ink: #172033;
      --muted: #667085;
      --line: #d9e1ee;
      --brand: #1f4e79;
      --brand2: #2b7a78;
      --danger: #b42318;
      --ok: #027a48;
      --shadow: 0 18px 45px rgba(16, 24, 40, .10);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Malgun Gothic", "맑은 고딕", Arial, sans-serif;
      color: var(--ink);
      background: var(--bg);
    }
    .app {
      min-height: 100vh;
      display: grid;
      grid-template-columns: 250px 1fr;
    }
    .side {
      background: #102a43;
      color: white;
      padding: 24px 18px;
    }
    .logo {
      display: flex;
      align-items: center;
      gap: 12px;
      font-weight: 800;
      font-size: 20px;
      margin-bottom: 28px;
    }
    .logo-mark {
      width: 34px;
      height: 34px;
      border-radius: 8px;
      background: #2b7a78;
      display: grid;
      place-items: center;
      font-weight: 900;
    }
    .nav-item {
      padding: 12px 13px;
      border-radius: 8px;
      color: #d9e7f3;
      margin-bottom: 6px;
      font-size: 14px;
      cursor: pointer;
      user-select: none;
    }
    .nav-item.active {
      background: rgba(255,255,255,.13);
      color: #fff;
      font-weight: 700;
    }
    .main {
      padding: 28px;
    }
    .topbar {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 22px;
    }
    h1 {
      font-size: 24px;
      margin: 0;
      letter-spacing: 0;
    }
    .sub {
      margin-top: 7px;
      color: var(--muted);
      font-size: 14px;
    }
    .grid {
      display: grid;
      grid-template-columns: minmax(420px, 1.1fr) minmax(320px, .9fr);
      gap: 18px;
      align-items: start;
    }
    .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }
    .panel-head {
      padding: 18px 20px;
      border-bottom: 1px solid var(--line);
      font-weight: 800;
    }
    .panel-body { padding: 20px; }
    label {
      display: block;
      font-weight: 700;
      margin-bottom: 8px;
      font-size: 14px;
    }
    input[type=file] {
      width: 100%;
      border: 1px dashed #9fb2c8;
      background: #f8fbff;
      border-radius: 8px;
      padding: 14px;
      margin-bottom: 18px;
    }
    .btn {
      border: 0;
      border-radius: 8px;
      padding: 13px 18px;
      background: var(--brand);
      color: white;
      font-weight: 800;
      cursor: pointer;
      width: 100%;
      font-size: 15px;
    }
    .btn:hover { background: #173c5e; }
    .note {
      background: #f7f9fc;
      border: 1px solid var(--line);
      padding: 12px 14px;
      border-radius: 8px;
      color: #344054;
      font-size: 13px;
      line-height: 1.55;
      margin-top: 12px;
    }
    .status {
      padding: 14px 16px;
      border-radius: 8px;
      font-size: 14px;
      line-height: 1.6;
      margin-bottom: 14px;
      border: 1px solid var(--line);
      background: #fff;
    }
    .status.ok { border-color: #abefc6; background: #ecfdf3; color: var(--ok); }
    .status.err { border-color: #fecdca; background: #fef3f2; color: var(--danger); }
    .download {
      display: inline-block;
      text-decoration: none;
      color: white;
      background: var(--brand2);
      padding: 12px 16px;
      border-radius: 8px;
      font-weight: 800;
      margin-top: 8px;
    }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }
    th, td {
      border-bottom: 1px solid var(--line);
      padding: 9px 8px;
      text-align: left;
    }
    th { color: #344054; background: #f8fafc; }
    @media (max-width: 880px) {
      .app { grid-template-columns: 1fr; }
      .side { display: none; }
      .main { padding: 18px; }
      .grid { grid-template-columns: 1fr; }
    }
  </style>
  <script>
    function showComingSoon(name) {
      alert(name + " 메뉴는 아직 준비 중입니다. 지금은 쿠팡 PO 변환 메뉴만 사용할 수 있습니다.");
    }
  </script>
</head>
<body>
  <div class="app">
    <aside class="side">
      <div class="logo"><div class="logo-mark">B</div><div>보니애가구<br><span style="font-size:13px;font-weight:500;color:#b8c8d9;">업무 시스템</span></div></div>
      <div class="nav-item" onclick="location.href='/master'">쿠팡 기초자료 관리</div>
      <div class="nav-item active" onclick="location.href='/'">PO변환</div>
      <div class="nav-item" onclick="location.href='/sales/folders'">월별납품관리</div>
      <div class="nav-item" onclick="location.href='/check'">수량검수</div>
      <div class="nav-item" onclick="location.href='/sales'">매출확인용</div>
      <div class="nav-item" onclick="location.href='/pallet'">파렛트/쉽먼트</div>
      <div class="nav-item" style="margin-top:28px;" onclick="showComingSoon('라벨 출력')">라벨 출력</div>
      <div class="nav-item" onclick="showComingSoon('배송완료 관리')">배송완료 관리</div>
    </aside>
    <main class="main">
      <div class="topbar">
        <div>
          <h1>쿠팡 PO 변환</h1>
          <div class="sub">기초자료는 저장해두고, 평소에는 PO만 올려 심플웍스 등록용 파일을 자동 생성합니다.</div>
        </div>
      </div>
      <div class="grid">
        <section class="panel {year_section_class}">
          <div class="panel-head">파일 업로드</div>
          <div class="panel-body">
            <form method="post" action="/convert" enctype="multipart/form-data">
              <label>쿠팡 PO 엑셀 파일</label>
              <input type="file" name="po_files" accept=".xlsx" multiple required />
              <button class="btn" type="submit">심플웍스 등록용 만들기</button>
            </form>
            <div class="note">
              {master_status}
              <br>기초자료 등록/교체/삭제는 왼쪽의 쿠팡 기초자료 관리에서 진행합니다.
              <br>PO는 한 개도 되고 여러 개도 됩니다. 결과는 ZIP 파일로 내려받습니다.
              ZIP은 여러 파일을 하나로 묶은 압축 파일입니다.
            </div>
          </div>
        </section>
        <section class="panel {upload_section_class}">
          <div class="panel-head">처리 기준</div>
          <div class="panel-body">
            <table>
              <tr><th>항목</th><th>기준</th></tr>
              <tr><td>스큐 아이디</td><td>PO의 상품코드와 기초자료의 SKU ID 연결</td></tr>
              <tr><td>납품불가</td><td>행 안에 납품불가 표시가 있으면 제외</td></tr>
              <tr><td>요청</td><td>요청 수량을 넘지 않게 PO 수량 조정</td></tr>
              <tr><td>결과</td><td>심플웍스 업로드용 PO 복사본</td></tr>
            </table>
          </div>
        </section>
      </div>
      {message}
    </main>
  </div>
</body>
</html>"""


PALLET_PAGE = """<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>보니애가구 파렛트/쉽먼트</title>
  <style>
    :root {
      --bg: #f3f6fb;
      --panel: #ffffff;
      --ink: #172033;
      --muted: #667085;
      --line: #d9e1ee;
      --brand: #1f4e79;
      --brand2: #2b7a78;
      --danger: #b42318;
      --ok: #027a48;
      --shadow: 0 18px 45px rgba(16, 24, 40, .10);
    }
    * { box-sizing: border-box; }
    body { margin: 0; font-family: "Malgun Gothic", "맑은 고딕", Arial, sans-serif; color: var(--ink); background: var(--bg); }
    .app { min-height: 100vh; display: grid; grid-template-columns: 250px minmax(0, 1fr); }
    .side { background: #102a43; color: white; padding: 24px 18px; }
    .logo { display: flex; align-items: center; gap: 12px; font-weight: 800; font-size: 20px; margin-bottom: 28px; }
    .logo-mark { width: 34px; height: 34px; border-radius: 8px; background: #2b7a78; display: grid; place-items: center; font-weight: 900; }
    .nav-item { padding: 12px 13px; border-radius: 8px; color: #d9e7f3; margin-bottom: 6px; font-size: 14px; cursor: pointer; user-select: none; }
    .nav-item.active { background: rgba(255,255,255,.13); color: #fff; font-weight: 700; }
    .main { padding: 28px; min-width: 0; }
    h1 { font-size: 24px; margin: 0; letter-spacing: 0; }
    .sub { margin-top: 7px; color: var(--muted); font-size: 14px; }
    .grid { display: grid; grid-template-columns: minmax(420px, 1fr) minmax(300px, .7fr); gap: 18px; align-items: start; margin-top: 22px; }
    .panel { background: var(--panel); border: 1px solid var(--line); border-radius: 8px; box-shadow: var(--shadow); overflow: hidden; }
    .panel-head { padding: 18px 20px; border-bottom: 1px solid var(--line); font-weight: 800; }
    .panel-body { padding: 20px; }
    label { display: block; font-weight: 700; margin-bottom: 8px; font-size: 14px; }
    input[type=file] { width: 100%; border: 1px dashed #9fb2c8; background: #f8fbff; border-radius: 8px; padding: 14px; margin-bottom: 18px; }
    .btn { border: 0; border-radius: 8px; padding: 13px 18px; background: var(--brand); color: white; font-weight: 800; cursor: pointer; width: 100%; font-size: 15px; }
    .btn:hover { background: #173c5e; }
    .note { background: #f7f9fc; border: 1px solid var(--line); padding: 12px 14px; border-radius: 8px; color: #344054; font-size: 13px; line-height: 1.55; margin-top: 12px; }
    .status { padding: 14px 16px; border-radius: 8px; font-size: 14px; line-height: 1.6; margin: 18px 0 0; border: 1px solid var(--line); background: #fff; }
    .status.ok { border-color: #abefc6; background: #ecfdf3; color: var(--ok); }
    .status.err { border-color: #fecdca; background: #fef3f2; color: var(--danger); }
    .download { display: inline-block; text-decoration: none; color: white; background: var(--brand2); padding: 12px 16px; border-radius: 8px; font-weight: 800; margin-top: 8px; }
    table { width: 100%; border-collapse: collapse; font-size: 13px; }
    th, td { border-bottom: 1px solid var(--line); padding: 9px 8px; text-align: left; vertical-align: top; }
    th { color: #344054; background: #f8fafc; width: 120px; }
    @media (max-width: 880px) {
      .app { grid-template-columns: 1fr; }
      .side { display: none; }
      .main { padding: 18px; }
      .grid { grid-template-columns: 1fr; }
    }
  </style>
  <script>
    function showComingSoon(name) {
      alert(name + " 메뉴는 아직 준비 중입니다.");
    }
  </script>
</head>
<body>
  <div class="app">
    <aside class="side">
      <div class="logo"><div class="logo-mark">B</div><div>보니애가구<br><span style="font-size:13px;font-weight:500;color:#b8c8d9;">업무 시스템</span></div></div>
      <div class="nav-item" onclick="location.href='/master'">쿠팡 기초자료 관리</div>
      <div class="nav-item" onclick="location.href='/'">PO변환</div>
      <div class="nav-item" onclick="location.href='/sales/folders'">월별납품관리</div>
      <div class="nav-item" onclick="location.href='/check'">수량검수</div>
      <div class="nav-item" onclick="location.href='/sales'">매출확인용</div>
      <div class="nav-item active" onclick="location.href='/pallet'">파렛트/쉽먼트</div>
      <div class="nav-item" style="margin-top:28px;" onclick="showComingSoon('라벨 출력')">라벨 출력</div>
      <div class="nav-item" onclick="showComingSoon('배송완료 관리')">배송완료 관리</div>
    </aside>
    <main class="main">
      <h1>파렛트/쉽먼트</h1>
      <div class="sub">PO 여러 개를 올리면 상품 크기와 파렛트 기준으로 납품 묶음 초안을 만듭니다.</div>
      {message}
      <div class="grid">
        <section class="panel">
          <div class="panel-head">PO 업로드</div>
          <div class="panel-body">
            <form method="post" action="/pallet/create" enctype="multipart/form-data">
              <label>파렛트/쉽먼트 초안을 만들 PO 엑셀</label>
              <input type="file" name="po_files" multiple accept=".xlsx" />
              <button class="btn" type="submit">파렛트/쉽먼트 초안 만들기</button>
            </form>
            <div class="note">{master_status}</div>
          </div>
        </section>
        <section class="panel">
          <div class="panel-head">결과 엑셀에 들어가는 내용</div>
          <div class="panel-body">
            <table>
              <tr><th>스큐 합산</th><td>같은 스큐 아이디는 수량을 합쳐 봅니다.</td></tr>
              <tr><th>PO별 상세</th><td>어떤 PO에서 온 상품인지 확인합니다.</td></tr>
              <tr><th>쉽먼트 초안</th><td>상품 크기, 박스 수량, 파렛트 수량으로 예상 파렛트 수를 봅니다.</td></tr>
            </table>
            <div class="note">
              `파렛트`는 상품을 쌓아 운반하는 받침 단위이고, `쉽먼트`는 쿠팡에 보낼 납품 묶음입니다.
              초안은 자동 계산 자료라서 최종 확정 전 사람이 한 번 확인하는 용도로 쓰면 됩니다.
            </div>
          </div>
        </section>
      </div>
    </main>
  </div>
</body>
</html>"""


CHECK_PAGE = """<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>보니애가구 수량검수</title>
  <style>
    :root { --bg:#f3f6fb; --panel:#fff; --ink:#172033; --muted:#667085; --line:#d9e1ee; --brand:#1f4e79; --danger:#b42318; --ok:#027a48; --warn:#b54708; --shadow:0 18px 45px rgba(16,24,40,.10); }
    * { box-sizing:border-box; }
    body { margin:0; font-family:"Malgun Gothic","맑은 고딕",Arial,sans-serif; color:var(--ink); background:var(--bg); }
    .app { min-height:100vh; display:grid; grid-template-columns:250px minmax(0,1fr); }
    .side { background:#102a43; color:white; padding:24px 18px; }
    .logo { display:flex; align-items:center; gap:12px; font-weight:800; font-size:20px; margin-bottom:28px; }
    .logo-mark { width:34px; height:34px; border-radius:8px; background:#2b7a78; display:grid; place-items:center; font-weight:900; }
    .nav-item { padding:12px 13px; border-radius:8px; color:#d9e7f3; margin-bottom:6px; font-size:14px; cursor:pointer; user-select:none; }
    .nav-item.active { background:rgba(255,255,255,.13); color:#fff; font-weight:700; }
    .main { padding:28px; min-width:0; }
    h1 { font-size:24px; margin:0; letter-spacing:0; }
    .sub { margin-top:7px; color:var(--muted); font-size:14px; }
    .cards { display:grid; grid-template-columns:minmax(0,1fr); gap:18px; margin-top:22px; }
    .panel { background:var(--panel); border:1px solid var(--line); border-radius:8px; box-shadow:var(--shadow); overflow:hidden; }
    .panel-head { padding:16px 18px; border-bottom:1px solid var(--line); font-weight:800; }
    .panel-body { padding:18px; }
    .form-grid { display:grid; grid-template-columns:180px 180px minmax(260px,1fr) minmax(260px,1fr); gap:12px; align-items:end; }
    label { display:flex; flex-direction:column; gap:6px; font-size:13px; color:#344054; font-weight:800; }
    input[type=date], input[type=file] { width:100%; border:1px solid #b9c6d8; border-radius:8px; padding:10px 11px; font:inherit; background:white; }
    .btn { border:0; border-radius:8px; padding:12px 16px; background:var(--brand); color:white; font-weight:800; cursor:pointer; font-size:14px; }
    .status { padding:14px 16px; border-radius:8px; font-size:14px; line-height:1.6; border:1px solid var(--line); background:#fff; }
    .status.ok { border-color:#abefc6; background:#ecfdf3; color:var(--ok); }
    .status.err { border-color:#fecdca; background:#fef3f2; color:var(--danger); }
    .note { margin-top:12px; color:#667085; font-size:13px; line-height:1.6; }
    .summary { display:grid; grid-template-columns:repeat(4,minmax(120px,1fr)); gap:10px; }
    .summary div { border:1px solid var(--line); border-radius:8px; padding:12px; background:#fbfcfe; }
    .summary b { display:block; font-size:20px; margin-top:4px; }
    .scroll { overflow:auto; max-height:620px; }
    table { width:100%; min-width:980px; border-collapse:collapse; table-layout:fixed; font-size:13px; }
    th, td { border-right:1px solid var(--line); border-bottom:1px solid var(--line); padding:9px 10px; text-align:left; vertical-align:middle; overflow-wrap:anywhere; word-break:keep-all; }
    th:last-child, td:last-child { border-right:0; }
    th { background:#f8fafc; color:#344054; position:sticky; top:0; z-index:1; }
    .ok-text { color:var(--ok); font-weight:800; }
    .bad-text { color:var(--danger); font-weight:800; }
    .warn-text { color:var(--warn); font-weight:800; }
    @media (max-width:880px) { .app{grid-template-columns:1fr;} .side{display:none;} .main{padding:18px;} .form-grid,.summary{grid-template-columns:1fr;} }
  </style>
  <script>
    function showComingSoon(name) { alert(name + " 메뉴는 아직 준비 중입니다."); }
  </script>
</head>
<body>
  <div class="app">
    <aside class="side">
      <div class="logo"><div class="logo-mark">B</div><div>보니애가구<br><span style="font-size:13px;font-weight:500;color:#b8c8d9;">업무 시스템</span></div></div>
      <div class="nav-item" onclick="location.href='/master'">쿠팡 기초자료 관리</div>
      <div class="nav-item" onclick="location.href='/'">PO변환</div>
      <div class="nav-item" onclick="location.href='/sales/folders'">월별납품관리</div>
      <div class="nav-item active" onclick="location.href='/check'">수량검수</div>
      <div class="nav-item" onclick="location.href='/sales'">매출확인용</div>
      <div class="nav-item" onclick="location.href='/pallet'">파렛트/쉽먼트</div>
      <div class="nav-item" style="margin-top:28px;" onclick="showComingSoon('라벨 출력')">라벨 출력</div>
      <div class="nav-item" onclick="showComingSoon('배송완료 관리')">배송완료 관리</div>
    </aside>
    <main class="main">
      <h1>수량검수</h1>
      <div class="sub">쿠팡에 등록된 스큐 납품수량과 심플웍스 No 기준 수량이 같은지 확인합니다.</div>
      {message}
      <div class="cards">
        <section class="panel">
          <div class="panel-head">심플웍스 엑셀 업로드</div>
          <div class="panel-body">
            <form method="post" action="/check/run" enctype="multipart/form-data">
              <div class="form-grid">
                <label>시작일
                  <input type="date" name="date_from" value="{date_from}">
                </label>
                <label>종료일
                  <input type="date" name="date_to" value="{date_to}">
                </label>
                <label>심플웍스 엑셀
                  <input type="file" name="simpleworks_file" accept=".xlsx" multiple>
                </label>
                <label>심플웍스 캡쳐 이미지
                  <input type="file" name="simpleworks_image" accept=".png,.jpg,.jpeg,.bmp,.webp" multiple>
                </label>
              </div>
              <div style="margin-top:14px;"><button class="btn" type="submit">수량 검수하기</button></div>
            </form>
            <div class="note">심플웍스 엑셀 또는 캡쳐 이미지를 여러 개 올릴 수 있습니다. 각 파일의 상품명 앞 노란 숫자를 `심플웍스 No`로 읽고, 같은 번호는 수량을 합산합니다. 캡쳐 이미지는 OCR로 읽기 때문에 엑셀보다 정확도가 낮을 수 있습니다.</div>
          </div>
        </section>
        {result}
      </div>
    </main>
  </div>
</body>
</html>"""


SALES_PAGE = """<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>보니애가구 매출확인용</title>
  <style>
    :root { --bg:#f3f6fb; --panel:#fff; --ink:#172033; --muted:#667085; --line:#d9e1ee; --brand:#1f4e79; --brand2:#2b7a78; --shadow:0 18px 45px rgba(16,24,40,.10); }
    * { box-sizing: border-box; }
    body { margin:0; font-family:"Malgun Gothic","맑은 고딕",Arial,sans-serif; color:var(--ink); background:var(--bg); }
    .app { min-height:100vh; display:grid; grid-template-columns:250px minmax(0, 1fr); }
    .side { background:#102a43; color:white; padding:24px 18px; }
    .logo { display:flex; align-items:center; gap:12px; font-weight:800; font-size:20px; margin-bottom:28px; }
    .logo-mark { width:34px; height:34px; border-radius:8px; background:#2b7a78; display:grid; place-items:center; font-weight:900; }
    .nav-item { padding:12px 13px; border-radius:8px; color:#d9e7f3; margin-bottom:6px; font-size:14px; cursor:pointer; user-select:none; }
    .nav-item.active { background:rgba(255,255,255,.13); color:#fff; font-weight:700; }
    .main { padding:28px; min-width:0; }
    h1 { font-size:24px; margin:0; letter-spacing:0; }
    .sub { margin-top:7px; color:var(--muted); font-size:14px; }
    .cards { display:grid; grid-template-columns:minmax(0, 1fr); gap:18px; margin-top:22px; min-width:0; }
    .upload-panel { order:1; }
    .summary-panel { order:2; }
    .detail-panel { order:3; }
    .year-panel { order:4; }
    .tester-panel { order:5; }
    .panel { background:var(--panel); border:1px solid var(--line); border-radius:8px; box-shadow:var(--shadow); overflow:hidden; min-width:0; }
    .panel-head { padding:16px 18px; border-bottom:1px solid var(--line); font-weight:800; display:flex; justify-content:space-between; align-items:center; gap:12px; flex-wrap:wrap; }
    .panel-body { padding:18px; }
    input[type=file] { width:100%; border:1px dashed #9fb2c8; background:#f8fbff; border-radius:8px; padding:14px; margin-bottom:12px; }
    .btn { text-decoration:none; border:0; border-radius:8px; padding:10px 14px; background:var(--brand); color:white; font-weight:800; cursor:pointer; font-size:14px; }
    table { width:100%; border-collapse:collapse; table-layout:fixed; font-size:13px; }
    th, td { border-right:1px solid var(--line); border-bottom:1px solid var(--line); padding:9px 10px; text-align:left; vertical-align:middle; overflow-wrap:anywhere; word-break:keep-all; }
    th:last-child, td:last-child { border-right:0; }
    th { background:#f8fafc; color:#344054; }
    .summary-table th, .summary-table td { text-align:center; white-space:normal; }
    .summary-table th:first-child, .summary-table td:first-child { text-align:left; }
    .summary-lookup { grid-template-columns:150px 150px 150px 130px; align-items:end; padding:10px 14px; gap:8px; background:linear-gradient(180deg,#fbfdff,#f6f9fc); }
    .summary-lookup label { gap:4px; font-size:11px; color:#344054; }
    .summary-lookup input { height:34px; padding:6px 9px; border-radius:7px; }
    .summary-lookup .lookup-reset { height:34px; padding:6px 12px; border-radius:7px; }
    .summary-months { display:grid; gap:12px; padding:14px; background:#f7fafc; }
    .summary-month-section { border:1px solid #d8e4ef; border-radius:8px; background:#fff; overflow:hidden; box-shadow:0 6px 16px rgba(16,24,40,.05); }
    .summary-month-head { display:flex; align-items:center; justify-content:space-between; gap:12px; padding:10px 12px; background:#eef6f7; color:#173f68; border-bottom:1px solid #d8e4ef; }
    .summary-month-title { font-size:14px; font-weight:900; letter-spacing:0; }
    .summary-month-title small, .summary-month-metrics { display:none; }
    .summary-month-toggle { border:1px solid #b9c6d8; background:#fff; color:#1f4e79; border-radius:6px; padding:5px 9px; font-size:12px; font-weight:900; cursor:pointer; }
    .summary-month-body { display:block; }
    .summary-month-section.is-collapsed .summary-month-body { display:none; }
    .summary-month-total { background:#fff2cc; font-weight:900; }
    .invoice-check { padding:12px; border-top:1px solid #d8e4ef; background:#fbfdff; }
    .confirm-form { display:flex; align-items:end; gap:10px; flex-wrap:wrap; }
    .confirm-form label { display:flex; flex-direction:column; gap:4px; color:#475467; font-size:12px; font-weight:700; }
    .confirm-form .confirm-check { flex-direction:row; align-items:center; padding-bottom:7px; }
    .confirm-money { width:170px; border:1px solid #b9c6d8; border-radius:6px; padding:7px 8px; text-align:right; }
    .confirm-action { border:0; border-radius:6px; padding:8px 11px; background:#1f4e79; color:#fff; font-weight:800; cursor:pointer; }
    .status-badge { display:inline-flex; align-items:center; border-radius:999px; padding:5px 9px; font-size:12px; font-weight:900; }
    .status-ok { background:#dcfae6; color:#067647; }
    .status-warn { background:#fff2cc; color:#854a0e; }
    .status-bad { background:#fee4e2; color:#b42318; }
    .confirm-states { display:flex; gap:6px; flex-wrap:wrap; margin-top:9px; }
    .confirm-meta { margin-top:7px; color:#667085; font-size:12px; }
    .history-details { margin-top:10px; font-size:12px; }
    .history-details summary { cursor:pointer; color:#1f4e79; font-weight:900; }
    .history-scroll { margin-top:8px; overflow:auto; }
    .history-table { min-width:720px; }
    .history-table th, .history-table td { font-size:12px; text-align:left; }
    .detail-table th:nth-child(1), .detail-table td:nth-child(1),
    .detail-table th:nth-child(3), .detail-table td:nth-child(3),
    .detail-table th:nth-child(4), .detail-table td:nth-child(4),
    .detail-table th:nth-child(5), .detail-table td:nth-child(5),
    .detail-table th:nth-child(6), .detail-table td:nth-child(6) { white-space:nowrap; }
    .detail-table td:nth-child(3), .detail-table td:nth-child(4), .detail-table td:nth-child(5) { text-align:right; }
    .qty-input, .memo-input { width:100%; border:1px solid #b9c6d8; border-radius:6px; padding:7px 8px; font:inherit; background:#fff; }
    .money-input { width:100%; border:1px solid #b9c6d8; border-radius:6px; padding:7px 8px; font:inherit; background:#fff; text-align:right; }
    .qty-input { text-align:right; min-width:64px; }
    .memo-input { min-width:120px; }
    .changed, .changed input { color:#c1121f; font-weight:800; }
    .save-row { display:flex; justify-content:flex-end; padding:12px 16px; border-bottom:1px solid var(--line); background:#fbfcfe; }
    .year-table th { text-align:center; background:#f8fafc; color:#344054; font-weight:800; }
    .year-table td { text-align:right; }
    .year-table td:first-child { text-align:left; font-weight:700; }
    .year-panel .lookup-row { grid-template-columns: minmax(160px, 220px) 160px; padding:10px 16px; }
    .year-panel .scroll { border-top:1px solid var(--line); }
    .auto-cell { background:#f7fbff; font-weight:700; }
    .over-cell.negative { color:#c1121f; font-weight:800; }
    .lookup-row { display:grid; grid-template-columns:repeat(5, minmax(120px, 1fr)); gap:10px; padding:12px 16px; border-bottom:1px solid var(--line); background:#fbfcfe; }
    .lookup-row label { display:flex; flex-direction:column; gap:5px; color:#475467; font-size:12px; font-weight:700; }
    .lookup-row input, .lookup-row select { width:100%; border:1px solid #b9c6d8; border-radius:6px; padding:8px 9px; font:inherit; background:#fff; }
    input[type="date"].date-empty::-webkit-datetime-edit { color:transparent; }
    .folder-year-filter { margin-top:18px; padding:12px 14px; border:1px solid #cbd9e8; border-radius:8px; background:#eef5fb; display:flex; align-items:center; gap:10px; flex-wrap:wrap; }
    .folder-year-filter label { color:#173f68; font-size:13px; font-weight:900; }
    .folder-year-filter select { min-width:120px; border:1px solid #9fb2c8; border-radius:7px; padding:8px 10px; background:#fff; color:#173f68; font:inherit; font-weight:800; }
    .folder-month-tabs { flex:1 1 620px; display:grid; grid-template-columns:repeat(12,minmax(48px,1fr)); gap:5px; }
    .folder-month-tab { border:1px solid #b9c9da; border-radius:6px; padding:7px 4px; background:#fff; color:#667085; font-size:12px; font-weight:800; cursor:pointer; }
    .folder-month-tab.has-data { color:#173f68; background:#f8fbff; }
    .folder-month-tab.is-active { border-color:#1f5d8f; background:#1f5d8f; color:#fff; box-shadow:0 3px 8px rgba(31,93,143,.2); }
    .folder-month-tab:not(.has-data) { opacity:.45; cursor:default; }
    .lookup-row .lookup-reset { align-self:end; border:1px solid #b9c6d8; background:#fff; color:#1f4e79; border-radius:6px; padding:8px 10px; font-weight:800; cursor:pointer; }
    .detail-result-bar { display:flex; align-items:center; justify-content:space-between; gap:12px; flex-wrap:wrap; padding:10px 16px; border-bottom:1px solid var(--line); background:#eef6ff; }
    .detail-result-summary { display:flex; align-items:center; gap:18px; flex-wrap:wrap; color:#1f4e79; font-size:13px; font-weight:800; }
    .detail-result-summary strong { color:#0b3155; font-size:15px; }
    .detail-download { border:0; border-radius:7px; padding:9px 14px; background:#1f5d8f; color:#fff; font-weight:900; cursor:pointer; }
    .inline-delete-form { align-self:end; margin:0; }
    .inline-delete-form .delete-btn { width:100%; }
    .panel-actions { display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
    .toggle-btn { border:1px solid #b9c6d8; background:#fff; color:#1f4e79; border-radius:6px; padding:8px 10px; font-weight:800; cursor:pointer; font-size:13px; }
    .delete-btn { border:1px solid #fecdca; background:#fff5f5; color:#b42318; border-radius:6px; padding:7px 9px; font-weight:800; cursor:pointer; font-size:12px; }
    .mode-note { display:none; position:fixed; left:0; top:0; z-index:40; width:280px; max-width:calc(100vw - 24px); padding:10px 12px; border:1px solid #c7d7e8; border-radius:8px; background:#fff; color:#26384d; box-shadow:0 10px 28px rgba(16,24,40,.14); font-size:12px; line-height:1.5; }
    .mode-note.is-visible { display:block; }
    .mode-note strong { display:block; margin-bottom:3px; color:#1f4e79; font-size:13px; }
    .collapsible-content.is-hidden { display:none; }
    .not-shown { display:none; }
    .month-folder th { background:#d7e8f7; color:#12385a; cursor:pointer; font-weight:900; }
    .month-hidden, .view-hidden { display:none !important; }
    .resizable-table th { position:relative; }
    .col-resizer { position:absolute; top:0; right:-4px; width:8px; height:100%; cursor:col-resize; user-select:none; touch-action:none; z-index:2; }
    .col-resizer:hover, .col-resizer.active { background:rgba(31,78,121,.22); }
    .scroll { max-height:520px; overflow:auto; }
    @media (max-width:880px) { .app{grid-template-columns:1fr;} .side{display:none;} .main{padding:18px;} table{font-size:12px;} th,td{padding:8px 7px;} .lookup-row{grid-template-columns:1fr 1fr;} .folder-month-tabs{grid-template-columns:repeat(6,minmax(46px,1fr));} }
  </style>
  <script>
    function showComingSoon(name) { alert(name + " 메뉴는 아직 준비 중입니다."); }
    function initResizableTables() {
      document.querySelectorAll(".resizable-table").forEach(function(table) {
        var cols = table.querySelectorAll("col");
        var headers = table.querySelectorAll("thead th");
        headers.forEach(function(header, index) {
          if (header.querySelector(".col-resizer")) return;
          var handle = document.createElement("span");
          handle.className = "col-resizer";
          header.appendChild(handle);
          handle.addEventListener("mousedown", function(event) {
            event.preventDefault();
            handle.classList.add("active");
            var startX = event.clientX;
            var tableWidth = table.getBoundingClientRect().width;
            var targetIndex = index;
            var partnerIndex = index + 1;
            var direction = 1;
            if (index === headers.length - 1) {
              targetIndex = index;
              partnerIndex = index - 1;
              direction = -1;
            }
            var targetWidth = headers[targetIndex].getBoundingClientRect().width;
            var partnerWidth = headers[partnerIndex].getBoundingClientRect().width;
            function move(moveEvent) {
              var delta = moveEvent.clientX - startX;
              var adjustedDelta = delta * direction;
              var newTarget = Math.max(58, targetWidth + adjustedDelta);
              var newPartner = Math.max(58, partnerWidth - adjustedDelta);
              var changedTotal = newTarget + newPartner;
              var originalTotal = targetWidth + partnerWidth;
              if (changedTotal !== originalTotal) {
                if (newTarget === 58) newPartner = originalTotal - 58;
                if (newPartner === 58) newTarget = originalTotal - 58;
              }
              cols[targetIndex].style.width = (newTarget / tableWidth * 100) + "%";
              cols[partnerIndex].style.width = (newPartner / tableWidth * 100) + "%";
            }
            function stop() {
              handle.classList.remove("active");
              document.removeEventListener("mousemove", move);
              document.removeEventListener("mouseup", stop);
            }
            document.addEventListener("mousemove", move);
            document.addEventListener("mouseup", stop);
          });
        });
      });
    }
    function parseMoney(text) {
      var cleaned = String(text || "").replaceAll(",", "").replaceAll("원", "").trim();
      var value = Number(cleaned);
      return Number.isFinite(value) ? value : 0;
    }
    function money(value) {
      return Math.round(value).toLocaleString("ko-KR") + "원";
    }
    function visibleDetailRows() {
      return Array.from(document.querySelectorAll(".detail-row")).filter(function(row) {
        return row.style.display !== "none" && !row.classList.contains("view-hidden") && !row.classList.contains("month-hidden");
      });
    }
    function detailRowValues(row) {
      var cells = row.children;
      var qtyInput = row.querySelector(".qty-input");
      var memoInput = row.querySelector(".memo-input");
      return {
        day: row.dataset.day || "",
        sku: (cells[0]?.textContent || "").trim(),
        name: (cells[1]?.textContent || "").trim(),
        qty: qtyInput ? Number(qtyInput.value || "0") : parseMoney(cells[2]?.textContent || "0"),
        unitPrice: parseMoney(cells[3]?.textContent || "0"),
        amount: parseMoney(row.querySelector(".row-amount")?.textContent || "0"),
        po: (row.querySelector(".po-cell")?.textContent || "").trim(),
        memo: memoInput ? memoInput.value.trim() : (cells[6]?.textContent || "").trim()
      };
    }
    function updateDetailResultSummary() {
      var rows = visibleDetailRows();
      var totals = rows.reduce(function(result, row) {
        var values = detailRowValues(row);
        result.qty += values.qty;
        result.amount += values.amount;
        return result;
      }, { qty: 0, amount: 0 });
      var countNode = document.getElementById("detail-result-count");
      var qtyNode = document.getElementById("detail-result-qty");
      var amountNode = document.getElementById("detail-result-amount");
      if (countNode) countNode.textContent = rows.length.toLocaleString("ko-KR") + "건";
      if (qtyNode) qtyNode.textContent = totals.qty.toLocaleString("ko-KR") + "개";
      if (amountNode) amountNode.textContent = money(totals.amount);
    }
    function excelCell(value) {
      return String(value ?? "").replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
    }
    function downloadDetailResults() {
      var rows = visibleDetailRows().map(detailRowValues);
      if (!rows.length) {
        alert("다운로드할 검색 결과가 없습니다.");
        return;
      }
      var totalQty = rows.reduce(function(sum, row) { return sum + row.qty; }, 0);
      var totalAmount = rows.reduce(function(sum, row) { return sum + row.amount; }, 0);
      var keyword = (document.getElementById("detail-keyword")?.value || "전체").trim() || "전체";
      var tableRows = rows.map(function(row) {
        return "<tr><td>" + [row.day, row.sku, row.name, row.qty, row.unitPrice, row.amount, row.po, row.memo].map(excelCell).join("</td><td>") + "</td></tr>";
      }).join("");
      var workbook = "<html xmlns:o='urn:schemas-microsoft-com:office:office' xmlns:x='urn:schemas-microsoft-com:office:excel'><head><meta charset='UTF-8'></head><body>" +
        "<table><tr><th>검색어</th><td>" + excelCell(keyword) + "</td></tr><tr><th>검색 결과</th><td>" + rows.length + "건</td><th>합계 수량</th><td>" + totalQty + "개</td><th>합계 금액</th><td>" + totalAmount + "원</td></tr></table><br>" +
        "<table border='1'><thead><tr><th>납품일</th><th>SKU ID</th><th>상품명</th><th>납품수량</th><th>단가</th><th>금액</th><th>PO</th><th>메모</th></tr></thead><tbody>" + tableRows + "</tbody></table></body></html>";
      var blob = new Blob(["\ufeff", workbook], { type: "application/vnd.ms-excel;charset=utf-8" });
      var link = document.createElement("a");
      link.href = URL.createObjectURL(blob);
      link.download = "쿠팡PO_검색결과_" + new Date().toISOString().slice(0, 10) + ".xls";
      document.body.appendChild(link);
      link.click();
      link.remove();
      URL.revokeObjectURL(link.href);
    }
    function recalcSalesScreen() {
      var dayTotals = {};
      var monthTotals = {};
      document.querySelectorAll(".detail-row").forEach(function(row) {
        var day = row.dataset.day;
        var month = row.dataset.month;
        var viewMode = row.dataset.viewMode || "sku";
        var key = day + "::" + viewMode;
        var originalQty = Number(row.dataset.originalQty || "0");
        var qtyInput = row.querySelector(".qty-input");
        var qty = qtyInput ? Number(qtyInput.value || "0") : Number((row.children[2]?.textContent || "0").replaceAll(",", ""));
        var unitPrice = Number(row.dataset.unitPrice || "0");
        var amountCell = row.querySelector(".row-amount");
        var amount = qtyInput ? qty * unitPrice : parseMoney(amountCell ? amountCell.textContent : "0");
        if (qtyInput && amountCell) amountCell.textContent = money(amount);
        var memoInput = row.querySelector(".memo-input");
        row.classList.toggle("changed", qty !== originalQty || !!(memoInput && memoInput.value.trim()));
        if (qtyInput) qtyInput.classList.toggle("changed", qty !== originalQty);
        if (!dayTotals[key]) dayTotals[key] = { qty: 0, amount: 0, day: day, viewMode: viewMode };
        dayTotals[key].qty += qty;
        dayTotals[key].amount += amount;
        if (viewMode === "sku") {
          if (!monthTotals[month]) monthTotals[month] = 0;
          monthTotals[month] += amount;
        }
      });
      Object.keys(dayTotals).forEach(function(key) {
        var total = dayTotals[key];
        document.querySelectorAll('[data-day-total="' + total.day + '"][data-view-mode="' + total.viewMode + '"]').forEach(function(cell) {
          cell.textContent = total.day + " 합계: 수량 " + total.qty.toLocaleString("ko-KR") + "개 / 금액 " + money(total.amount);
        });
        if (total.viewMode !== "sku") return;
        var summaryRow = document.querySelector('.summary-row[data-day="' + total.day + '"]');
        if (summaryRow) {
          var vat = Math.round(total.amount / 1.1);
          var budget = Math.round(vat * 0.035);
          summaryRow.querySelector(".summary-qty").textContent = total.qty.toLocaleString("ko-KR");
          summaryRow.querySelector(".summary-amount").textContent = money(total.amount);
          summaryRow.querySelector(".summary-vat").textContent = money(vat);
          summaryRow.querySelector(".summary-budget").textContent = money(budget);
        }
      });
      var year = new Date().getFullYear();
      document.querySelectorAll(".summary-month-section").forEach(function(section) {
        var month = section.dataset.month || "";
        if (!month.startsWith(String(year) + "-")) return;
        var monthAmount = 0;
        section.querySelectorAll(".summary-row").forEach(function(row) {
          monthAmount += parseMoney(row.querySelector(".summary-amount").textContent);
        });
        monthTotals[month] = monthAmount;
      });
      document.querySelectorAll(".year-row").forEach(function(row) {
        var monthText = row.querySelector("td:first-child").textContent.replace("월", "").padStart(2, "0");
        var key = year + "-" + monthText;
        var monthAmount = monthTotals[key] || 0;
        var vat = Math.round(monthAmount / 1.1);
        var budget = Math.round(vat * 0.035);
        row.dataset.sales = String(vat);
        row.dataset.budget = String(budget);
        row.querySelector(".year-sales").textContent = money(vat);
        row.querySelector(".year-budget").textContent = money(budget);
      });
      recalcYearScreen();
      applyLookups();
    }
    function recalcYearScreen() {
      var totalSales = 0;
      var totalBudget = 0;
      var totalTester = 0;
      var totalAd = 0;
      var totalOver = 0;
      var totalSupport = 0;
      var totalDiscount = 0;
      var totalPartnerDiscount = 0;
      var totalExtraAd = 0;
      var totalNet = 0;
      document.querySelectorAll(".year-row").forEach(function(row) {
        if (row.style.display === "none") return;
        var sales = Number(row.dataset.sales || "0");
        var budget = Number(row.dataset.budget || "0");
        var tester = parseMoney(row.querySelector(".tester-input").value);
        var ad = parseMoney(row.querySelector(".ad-input").value);
        var support = Number(row.dataset.support || "0");
        var discount = parseMoney(row.querySelector(".discount-input").value);
        var partnerDiscount = parseMoney(row.querySelector(".partner-discount-input").value);
        var extraAd = parseMoney(row.querySelector(".extra-ad-input").value);
        var over = budget - tester - ad;
        var net = sales - tester - ad - support - discount - partnerDiscount - extraAd;
        row.querySelector(".over-cell").textContent = money(over);
        row.querySelector(".over-cell").classList.toggle("negative", over < 0);
        row.querySelector(".net-cell").textContent = money(net);
        row.querySelector(".net-cell").classList.toggle("negative", net < 0);
        totalSales += sales;
        totalBudget += budget;
        totalTester += tester;
        totalAd += ad;
        totalOver += over;
        totalSupport += support;
        totalDiscount += discount;
        totalPartnerDiscount += partnerDiscount;
        totalExtraAd += extraAd;
        totalNet += net;
        row.querySelector(".discount-display").textContent = money(discount);
        row.querySelector(".partner-discount-display").textContent = money(partnerDiscount);
        row.querySelector(".extra-ad-display").textContent = money(extraAd);
      });
      var totalRow = document.querySelector(".year-total-row");
      if (totalRow) {
        totalRow.querySelector(".year-total-sales").textContent = money(totalSales);
        totalRow.querySelector(".year-total-budget").textContent = money(totalBudget);
        totalRow.querySelector(".year-total-tester").textContent = money(totalTester);
        totalRow.querySelector(".year-total-ad").textContent = money(totalAd);
        totalRow.querySelector(".year-total-over").textContent = money(totalOver);
        totalRow.querySelector(".year-total-support").textContent = money(totalSupport);
        totalRow.querySelector(".year-total-discount").textContent = money(totalDiscount);
        totalRow.querySelector(".year-total-partner-discount").textContent = money(totalPartnerDiscount);
        totalRow.querySelector(".year-total-extra-ad").textContent = money(totalExtraAd);
        totalRow.querySelector(".year-total-net").textContent = money(totalNet);
      }
    }
    function inDateRange(day, from, to) {
      if (from && day < from) return false;
      if (to && day > to) return false;
      return true;
    }
    function applyLookups() {
      var yearMonth = document.getElementById("year-month-lookup")?.value || "";
      var folderYearSelect = document.getElementById("folder-year-select");
      var selectedFolderYear = folderYearSelect?.value || "";
      var selectedFolderMonth = folderYearSelect?.dataset.selectedMonth || "";
      document.querySelectorAll(".year-row").forEach(function(row) {
        row.style.display = !yearMonth || row.dataset.month === yearMonth ? "" : "none";
      });

      var summaryMonth = document.getElementById("summary-month")?.value || "";
      var summaryFrom = document.getElementById("summary-from")?.value || "";
      var summaryTo = document.getElementById("summary-to")?.value || "";
      document.querySelectorAll(".summary-month-section").forEach(function(section) {
        var sectionMonth = section.dataset.month || "";
        var folderMonthMatched = !selectedFolderMonth || sectionMonth === selectedFolderMonth;
        var monthSummary = { po: 0, qty: 0, amount: 0, rows: 0 };
        section.querySelectorAll(".summary-row").forEach(function(row) {
          var show = folderMonthMatched && (!summaryMonth || sectionMonth === summaryMonth) && inDateRange(row.dataset.day, summaryFrom, summaryTo);
          row.style.display = show ? "" : "none";
          if (show) {
            monthSummary.rows += 1;
            monthSummary.po += Number(row.dataset.poCount || "0");
            monthSummary.qty += parseMoney(row.querySelector(".summary-qty").textContent);
            monthSummary.amount += parseMoney(row.querySelector(".summary-amount").textContent);
          }
        });
        section.style.display = monthSummary.rows ? "" : "none";
        var vat = Math.round(monthSummary.amount / 1.1);
        var budget = Math.round(vat * 0.035);
        section.querySelectorAll(".summary-month-po").forEach(function(node) { node.textContent = monthSummary.po.toLocaleString("ko-KR"); });
        section.querySelectorAll(".summary-month-qty").forEach(function(node) { node.textContent = monthSummary.qty.toLocaleString("ko-KR"); });
        section.querySelectorAll(".summary-month-amount").forEach(function(node) { node.textContent = money(monthSummary.amount); });
        section.querySelectorAll(".summary-month-vat").forEach(function(node) { node.textContent = money(vat); });
        section.querySelectorAll(".summary-month-budget").forEach(function(node) { node.textContent = money(budget); });
      });

      var detailFrom = document.getElementById("detail-from")?.value || "";
      var detailTo = document.getElementById("detail-to")?.value || "";
      var detailKeywordNode = document.getElementById("detail-keyword");
      var detailKeyword = (detailKeywordNode?.value || "").trim().toLowerCase();
      var detailPoSelect = document.getElementById("detail-po-select");
      var detailViewMode = document.getElementById("detail-view-mode")?.value || "sku";
      function getRowPoList(row) {
        var poText = row.dataset.poList || row.querySelector(".po-cell")?.dataset.originalPo || "";
        return poText.split(",").map(function(po) { return po.trim(); }).filter(Boolean);
      }
      function cleanAttr(value) {
        return String(value).replace(/&/g, "&amp;").replace(/"/g, "&quot;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
      }
      function cleanText(value) {
        return String(value).replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
      }
      var selectedPo = (detailPoSelect?.value || "").trim();
      var availablePo = {};
      document.querySelectorAll('.detail-row[data-view-mode="po"]').forEach(function(row) {
        if ((row.dataset.month || "") < "2026-07") return;
        if (!inDateRange(row.dataset.day, detailFrom, detailTo)) return;
        getRowPoList(row).forEach(function(po) { availablePo[po] = true; });
      });
      if (detailPoSelect) {
        var previousPo = selectedPo;
        var poOptions = Object.keys(availablePo).sort();
        detailPoSelect.innerHTML = '<option value="">전체 PO</option>' + poOptions.map(function(po) {
          return '<option value="' + cleanAttr(po) + '">' + cleanText(po) + '</option>';
        }).join("");
        detailPoSelect.value = poOptions.includes(previousPo) ? previousPo : "";
        selectedPo = detailPoSelect.value;
      }
      var keywordLooksPo = /^\\d{6,}$/.test(detailKeyword);
      var requestedPo = selectedPo || (keywordLooksPo ? detailKeyword : "");
      if (selectedPo && detailKeywordNode && detailKeywordNode.value !== selectedPo) {
        detailKeywordNode.value = selectedPo;
        detailKeyword = selectedPo.toLowerCase();
        keywordLooksPo = true;
      }
      var visibleByDay = {};
      var visibleByMonth = {};
      document.querySelectorAll(".detail-row").forEach(function(row) {
        var haystack = (row.dataset.search || "").toLowerCase();
        var rowMode = row.dataset.viewMode || "sku";
        var poList = getRowPoList(row);
        var allowPoMonth = detailViewMode !== "po" || (row.dataset.month || "") >= "2026-07";
        var poMatched = detailViewMode !== "po" || !requestedPo || poList.includes(requestedPo);
        var keywordMatched = !detailKeyword || (keywordLooksPo ? poList.includes(detailKeyword) : haystack.includes(detailKeyword));
        var periodMatched = !selectedFolderMonth || (row.dataset.month || "") === selectedFolderMonth;
        var show = rowMode === detailViewMode && allowPoMonth && poMatched && periodMatched && inDateRange(row.dataset.day, detailFrom, detailTo) && keywordMatched;
        var displayPo = requestedPo && poList.includes(requestedPo) ? requestedPo : "";
        var poCell = row.querySelector(".po-cell");
        if (poCell) poCell.textContent = displayPo || (poCell.dataset.originalPo || poCell.textContent);
        row.classList.toggle("view-hidden", rowMode !== detailViewMode);
        row.style.display = show ? "" : "none";
        if (show) {
          var dayKey = row.dataset.day + "::" + rowMode;
          var qtyInput = row.querySelector(".qty-input");
          var qty = qtyInput ? Number(qtyInput.value || "0") : Number((row.children[2]?.textContent || "0").replaceAll(",", ""));
          var amount = parseMoney(row.querySelector(".row-amount")?.textContent || "0");
          if (!visibleByDay[dayKey]) visibleByDay[dayKey] = { qty: 0, amount: 0 };
          visibleByDay[dayKey].qty += qty;
          visibleByDay[dayKey].amount += amount;
          visibleByMonth[row.dataset.month + "::" + rowMode] = true;
        }
      });
      document.querySelectorAll(".detail-day-row").forEach(function(row) {
        var rowMode = row.dataset.viewMode || "sku";
        var dayKey = row.dataset.day + "::" + rowMode;
        var total = visibleByDay[dayKey];
        row.classList.toggle("view-hidden", rowMode !== detailViewMode);
        row.style.display = total ? "" : "none";
        var totalCell = row.querySelector("[data-day-total]");
        if (totalCell && total) {
          totalCell.textContent = row.dataset.day + " 합계: 수량 " + total.qty.toLocaleString("ko-KR") + "개 / 금액 " + money(total.amount);
        }
      });
      document.querySelectorAll(".month-folder").forEach(function(row) {
        var rowMode = row.dataset.viewMode || "sku";
        row.classList.toggle("view-hidden", rowMode !== detailViewMode);
        row.style.display = visibleByMonth[row.dataset.month + "::" + rowMode] ? "" : "none";
      });
      updateDetailResultSummary();
      recalcYearScreen();
    }
    function toggleSummaryMonth(button) {
      var section = button.closest(".summary-month-section");
      if (!section) return;
      var collapsed = section.classList.toggle("is-collapsed");
      button.textContent = collapsed ? "펼치기" : "숨기기";
    }    function resetLookup(group) {
      document.querySelectorAll('[data-lookup-group="' + group + '"] input, [data-lookup-group="' + group + '"] select').forEach(function(input) {
        input.value = "";
      });
      applyLookups();
    }
    function confirmExactDelete(label, expected) {
      var typed = prompt(label + " 삭제를 진행하려면 아래 값을 그대로 입력하세요.\\n\\n" + expected);
      return typed === expected;
    }
    function confirmDeleteDay(day) {
      return confirmExactDelete(day + " 일자", day);
    }
    function confirmDeleteMonth(month) {
      return confirmExactDelete(month + " 월 전체", month);
    }
    function setLookupDeleteDay(form) {
      var from = document.getElementById("detail-from")?.value || "";
      var to = document.getElementById("detail-to")?.value || "";
      if (!from || !to || from !== to) {
        alert("일자삭제는 시작일과 종료일을 같은 날짜로 선택해야 합니다.");
        return false;
      }
      form.querySelector('input[name="delete_day"]').value = from;
      return confirmDeleteDay(from);
    }
    function setLookupDeleteMonth(form) {
      var from = document.getElementById("detail-from")?.value || "";
      var to = document.getElementById("detail-to")?.value || "";
      var month = form.querySelector('input[name="delete_month"]').value;
      if (from && to && from.slice(0, 7) !== to.slice(0, 7)) {
        alert("월삭제는 시작일과 종료일이 같은 월 안에 있어야 합니다.");
        return false;
      }
      if (from) month = from.slice(0, 7);
      else if (to) month = to.slice(0, 7);
      form.querySelector('input[name="delete_month"]').value = month;
      return confirmDeleteMonth(month);
    }
    function toggleSection(targetId, button) {
      var target = document.getElementById(targetId);
      if (!target) return;
      var hidden = target.classList.toggle("is-hidden");
      button.textContent = hidden ? "펼치기" : "숨기기";
    }
    function toggleMonthFolder(month, button) {
      var folder = button.closest(".month-folder");
      var viewMode = folder?.dataset.viewMode || "sku";
      var closed = folder.classList.toggle("is-closed");
      document.querySelectorAll('[data-month="' + month + '"][data-view-mode="' + viewMode + '"]').forEach(function(row) {
        if (!row.classList.contains("month-folder")) row.classList.toggle("month-hidden", closed);
      });
      button.textContent = closed ? "펼치기" : "숨기기";
      applyLookups();
    }
    function prepareSalesSave(form) {
      var changedConfirmedMonths = new Set();
      form.querySelectorAll('.detail-row[data-view-mode="po"]').forEach(function(row) {
        var qty = row.querySelector('.qty-input');
        var memo = row.querySelector('.memo-input');
        var qtyChanged = qty && Number(qty.value || "0") !== Number(row.dataset.savedQty || row.dataset.originalQty || "0");
        var memoChanged = memo && memo.value.trim() !== String(row.dataset.savedMemo || "");
        if ((qtyChanged || memoChanged) && row.dataset.confirmed === "true") changedConfirmedMonths.add(row.dataset.month);
      });
      if (!changedConfirmedMonths.size) return true;
      var warning = "계산서 발행 확인이 완료된 건입니다. 수정하면 확인 당시의 내용과 달라질 수 있습니다. 그래도 수정하시겠습니까?";
      if (!confirm(warning)) return false;
      var reason = prompt("수정 사유를 입력해 주세요. 수정 사유가 없으면 저장할 수 없습니다.");
      if (!reason || !reason.trim()) { alert("수정 사유를 입력해야 합니다."); return false; }
      form.querySelector('input[name="override_reason"]').value = reason.trim();
      return true;
    }
    function syncBlankDateInputs() {
      document.querySelectorAll('input[type="date"]').forEach(function(input) {
        function sync() { input.classList.toggle("date-empty", !input.value); }
        sync();
        input.addEventListener("input", sync);
        input.addEventListener("change", sync);
      });
    }
    function applyFolderYearFilter() {
      var select = document.getElementById("folder-year-select");
      if (!select) return;
      var month = select.dataset.selectedMonth || "";
      document.querySelectorAll(".summary-month-section[data-month], .tester-panel details[data-month]").forEach(function(node) {
        node.style.display = !month || (node.dataset.month || "") === month ? "" : "none";
      });
      applyLookups();
    }
    function renderFolderMonthTabs(resetMonth) {
      var select = document.getElementById("folder-year-select");
      var container = document.getElementById("folder-month-tabs");
      if (!select || !container) return;
      var year = select.value;
      var available = new Set();
      document.querySelectorAll(".summary-month-section[data-month], .month-folder[data-month], .tester-panel details[data-month]").forEach(function(node) {
        var month = node.dataset.month || "";
        if (month.startsWith(year + "-")) available.add(month);
      });
      var current = resetMonth ? "" : (select.dataset.selectedMonth || "");
      if (!available.has(current)) current = Array.from(available).sort().reverse()[0] || "";
      select.dataset.selectedMonth = current;
      container.innerHTML = "";
      for (var number = 1; number <= 12; number++) {
        var month = year + "-" + String(number).padStart(2, "0");
        var button = document.createElement("button");
        button.type = "button";
        button.className = "folder-month-tab" + (available.has(month) ? " has-data" : "") + (month === current ? " is-active" : "");
        button.textContent = number + "월";
        button.disabled = !available.has(month);
        button.dataset.month = month;
        button.addEventListener("click", function() {
          select.dataset.selectedMonth = this.dataset.month;
          renderFolderMonthTabs(false);
          applyFolderYearFilter();
        });
        container.appendChild(button);
      }
    }
    function initFolderYearFilter() {
      var select = document.getElementById("folder-year-select");
      if (!select) return;
      var years = new Set();
      document.querySelectorAll("[data-month]").forEach(function(node) {
        var month = node.dataset.month || "";
        if (/^\\d{4}-\\d{2}$/.test(month)) years.add(month.slice(0, 4));
      });
      var sorted = Array.from(years).sort().reverse();
      select.innerHTML = sorted.map(function(year) { return '<option value="' + year + '">' + year + '년</option>'; }).join("");
      if (sorted.length) select.value = sorted[0];
      select.addEventListener("change", function() { renderFolderMonthTabs(true); applyFolderYearFilter(); });
      renderFolderMonthTabs(true);
      applyFolderYearFilter();
    }
    document.addEventListener("DOMContentLoaded", function() {
      initResizableTables();
      syncBlankDateInputs();
      initFolderYearFilter();
      document.querySelectorAll(".qty-input, .memo-input").forEach(function(input) {
        input.addEventListener("input", function() { recalcSalesScreen(); updateDetailResultSummary(); });
      });
      document.querySelectorAll(".money-input").forEach(function(input) {
        input.addEventListener("input", recalcYearScreen);
      });
      document.querySelectorAll(".lookup-row input, .lookup-row select").forEach(function(input) {
        input.addEventListener("input", applyLookups);
        input.addEventListener("change", applyLookups);
      });
      var detailPoSelect = document.getElementById("detail-po-select");
      var detailKeywordInput = document.getElementById("detail-keyword");
      if (detailPoSelect && detailKeywordInput) {
        detailPoSelect.addEventListener("change", function() {
          detailKeywordInput.value = detailPoSelect.value;
          applyLookups();
        });
      }
      var detailViewMode = document.getElementById("detail-view-mode");
      if (detailViewMode) {
        var modeNote = document.getElementById("mode-note");
        var detailSaveRow = document.getElementById("detail-save-row");
        var noteTimer;
        function showModeNote() {
          if (!modeNote) return;
          var rect = detailViewMode.getBoundingClientRect();
          modeNote.classList.add("is-visible");
          var noteRect = modeNote.getBoundingClientRect();
          var left = Math.min(Math.max(12, rect.left), window.innerWidth - noteRect.width - 12);
          var top = rect.top - noteRect.height - 8;
          if (top < 12) top = rect.bottom + 8;
          modeNote.style.left = left + "px";
          modeNote.style.top = top + "px";
          window.clearTimeout(noteTimer);
          noteTimer = window.setTimeout(function() { modeNote.classList.remove("is-visible"); }, 3000);
        }
        function syncDetailModeControls() {
          if (detailSaveRow) detailSaveRow.style.display = detailViewMode.value === "po" ? "flex" : "none";
        }
        detailViewMode.addEventListener("click", showModeNote);
        detailViewMode.addEventListener("focus", showModeNote);
        detailViewMode.addEventListener("change", function() {
          syncDetailModeControls();
          showModeNote();
        });
        syncDetailModeControls();
      }
      recalcSalesScreen();
      recalcYearScreen();
      applyLookups();
    });
  </script>
</head>
<body>
  <div class="app">
    <aside class="side">
      <div class="logo"><div class="logo-mark">B</div><div>보니애가구<br><span style="font-size:13px;font-weight:500;color:#b8c8d9;">업무 시스템</span></div></div>
      <div class="nav-item" onclick="location.href='/master'">쿠팡 기초자료 관리</div>
      <div class="nav-item" onclick="location.href='/'">PO변환</div>
      <div class="nav-item {folders_active}" onclick="location.href='/sales/folders'">월별납품관리</div>
      <div class="nav-item" onclick="location.href='/check'">수량검수</div>
      <div class="nav-item {sales_active}" onclick="location.href='/sales'">매출확인용</div>
      <div class="nav-item" onclick="location.href='/pallet'">파렛트/쉽먼트</div>
      <div class="nav-item" style="margin-top:28px;" onclick="showComingSoon('라벨 출력')">라벨 출력</div>
      <div class="nav-item" onclick="showComingSoon('배송완료 관리')">배송완료 관리</div>
    </aside>
    <main class="main">
      <h1>{page_title}</h1>
      <div class="sub">{page_sub}</div>
      {message}
      <div class="folder-year-filter {folder_filter_class}">
        <label for="folder-year-select">표시 연도</label>
        <select id="folder-year-select" aria-label="표시 연도"></select>
        <div id="folder-month-tabs" class="folder-month-tabs" aria-label="표시 월"></div>
      </div>
      <div class="cards">
        <section class="panel year-panel {year_section_class}">
          <div class="panel-head">
            <span>연도총매출</span>
            <div class="panel-actions">
              <span style="font-size:12px;color:#667085;">매출/VAT 별도/광고비예산/성장장려금은 월매출과 기초자료에서 자동 반영</span>
              <button class="toggle-btn" type="button" onclick="toggleSection('year-sales-content', this)">숨기기</button>
            </div>
          </div>
          <div id="year-sales-content" class="collapsible-content">
          <div class="lookup-row" data-lookup-group="year">
            <label>월 조회
              <select id="year-month-lookup">
                <option value="">전체</option>
                <option value="01">1월</option><option value="02">2월</option><option value="03">3월</option>
                <option value="04">4월</option><option value="05">5월</option><option value="06">6월</option>
                <option value="07">7월</option><option value="08">8월</option><option value="09">9월</option>
                <option value="10">10월</option><option value="11">11월</option><option value="12">12월</option>
              </select>
            </label>
            <button class="lookup-reset" type="button" onclick="resetLookup('year')">조회 초기화</button>
          </div>
          <form method="post" action="/sales/year/save">
          <div class="save-row"><button class="btn" type="submit">연도총매출 저장</button></div>
          <div class="scroll" style="max-height:360px;">
            <table class="year-table resizable-table">
              <colgroup>
                <col style="width:7%;">
                <col style="width:11%;">
                <col style="width:11%;">
                <col style="width:10%;">
                <col style="width:10%;">
                <col style="width:11%;">
                <col style="width:10%;">
                <col style="width:10%;">
                <col style="width:10%;">
                <col style="width:10%;">
                <col style="width:12%;">
              </colgroup>
              <thead><tr><th>월</th><th>매출<br>(VAT 미포함)</th><th>광고비예산<br>(VAT 미포함)</th><th>체험단<br>(VAT 미포함)</th><th>광고</th><th>광고비초과</th><th>성장장려금<br>(자동)</th><th>즉시할인<br>(VAT 미포함)</th><th>즉시할인<br>(다연채)</th><th>광고비</th><th>차감 후 금액<br>(자동)</th></tr></thead>
              <tbody>{year_rows}</tbody>
            </table>
          </div>
          </form>
          </div>
        </section>
        <section class="panel tester-panel {upload_section_class}">
          <div class="panel-head">
            <span>체험단 자료</span>
            <span style="font-size:12px;color:#667085;">엑셀 원본 보관 · 월별 금액 자동 반영 · 수기 수정 가능</span>
          </div>
          <div class="panel-body">
            <form method="post" action="/sales/tester/upload" enctype="multipart/form-data">
              <input type="file" name="tester_files" accept=".xlsx" multiple required>
              <button class="btn" type="submit">체험단 엑셀 올리기</button>
            </form>
            <div style="margin-top:9px;color:#667085;font-size:12px;">파일명에서 연·월을 확인하고 ‘총 진행 금액’을 체험단 비용에 자동 반영합니다. 자동 반영 후 위 연도총매출 표에서 금액을 수기로 고쳐 저장할 수도 있습니다.</div>
            <div style="margin-top:14px;">{tester_files}</div>
          </div>
        </section>
        <section class="panel upload-panel {upload_section_class}">
          <div class="panel-head">납품 PO 업로드</div>
          <div class="panel-body">
            <form method="post" action="/sales/upload" enctype="multipart/form-data">
              <input type="file" name="sales_po_files" accept=".xlsx" multiple required>
              <button class="btn" type="submit">납품 상품 매출에 반영</button>
            </form>
            <div style="margin-top:10px;color:#667085;font-size:13px;line-height:1.6;">
              심플웍스 등록 전후로 직접 수정한 최종 PO 복사본을 여기에 올리면, 스큐 아이디 기준으로 월별 납품 내역과 매출에 누적됩니다.
            </div>
            <div style="margin-top:8px;color:#667085;font-size:12px;line-height:1.5;">
              기존에 여러 PO가 한 줄로 묶인 자료는 해당 PO 원본 파일들을 한 번에 다시 올리면 PO별 상세로 교체됩니다.
            </div>
          </div>
        </section>
        <section class="panel summary-panel {summary_section_class}">
          <div class="panel-head">
            <span>일자별 합계</span>
            <div class="panel-actions">
              <a class="btn" href="/sales/download">월매출 엑셀 다운로드</a>
              <button class="toggle-btn" type="button" onclick="toggleSection('monthly-summary-content', this)">숨기기</button>
            </div>
          </div>
          <div id="monthly-summary-content" class="collapsible-content">
          <div class="lookup-row summary-lookup" data-lookup-group="summary">
            <label>조회월
              <input id="summary-month" type="month">
            </label>
            <label>시작일
              <input id="summary-from" type="date">
            </label>
            <label>종료일
              <input id="summary-to" type="date">
            </label>
            <button class="lookup-reset" type="button" onclick="resetLookup('summary')">조회 초기화</button>
          </div>
          <div class="summary-months">{summary_rows}</div>
          </div>
        </section>
        <section class="panel detail-panel">
          <div class="panel-head">{detail_title}</div>
          <div class="lookup-row" data-lookup-group="detail">
            <label>시작일
              <input id="detail-from" type="date">
            </label>
            <label>종료일
              <input id="detail-to" type="date">
            </label>
            <label>보기방식
              <select id="detail-view-mode">
                <option value="sku">SKU 합계</option>
                <option value="po">PO별 상세(2026년 7월부터)</option>
              </select>
            </label>
            <label>PO 선택
              <select id="detail-po-select">
                <option value="">전체 PO</option>
              </select>
            </label>
            <label>스큐/상품명/PO/메모
              <input id="detail-keyword" type="search" placeholder="찾을 내용 입력">
            </label>
            <button class="lookup-reset" type="button" onclick="resetLookup('detail')">조회 초기화</button>
            <form class="inline-delete-form" method="post" action="/sales/delete" onsubmit="return setLookupDeleteDay(this)">
              <input type="hidden" name="delete_day" value="">
              <button class="delete-btn" type="submit">조회 일자 삭제</button>
            </form>
            <form class="inline-delete-form" method="post" action="/sales/delete" onsubmit="return setLookupDeleteMonth(this)">
              <input type="hidden" name="delete_month" value="{current_month}">
              <button class="delete-btn" type="submit">조회 월 삭제</button>
            </form>
          </div>
          <div class="detail-result-bar">
            <div class="detail-result-summary">
              <span>검색 결과 <strong id="detail-result-count">0건</strong></span>
              <span>합계 수량 <strong id="detail-result-qty">0개</strong></span>
              <span>합계 금액 <strong id="detail-result-amount">0원</strong></span>
            </div>
            <button class="detail-download" type="button" onclick="downloadDetailResults()">검색결과 엑셀 다운로드</button>
          </div>
          <form method="post" action="/sales/save" onsubmit="return prepareSalesSave(this)">
          <input type="hidden" name="override_reason" value="">
          <div id="detail-save-row" class="save-row"><button class="btn" type="submit">수량/메모 저장</button></div><div id="mode-note" class="mode-note"><strong>보기방식 안내</strong>SKU 합계는 확인용이며, 저장/수정/삭제는 PO별 상세에서만 가능합니다.</div>
          <div class="scroll">
            <table class="detail-table resizable-table">
              <colgroup>
                <col style="width:10%;">
                <col style="width:37%;">
                <col style="width:8%;">
                <col style="width:9%;">
                <col style="width:10%;">
                <col style="width:8%;">
                <col style="width:13%;">
                <col style="width:5%;">
              </colgroup>
              <thead><tr><th>SKU ID</th><th>상품명</th><th>납품수량</th><th>단가</th><th>금액</th><th>비고</th><th>메모</th><th>삭제</th></tr></thead>
              <tbody>{detail_rows}</tbody>
            </table>
          </div>
          </form>
        </section>
      </div>
    </main>
  </div>
</body>
</html>"""


MASTER_PAGE = """<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>보니애가구 쿠팡 기초자료 관리</title>
  <style>
    :root {
      --bg: #f3f6fb;
      --panel: #ffffff;
      --ink: #172033;
      --muted: #667085;
      --line: #d9e1ee;
      --brand: #1f4e79;
      --brand2: #2b7a78;
      --danger: #b42318;
      --ok: #027a48;
      --shadow: 0 18px 45px rgba(16, 24, 40, .10);
    }
    * { box-sizing: border-box; }
    body { margin: 0; font-family: "Malgun Gothic", "맑은 고딕", Arial, sans-serif; color: var(--ink); background: var(--bg); }
    .app { min-height: 100vh; display: grid; grid-template-columns: 250px minmax(0, 1fr); }
    .side { background: #102a43; color: white; padding: 24px 18px; }
    .logo { display: flex; align-items: center; gap: 12px; font-weight: 800; font-size: 20px; margin-bottom: 28px; }
    .logo-mark { width: 34px; height: 34px; border-radius: 8px; background: #2b7a78; display: grid; place-items: center; font-weight: 900; }
    .nav-item { padding: 12px 13px; border-radius: 8px; color: #d9e7f3; margin-bottom: 6px; font-size: 14px; cursor: pointer; user-select: none; }
    .nav-item.active { background: rgba(255,255,255,.13); color: #fff; font-weight: 700; }
    .main { padding: 28px; min-width: 0; }
    h1 { font-size: 24px; margin: 0; letter-spacing: 0; }
    .sub { margin-top: 7px; color: var(--muted); font-size: 14px; }
    .panel { margin-top: 22px; background: var(--panel); border: 1px solid var(--line); border-radius: 8px; box-shadow: var(--shadow); overflow: hidden; }
    .add-grid { display: grid; grid-template-columns: 130px minmax(260px, 1.4fr) repeat(7, minmax(95px, 1fr)) 110px; gap: 10px; padding: 16px 18px; border-bottom: 1px solid var(--line); background: #fbfdff; align-items: end; }
    .field label { display: block; font-size: 12px; font-weight: 800; color: #344054; margin-bottom: 6px; }
    .field input[type=text], .field input[type=number] { width: 100%; border: 1px solid var(--line); border-radius: 8px; padding: 10px 11px; font-size: 14px; background: white; }
    .check-field { display: flex; gap: 7px; align-items: center; height: 39px; font-size: 13px; font-weight: 700; }
    .toolbar { display: grid; grid-template-columns: minmax(260px, 460px) auto; gap: 10px; align-items: end; padding: 16px 18px; border-bottom: 1px solid var(--line); }
    .search { width: min(460px, 100%); border: 1px solid var(--line); border-radius: 8px; padding: 11px 12px; font-size: 14px; }
    .toolbar-sort { display: flex; flex-direction: column; gap: 5px; color: #475467; font-size: 12px; font-weight: 800; }
    .toolbar-sort select { width: 100%; border: 1px solid #b9c6d8; border-radius: 8px; padding: 9px 10px; font: inherit; background: white; }
    .toolbar-actions { display: flex; gap: 10px; justify-content: flex-end; flex-wrap: wrap; }
    .filter-row { display: grid; grid-template-columns: repeat(4, minmax(130px, 1fr)); gap: 10px; padding: 12px 18px; border-bottom: 1px solid var(--line); background: #fbfcfe; }
    .filter-row label { display: flex; flex-direction: column; gap: 5px; color: #475467; font-size: 12px; font-weight: 800; }
    .filter-row select { width: 100%; border: 1px solid #b9c6d8; border-radius: 8px; padding: 9px 10px; font: inherit; background: white; }
    .filter-buttons { display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }
    .filter-chip { border: 1px solid #b9c6d8; background: #fff; color: #1f4e79; border-radius: 8px; padding: 11px 14px; font-weight: 900; cursor: pointer; font-size: 14px; }
    .filter-chip.active { background: var(--brand); color: #fff; border-color: var(--brand); }
    .filter-count { color: #475467; font-size: 13px; font-weight: 800; padding: 8px 0; }
    .btn { border: 0; border-radius: 8px; padding: 11px 16px; background: var(--brand); color: white; font-weight: 800; cursor: pointer; font-size: 14px; }
    .btn.secondary { background: #475467; }
    .status { padding: 13px 16px; border-radius: 8px; font-size: 14px; line-height: 1.6; margin-top: 16px; border: 1px solid var(--line); background: #fff; }
    .status.ok { border-color: #abefc6; background: #ecfdf3; color: var(--ok); }
    .status.err { border-color: #fecdca; background: #fef3f2; color: var(--danger); }
    .table-wrap { max-height: calc(100vh - 230px); overflow: auto; }
    table { width: 100%; min-width: 900px; border-collapse: collapse; table-layout: fixed; font-size: 13px; }
    th, td { border-right: 1px solid var(--line); border-bottom: 1px solid var(--line); padding: 8px; text-align: left; vertical-align: middle; overflow-wrap: anywhere; word-break: keep-all; }
    th:last-child, td:last-child { border-right: 0; }
    th { position: sticky; top: 0; background: #f8fafc; z-index: 1; color: #344054; }
    td.sku { font-weight: 700; color: #1f4e79; white-space: nowrap; }
    input.qty { width: 100%; min-width: 72px; border: 1px solid var(--line); border-radius: 6px; padding: 7px 8px; text-align: right; }
    .product { min-width: 420px; }
    .resizable-table th { position: sticky; }
    .col-resizer { position:absolute; top:0; right:-4px; width:8px; height:100%; cursor:col-resize; user-select:none; touch-action:none; z-index:3; }
    .col-resizer:hover, .col-resizer.active { background:rgba(31,78,121,.22); }
    .muted { color: var(--muted); font-size: 12px; }
    @media (max-width: 880px) {
      .app { grid-template-columns: 1fr; }
      .side { display: none; }
      .main { padding: 18px; }
      .toolbar { grid-template-columns: 1fr; align-items: stretch; }
      .filter-row { grid-template-columns: 1fr 1fr; }
    }
  </style>
  <script>
    function numberValue(value) {
      const cleaned = String(value || '').replace(/[^0-9.-]/g, '');
      const parsed = Number(cleaned);
      return Number.isFinite(parsed) ? parsed : 0;
    }
    function rowText(tr, selector) {
      const cell = tr.querySelector(selector);
      return cell ? cell.textContent.trim() : '';
    }
    function rowInputNumber(tr, inputNamePrefix) {
      const input = tr.querySelector('input[name^="' + inputNamePrefix + '"]');
      if (!input || !String(input.value || '').trim()) return null;
      return numberValue(input.value);
    }
    function compareNullableNumber(a, b, direction) {
      const aBlank = a === null;
      const bBlank = b === null;
      if (aBlank && bBlank) return 0;
      if (aBlank) return 1;
      if (bBlank) return -1;
      const result = a - b;
      return direction === 'desc' ? -result : result;
    }
    function sortMasterRows(rows) {
      const sort = document.getElementById('sort-master') ? document.getElementById('sort-master').value : '';
      if (!sort) return rows;
      const sorted = rows.slice();
      sorted.sort((a, b) => {
        if (sort === 'name-asc' || sort === 'name-desc') {
          const result = rowText(a, '.product').localeCompare(rowText(b, '.product'), 'ko-KR', { numeric: true });
          return sort === 'name-desc' ? -result : result;
        }
        if (sort === 'amount-asc' || sort === 'amount-desc') {
          return compareNullableNumber(rowInputNumber(a, 'amount_'), rowInputNumber(b, 'amount_'), sort === 'amount-desc' ? 'desc' : 'asc');
        }
        if (sort === 'simple-desc' || sort === 'simple-asc') {
          return compareNullableNumber(rowInputNumber(a, 'simple_no_'), rowInputNumber(b, 'simple_no_'), sort === 'simple-desc' ? 'desc' : 'asc');
        }
        return 0;
      });
      return sorted;
    }
    const masterFilterStorageKey = 'bonie-master-filters';
    function saveMasterFilterState() {
      const state = {
        q: document.getElementById('q') ? document.getElementById('q').value : '',
        unavailable: document.getElementById('filter-unavailable') ? document.getElementById('filter-unavailable').value : '',
        simpleNo: document.getElementById('filter-simple-no') ? document.getElementById('filter-simple-no').value : '',
        amount: document.getElementById('filter-amount') ? document.getElementById('filter-amount').value : '',
        sort: document.getElementById('sort-master') ? document.getElementById('sort-master').value : ''
      };
      localStorage.setItem(masterFilterStorageKey, JSON.stringify(state));
    }
    function restoreMasterFilterState() {
      let state = null;
      try {
        state = JSON.parse(localStorage.getItem(masterFilterStorageKey) || 'null');
      } catch (error) {
        state = null;
      }
      if (!state) return;
      if (document.getElementById('q')) document.getElementById('q').value = state.q || '';
      if (document.getElementById('filter-unavailable')) document.getElementById('filter-unavailable').value = state.unavailable || '';
      if (document.getElementById('filter-simple-no')) document.getElementById('filter-simple-no').value = state.simpleNo || '';
      if (document.getElementById('filter-amount')) document.getElementById('filter-amount').value = state.amount || '';
      if (document.getElementById('sort-master')) document.getElementById('sort-master').value = state.sort || '';
    }
    function filterRows() {
      const q = document.getElementById('q').value.trim().toLowerCase();
      const unavailable = document.getElementById('filter-unavailable').value;
      const simpleNo = document.getElementById('filter-simple-no').value;
      const amount = document.getElementById('filter-amount').value;
      const tbody = document.querySelector('tbody');
      const rows = Array.from(tbody.querySelectorAll('tr'));
      sortMasterRows(rows).forEach(tr => tbody.appendChild(tr));
      rows.forEach(tr => {
        const matchText = tr.innerText.toLowerCase().includes(q);
        const matchUnavailable = !unavailable || tr.dataset.unavailable === unavailable;
        const hasSimpleNo = rowInputNumber(tr, 'simple_no_') !== null ? '1' : '0';
        const hasAmount = rowInputNumber(tr, 'amount_') !== null ? '1' : '0';
        const matchSimple = !simpleNo || hasSimpleNo === simpleNo;
        const matchAmount = !amount || hasAmount === amount;
        tr.style.display = matchText && matchUnavailable && matchSimple && matchAmount ? '' : 'none';
      });
      saveMasterFilterState();
      updateFilterCount();
    }
    function resetMasterFilters() {
      document.getElementById('q').value = '';
      document.getElementById('filter-unavailable').value = '';
      document.getElementById('filter-simple-no').value = '';
      document.getElementById('filter-amount').value = '';
      const sortSelect = document.getElementById('sort-master');
      if (sortSelect) sortSelect.value = '';
      localStorage.removeItem(masterFilterStorageKey);
      updateFilterChips();
      filterRows();
    }
    function setMasterFilter(filterId, value) {
      document.getElementById(filterId).value = value;
      updateFilterChips();
      filterRows();
    }
    function updateFilterChips() {
      document.querySelectorAll('.filter-chip[data-filter-id]').forEach(btn => {
        const select = document.getElementById(btn.dataset.filterId);
        btn.classList.toggle('active', select && select.value === btn.dataset.value);
      });
    }
    function updateFilterCount() {
      const rows = Array.from(document.querySelectorAll('tbody tr'));
      const visible = rows.filter(tr => tr.style.display !== 'none').length;
      const counter = document.getElementById('filter-count');
      if (counter) counter.textContent = '현재 ' + visible.toLocaleString('ko-KR') + '개 표시 / 전체 ' + rows.length.toLocaleString('ko-KR') + '개';
    }
    function initResizableTables() {
      document.querySelectorAll(".resizable-table").forEach(function(table) {
        var cols = table.querySelectorAll("col");
        var headers = table.querySelectorAll("thead th");
        headers.forEach(function(header, index) {
          if (header.querySelector(".col-resizer")) return;
          var handle = document.createElement("span");
          handle.className = "col-resizer";
          header.appendChild(handle);
          handle.addEventListener("mousedown", function(event) {
            event.preventDefault();
            handle.classList.add("active");
            var startX = event.clientX;
            var startWidth = cols[index].getBoundingClientRect().width;
            function move(moveEvent) {
              var nextWidth = Math.max(70, startWidth + moveEvent.clientX - startX);
              cols[index].style.width = nextWidth + "px";
            }
            function stop() {
              handle.classList.remove("active");
              document.removeEventListener("mousemove", move);
              document.removeEventListener("mouseup", stop);
            }
            document.addEventListener("mousemove", move);
            document.addEventListener("mouseup", stop);
          });
        });
      });
    }
    window.addEventListener("load", initResizableTables);
    window.addEventListener("load", function() {
      restoreMasterFilterState();
      updateFilterChips();
      filterRows();
    });
    function showComingSoon(name) {
      alert(name + " 메뉴는 아직 준비 중입니다. 지금은 쿠팡 PO 변환과 기초자료 관리를 사용할 수 있습니다.");
    }
  </script>
</head>
<body>
  <div class="app">
    <aside class="side">
      <div class="logo"><div class="logo-mark">B</div><div>보니애가구<br><span style="font-size:13px;font-weight:500;color:#b8c8d9;">업무 시스템</span></div></div>
      <div class="nav-item active" onclick="location.href='/master'">쿠팡 기초자료 관리</div>
      <div class="nav-item" onclick="location.href='/'">PO변환</div>
      <div class="nav-item" onclick="location.href='/sales/folders'">월별납품관리</div>
      <div class="nav-item" onclick="location.href='/check'">수량검수</div>
      <div class="nav-item" onclick="location.href='/sales'">매출확인용</div>
      <div class="nav-item" onclick="location.href='/pallet'">파렛트/쉽먼트</div>
      <div class="nav-item" style="margin-top:28px;" onclick="showComingSoon('라벨 출력')">라벨 출력</div>
      <div class="nav-item" onclick="showComingSoon('배송완료 관리')">배송완료 관리</div>
    </aside>
    <main class="main">
      <h1>쿠팡 기초자료 관리</h1>
      <div class="sub">저장된 기초자료에서 금액, 심플웍스 No, 상품 사이즈, 무게, 납품불가 표시를 수정합니다.</div>
      {message}
      <section class="panel">
        <form method="post" action="/master/upload" enctype="multipart/form-data">
          <input type="hidden" name="next" value="master">
          <div class="toolbar">
            <label>기초자료 엑셀 다시 올리기
              <input type="file" name="master" accept=".xlsx" required>
            </label>
            <div class="toolbar-actions">
              <button class="btn secondary" type="submit">현재 기초자료 교체</button>
            </div>
          </div>
        </form>
      </section>
      <section class="panel">
        <form method="post" action="/master/growth-incentive/save">
          <div class="toolbar">
            <div>
              <strong style="font-size:17px;">성장장려금 기초자료</strong>
              <div style="margin-top:6px;color:#667085;font-size:13px;">계약서의 월별 기본계약과 분기 타입B 구간입니다. 저장 후 연도총매출에 자동 반영됩니다.</div>
            </div>
            <div class="toolbar-actions"><button class="btn" type="submit">성장장려금 기초자료 저장</button></div>
          </div>
          {incentive_tables}
        </form>
      </section>
      <form method="post" action="/master/save">
        <section class="panel">
          <div class="add-grid">
            <div class="field">
              <label>새 SKU ID</label>
              <input type="text" name="new_sku" placeholder="예: 12345678">
            </div>
            <div class="field">
              <label>상품명</label>
              <input type="text" name="new_name" placeholder="새 상품명을 입력">
            </div>
            <div class="field">
              <label>금액</label>
              <input type="number" name="new_amount" min="0" placeholder="0">
            </div>
            <div class="field">
              <label>심플웍스 No</label>
              <input type="text" name="new_simple_no" placeholder="예: 7463">
            </div>
            <div class="field">
              <label>가로(mm)</label>
              <input type="number" name="new_width_mm" min="0" placeholder="0">
            </div>
            <div class="field">
              <label>세로(mm)</label>
              <input type="number" name="new_depth_mm" min="0" placeholder="0">
            </div>
            <div class="field">
              <label>높이(mm)</label>
              <input type="number" name="new_height_mm" min="0" placeholder="0">
            </div>
            <div class="field">
              <label>무게(kg)</label>
              <input type="number" name="new_weight_kg" min="0" step="0.1" placeholder="0">
            </div>
            <div class="field">
              <label>바코드</label>
              <input type="text" name="new_barcode" placeholder="R...">
            </div>
            <label class="check-field">
              <input type="checkbox" name="new_unavailable" value="1"> 납품불가
            </label>
          </div>
          <div class="toolbar">
            <input class="search" id="q" oninput="filterRows()" placeholder="스큐 아이디 또는 상품명 검색" />
            <div class="toolbar-actions">
              <button class="btn secondary" type="button" onclick="location.href='/'">PO 변환으로 이동</button>
              <button class="btn" type="submit">수정 내용 저장</button>
            </div>
          </div>
          <div class="filter-row">
            <label>납품상태
              <select id="filter-unavailable" onchange="filterRows()">
                <option value="">전체 상품</option>
                <option value="0">납품가능 상품</option>
                <option value="1">납품불가 상품</option>
              </select>
              <span id="filter-count" class="filter-count"></span>
            </label>
            <label>정렬
              <select id="sort-master" onchange="filterRows()">
                <option value="">기본순</option>
                <option value="name-asc">상품명 오름차순</option>
                <option value="name-desc">상품명 내림차순</option>
                <option value="amount-asc">금액 낮은순</option>
                <option value="amount-desc">금액 높은순</option>
                <option value="simple-desc">심플웍스 No 큰수</option>
                <option value="simple-asc">심플웍스 No 작은수</option>
              </select>
            </label>
            <label>심플웍스 No
              <select id="filter-simple-no" onchange="filterRows()">
                <option value="">전체</option>
                <option value="1">입력 있음</option>
                <option value="0">입력 없음</option>
              </select>
            </label>
            <label>금액
              <select id="filter-amount" onchange="filterRows()">
                <option value="">전체</option>
                <option value="1">입력 있음</option>
                <option value="0">입력 없음</option>
              </select>
            </label>
          </div>
          <div class="table-wrap">
            <table class="resizable-table">
              <colgroup>
                <col style="width:130px;">
                <col style="width:420px;">
                <col style="width:120px;">
                <col style="width:120px;">
                <col style="width:95px;">
                <col style="width:95px;">
                <col style="width:95px;">
                <col style="width:95px;">
                <col style="width:100px;">
                <col style="width:170px;">
              </colgroup>
              <thead>
                <tr>
                  <th>SKU ID<br><span class="muted">스큐 아이디</span></th>
                  <th class="product">상품명</th>
                  <th>금액</th>
                  <th>심플웍스 No</th>
                  <th>가로(mm)</th>
                  <th>세로(mm)</th>
                  <th>높이(mm)</th>
                  <th>무게(kg)</th>
                  <th>납품불가</th>
                  <th>바코드</th>
                </tr>
              </thead>
              <tbody>
                {rows}
              </tbody>
            </table>
          </div>
        </section>
      </form>
    </main>
  </div>
</body>
</html>"""


def safe_name(name: str) -> str:
    return Path(name).name.replace("/", "_").replace("\\", "_")


def get_local_master_path() -> Path | None:
    if not MASTER_DIR.exists():
        return None
    files = [
        path
        for path in MASTER_DIR.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def get_saved_master_path() -> Path | None:
    master_path = get_local_master_path()
    if master_path is not None:
        return master_path
    restore_file_from_supabase("master_file", MASTER_DIR / "기초자료.xlsx")
    return get_local_master_path()


def norm_header(value: object) -> str:
    return "".join(str(value or "").lower().split())


def find_header_col(ws, candidates: list[str]) -> int | None:
    headers = {norm_header(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    normalized = [norm_header(candidate) for candidate in candidates]
    for candidate in normalized:
        if candidate in headers:
            return headers[candidate]
    for header, col in headers.items():
        if any(candidate in header for candidate in normalized):
            return col
    return None


def find_master_columns(ws) -> dict[str, int]:
    sku_col = find_header_col(ws, ["SKU ID", "상품코드", "스큐 아이디", "스큐"])
    qty_col = find_header_col(ws, ["요청", "납품가능수량", "수량"])
    name_col = find_header_col(ws, ["상품명"])
    amount_col = find_header_col(ws, ["금액", "단가", "매입가", "공급가"])
    simple_no_col = find_header_col(ws, ["심플웍스 No", "심플웍스번호", "심플웍스NO", "심플 No", "심플번호"])
    barcode_col = find_header_col(ws, ["바코드"])
    unavailable_col = find_header_col(ws, ["납품불가", "납품불가여부", "제외"])
    width_col = find_header_col(ws, ["길이 (mm)", "길이(mm)", "길이", "가로(mm)", "가로", "폭", "width", "width_mm"])
    depth_col = find_header_col(ws, ["넓이 (mm)", "넓이(mm)", "넓이", "세로(mm)", "세로", "깊이", "depth", "depth_mm"])
    height_col = find_header_col(ws, ["높이(mm)", "높이", "height", "height_mm"])
    weight_col = find_header_col(ws, ["무게(kg)", "무게", "중량", "weight", "weight_kg"])
    if unavailable_col is None:
        for col in range(1, ws.max_column + 1):
            if any("납품불가" in str(ws.cell(row, col).value or "") for row in range(2, min(ws.max_row, 80) + 1)):
                unavailable_col = col
                break
    if unavailable_col is None:
        insert_at = (name_col or 3) + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "납품불가"
        unavailable_col = insert_at
        if barcode_col and barcode_col >= insert_at:
            barcode_col += 1
        if amount_col and amount_col >= insert_at:
            amount_col += 1
        if simple_no_col and simple_no_col >= insert_at:
            simple_no_col += 1
    if amount_col is None:
        insert_at = (name_col or 3) + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "금액"
        amount_col = insert_at
        if barcode_col and barcode_col >= insert_at:
            barcode_col += 1
        if unavailable_col and unavailable_col >= insert_at:
            unavailable_col += 1
        if simple_no_col and simple_no_col >= insert_at:
            simple_no_col += 1
    if simple_no_col is None:
        insert_at = (amount_col or name_col or 3) + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "심플웍스 No"
        simple_no_col = insert_at
        if barcode_col and barcode_col >= insert_at:
            barcode_col += 1
        if unavailable_col and unavailable_col >= insert_at:
            unavailable_col += 1
    width_col = find_header_col(ws, ["길이 (mm)", "길이(mm)", "길이", "가로(mm)", "가로", "폭", "width", "width_mm"])
    depth_col = find_header_col(ws, ["넓이 (mm)", "넓이(mm)", "넓이", "세로(mm)", "세로", "깊이", "depth", "depth_mm"])
    height_col = find_header_col(ws, ["높이(mm)", "높이", "height", "height_mm"])
    weight_col = find_header_col(ws, ["무게(kg)", "무게", "중량", "weight", "weight_kg"])
    size_columns = [
        ("width_mm", width_col, "가로(mm)"),
        ("depth_mm", depth_col, "세로(mm)"),
        ("height_mm", height_col, "높이(mm)"),
        ("weight_kg", weight_col, "무게(kg)"),
    ]
    resolved_size_columns: dict[str, int] = {}
    for key, col, header in size_columns:
        if col is None:
            col = ws.max_column + 1
            ws.cell(1, col).value = header
        resolved_size_columns[key] = col
    if sku_col is None or qty_col is None or name_col is None:
        raise ValueError("기초자료에서 SKU ID, 요청, 상품명 열을 찾지 못했습니다.")
    return {
        "sku": sku_col,
        "qty": qty_col,
        "name": name_col,
        "amount": amount_col,
        "simple_no": simple_no_col,
        **resolved_size_columns,
        "barcode": barcode_col or 0,
        "unavailable": unavailable_col,
    }


def normalize_master_file(master_path: Path) -> None:
    wb = load_workbook(master_path)
    ws = wb.active
    find_master_columns(ws)
    wb.save(master_path)


def replace_saved_master_file(name: str, master_bytes: bytes) -> Path:
    MASTER_DIR.mkdir(parents=True, exist_ok=True)
    current_master = get_local_master_path()
    if current_master is not None:
        backup_dir = MASTER_DIR / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"{current_master.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(current_master, backup_path)
    for path in MASTER_DIR.glob("*.xlsx"):
        if path.is_file() and not path.name.startswith("~$"):
            path.unlink(missing_ok=True)
    target_path = MASTER_DIR / name
    target_path.write_bytes(master_bytes)
    normalize_master_file(target_path)
    backup_master_file_to_supabase(target_path)
    return target_path


def render_master_rows(master_path: Path) -> str:
    wb = load_workbook(master_path, data_only=True)
    ws = wb.active
    cols = find_master_columns(ws)
    rows: list[str] = []
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row, cols["sku"]).value or "").strip()
        name = str(ws.cell(row, cols["name"]).value or "").strip()
        if not sku and not name:
            continue
        amount = str(ws.cell(row, cols["amount"]).value or "").strip()
        simple_no = str(ws.cell(row, cols["simple_no"]).value or "").strip()
        width_mm = str(ws.cell(row, cols["width_mm"]).value or "").strip()
        depth_mm = str(ws.cell(row, cols["depth_mm"]).value or "").strip()
        height_mm = str(ws.cell(row, cols["height_mm"]).value or "").strip()
        weight_kg = str(ws.cell(row, cols["weight_kg"]).value or "").strip()
        barcode = str(ws.cell(row, cols["barcode"]).value or "").strip() if cols["barcode"] else ""
        unavailable = "납품불가" in str(ws.cell(row, cols["unavailable"]).value or "")
        checked = " checked" if unavailable else ""
        simple_flag = "1" if simple_no else "0"
        amount_flag = "1" if parse_int(amount) > 0 else "0"
        unavailable_flag = "1" if unavailable else "0"
        rows.append(
            f'<tr data-unavailable="{unavailable_flag}" data-simple-no="{simple_flag}" data-amount="{amount_flag}">'
            f'<td class="sku">{html.escape(sku)}<input type="hidden" name="row" value="{row}"></td>'
            f'<td class="product">{html.escape(name)}</td>'
            f'<td><input class="qty" name="amount_{row}" value="{html.escape(amount)}" inputmode="numeric"></td>'
            f'<td><input class="qty" name="simple_no_{row}" value="{html.escape(simple_no)}" inputmode="numeric"></td>'
            f'<td><input class="qty" name="width_mm_{row}" value="{html.escape(width_mm)}" inputmode="numeric"></td>'
            f'<td><input class="qty" name="depth_mm_{row}" value="{html.escape(depth_mm)}" inputmode="numeric"></td>'
            f'<td><input class="qty" name="height_mm_{row}" value="{html.escape(height_mm)}" inputmode="numeric"></td>'
            f'<td><input class="qty" name="weight_kg_{row}" value="{html.escape(weight_kg)}" inputmode="decimal"></td>'
            f'<td><input type="checkbox" name="unavailable_{row}" value="1"{checked}></td>'
            f"<td>{html.escape(barcode)}</td>"
            "</tr>"
        )
    return "\n".join(rows)


def save_master_form(form: cgi.FieldStorage) -> int:
    master_path = get_saved_master_path()
    if master_path is None:
        raise ValueError("저장된 기초자료가 없습니다.")

    backup_dir = MASTER_DIR / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{master_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(master_path, backup_path)

    wb = load_workbook(master_path)
    ws = wb.active
    cols = find_master_columns(ws)
    if not ws.cell(1, cols["unavailable"]).value:
        ws.cell(1, cols["unavailable"]).value = "납품불가"

    changed = 0
    new_sku = form["new_sku"].value.strip() if "new_sku" in form else ""
    new_name = form["new_name"].value.strip() if "new_name" in form else ""
    if new_sku or new_name:
        if not new_sku:
            raise ValueError("새 상품을 추가하려면 SKU ID를 입력해야 합니다.")
        if not new_name:
            raise ValueError("새 상품을 추가하려면 상품명을 입력해야 합니다.")
        existing_skus = {
            str(ws.cell(row, cols["sku"]).value or "").strip()
            for row in range(2, ws.max_row + 1)
        }
        if new_sku in existing_skus:
            raise ValueError(f"이미 등록된 SKU ID입니다: {new_sku}")
        new_row = ws.max_row + 1
        ws.cell(new_row, cols["sku"]).value = new_sku
        ws.cell(new_row, cols["qty"]).value = ""
        ws.cell(new_row, cols["name"]).value = new_name
        ws.cell(new_row, cols["amount"]).value = form["new_amount"].value.strip() if "new_amount" in form else ""
        ws.cell(new_row, cols["simple_no"]).value = form["new_simple_no"].value.strip() if "new_simple_no" in form else ""
        ws.cell(new_row, cols["width_mm"]).value = form["new_width_mm"].value.strip() if "new_width_mm" in form else ""
        ws.cell(new_row, cols["depth_mm"]).value = form["new_depth_mm"].value.strip() if "new_depth_mm" in form else ""
        ws.cell(new_row, cols["height_mm"]).value = form["new_height_mm"].value.strip() if "new_height_mm" in form else ""
        ws.cell(new_row, cols["weight_kg"]).value = form["new_weight_kg"].value.strip() if "new_weight_kg" in form else ""
        if cols["barcode"]:
            ws.cell(new_row, cols["barcode"]).value = form["new_barcode"].value.strip() if "new_barcode" in form else ""
        ws.cell(new_row, cols["unavailable"]).value = "납품불가" if "new_unavailable" in form else ""
        changed += 1

    row_values = form["row"] if "row" in form else []
    if not isinstance(row_values, list):
        row_values = [row_values]
    for item in row_values:
        row = int(item.value)
        amount_key = f"amount_{row}"
        simple_no_key = f"simple_no_{row}"
        width_key = f"width_mm_{row}"
        depth_key = f"depth_mm_{row}"
        height_key = f"height_mm_{row}"
        weight_key = f"weight_kg_{row}"
        unavailable_key = f"unavailable_{row}"
        new_amount = form[amount_key].value.strip() if amount_key in form else ""
        new_simple_no = form[simple_no_key].value.strip() if simple_no_key in form else ""
        new_width = form[width_key].value.strip() if width_key in form else ""
        new_depth = form[depth_key].value.strip() if depth_key in form else ""
        new_height = form[height_key].value.strip() if height_key in form else ""
        new_weight = form[weight_key].value.strip() if weight_key in form else ""
        new_unavailable = "납품불가" if unavailable_key in form else ""
        old_amount = str(ws.cell(row, cols["amount"]).value or "").strip()
        old_simple_no = str(ws.cell(row, cols["simple_no"]).value or "").strip()
        old_width = str(ws.cell(row, cols["width_mm"]).value or "").strip()
        old_depth = str(ws.cell(row, cols["depth_mm"]).value or "").strip()
        old_height = str(ws.cell(row, cols["height_mm"]).value or "").strip()
        old_weight = str(ws.cell(row, cols["weight_kg"]).value or "").strip()
        old_unavailable = "납품불가" if "납품불가" in str(ws.cell(row, cols["unavailable"]).value or "") else ""
        if new_amount != old_amount:
            ws.cell(row, cols["amount"]).value = new_amount
            changed += 1
        if new_simple_no != old_simple_no:
            ws.cell(row, cols["simple_no"]).value = new_simple_no
            changed += 1
        if new_width != old_width:
            ws.cell(row, cols["width_mm"]).value = new_width
            changed += 1
        if new_depth != old_depth:
            ws.cell(row, cols["depth_mm"]).value = new_depth
            changed += 1
        if new_height != old_height:
            ws.cell(row, cols["height_mm"]).value = new_height
            changed += 1
        if new_weight != old_weight:
            ws.cell(row, cols["weight_kg"]).value = new_weight
            changed += 1
        if new_unavailable != old_unavailable:
            ws.cell(row, cols["unavailable"]).value = new_unavailable
            changed += 1
    wb.save(master_path)
    backup_master_file_to_supabase(master_path)
    return changed


def update_master_amounts_from_lines(lines) -> int:
    master_path = get_saved_master_path()
    if master_path is None:
        return 0

    price_by_sku: dict[str, int] = {}
    for line in lines:
        sku = str(line.sku_id or "").strip()
        if not sku or sku in price_by_sku:
            continue
        prefer_inbound = has_inbound_sales_data(lines)
        sales_qty = get_sales_qty(line, prefer_inbound)
        sales_amount = get_sales_amount(line, prefer_inbound)
        if sales_qty <= 0 or sales_amount <= 0:
            continue
        unit_price = round(sales_amount / sales_qty) if sales_qty else parse_int(getattr(line, "purchase_price", 0))
        if unit_price <= 0 and line.available_qty:
            unit_price = round(parse_int(getattr(line, "order_amount", 0)) / line.available_qty)
        if unit_price > 0:
            price_by_sku[sku] = unit_price

    if not price_by_sku:
        return 0

    wb = load_workbook(master_path)
    ws = wb.active
    cols = find_master_columns(ws)
    changed = 0
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row, cols["sku"]).value or "").strip()
        if sku not in price_by_sku:
            continue
        current_amount = parse_int(ws.cell(row, cols["amount"]).value)
        if current_amount > 0:
            continue
        ws.cell(row, cols["amount"]).value = price_by_sku[sku]
        changed += 1

    if changed:
        wb.save(master_path)
    return changed


def get_master_amounts_by_sku() -> dict[str, int]:
    master_path = get_saved_master_path()
    if master_path is None:
        return {}
    wb = load_workbook(master_path, data_only=True)
    ws = wb.active
    cols = find_master_columns(ws)
    amounts: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row, cols["sku"]).value or "").strip()
        amount = parse_int(ws.cell(row, cols["amount"]).value)
        if sku and amount > 0:
            amounts[sku] = amount
    return amounts


def parse_month(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return datetime.now().strftime("%Y-%m")
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19], fmt).strftime("%Y-%m")
        except ValueError:
            continue
    if len(text) >= 7:
        return text[:7].replace("/", "-")
    return datetime.now().strftime("%Y-%m")


def parse_date(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return datetime.now().strftime("%Y-%m-%d")
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    if len(text) >= 10:
        return text[:10].replace("/", "-")
    return datetime.now().strftime("%Y-%m-%d")


def parse_int(value, default: int = 0) -> int:
    text = str(value or "").replace(",", "").replace("원", "").strip()
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def has_inbound_sales_data(lines) -> bool:
    return any(
        parse_int(getattr(line, "inbound_qty", 0)) > 0
        or parse_int(getattr(line, "inbound_amount", 0)) > 0
        for line in lines
    )


def get_sales_qty(line, prefer_inbound: bool = False) -> int:
    inbound_qty = parse_int(getattr(line, "inbound_qty", 0))
    if prefer_inbound:
        return inbound_qty
    return parse_int(getattr(line, "available_qty", 0))


def get_sales_amount(line, prefer_inbound: bool = False) -> int:
    inbound_amount = parse_int(getattr(line, "inbound_amount", 0))
    if prefer_inbound:
        return inbound_amount
    return parse_int(getattr(line, "order_amount", 0))


def describe_sales_date_breakdown(lines) -> str:
    prefer_inbound = has_inbound_sales_data(lines)
    by_day = defaultdict(lambda: {"qty": 0, "amount": 0, "po_numbers": set()})
    for line in lines:
        qty = get_sales_qty(line, prefer_inbound)
        amount = get_sales_amount(line, prefer_inbound)
        if qty <= 0 and amount <= 0:
            continue
        day = parse_date(line.inbound_date)
        by_day[day]["qty"] += qty
        by_day[day]["amount"] += amount
        by_day[day]["po_numbers"].add(line.po_no)
    if not by_day:
        return ""
    parts = []
    for day in sorted(by_day):
        values = by_day[day]
        parts.append(
            f"{day}: PO {len(values['po_numbers'])}개, 수량 {values['qty']:,}개, 금액 {values['amount']:,}원"
        )
    return " 입고예정일 기준으로 " + " / ".join(parts) + "으로 나누어 등록했습니다."



def split_po_numbers(value: object) -> list[str]:
    parts = []
    for part in str(value or "").replace("PO:", "").split(","):
        po_no = part.strip()
        if po_no and po_no not in parts:
            parts.append(po_no)
    return parts


def find_uploaded_sales_po_files() -> list[Path]:
    if not RUNS_DIR.exists():
        return []
    files = []
    for path in RUNS_DIR.glob("sales_*/uploaded_po/*.xlsx"):
        if path.name.startswith("~$"):
            continue
        files.append(path)
    return sorted(files, key=lambda p: (p.stat().st_mtime, str(p)))

def ensure_monthly_sales_book():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    restore_sales_files_from_supabase()
    headers = ["일자", "월", "입고예정일", "SKU ID", "상품명", "납품수량", "금액", "바코드", "비고", "수정수량", "수정메모", "PO번호"]
    if SALES_LEDGER_PATH.exists():
        wb = load_workbook(SALES_LEDGER_PATH)
        ws = wb.active
        changed = False
        for col, header in enumerate(headers, start=1):
            if ws.cell(1, col).value != header:
                ws.cell(1, col).value = header
                changed = True
        if changed:
            wb.save(SALES_LEDGER_PATH)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "월매출"
        ws.append(headers)
        wb.save(SALES_LEDGER_PATH)
    return wb, ws


def update_monthly_sales(lines) -> tuple[int, int]:
    wb, ws = ensure_monthly_sales_book()
    po_numbers = {po for line in lines for po in split_po_numbers(getattr(line, "po_no", ""))}
    prefer_inbound = has_inbound_sales_data(lines)
    master_amounts = get_master_amounts_by_sku()

    existing_adjustments = {}
    rows_to_keep = []
    for row in range(2, ws.max_row + 1):
        day = str(ws.cell(row, 1).value or "").strip()
        sku = str(ws.cell(row, 4).value or "").strip()
        remarks = str(ws.cell(row, 9).value or "")
        row_po_no = str(ws.cell(row, 12).value or "")
        existing_po_numbers = set(split_po_numbers(row_po_no) or split_po_numbers(remarks))
        if existing_po_numbers & po_numbers and not existing_po_numbers <= po_numbers:
            needed = ", ".join(sorted(existing_po_numbers))
            raise ValueError(f"기존 자료에 PO {needed}가 한 줄로 묶여 있습니다. 정확히 재정리하려면 이 PO 원본 파일들을 한 번에 같이 올려주세요.")
        adjusted_qty = str(ws.cell(row, 10).value or "").strip()
        adjusted_memo = str(ws.cell(row, 11).value or "").strip()
        for existing_po in existing_po_numbers:
            if adjusted_qty or adjusted_memo:
                existing_adjustments[(day, existing_po, sku)] = (adjusted_qty, adjusted_memo)
        if not (existing_po_numbers & po_numbers):
            rows_to_keep.append([ws.cell(row, col).value for col in range(1, 13)])

    ws.delete_rows(2, max(ws.max_row - 1, 0))
    for row_values in rows_to_keep:
        ws.append(row_values)

    grouped = defaultdict(lambda: {
        "month": "",
        "inbound_date": "",
        "po_no": "",
        "sku": "",
        "name": "",
        "qty": 0,
        "amount": 0,
        "barcode": "",
    })
    for line in lines:
        sku = str(line.sku_id or "").strip()
        sales_qty = get_sales_qty(line, prefer_inbound)
        master_unit_price = master_amounts.get(sku, 0)
        sales_amount = master_unit_price * sales_qty if master_unit_price > 0 else get_sales_amount(line, prefer_inbound)
        if sales_qty <= 0 and sales_amount <= 0:
            continue
        day = parse_date(line.inbound_date)
        for po_no in split_po_numbers(getattr(line, "po_no", "")):
            key = (day, po_no, sku)
            item = grouped[key]
            item["month"] = parse_month(line.inbound_date)
            item["inbound_date"] = line.inbound_date
            item["po_no"] = po_no
            item["sku"] = sku
            item["name"] = item["name"] or line.product_name
            item["qty"] += sales_qty
            item["amount"] += sales_amount
            item["barcode"] = item["barcode"] or line.barcode

    saved_count = 0
    for (_day, _po_no, _sku), item in sorted(grouped.items()):
        adjusted_qty, adjusted_memo = existing_adjustments.get((_day, _po_no, _sku), ("", ""))
        ws.append([
            _day,
            item["month"],
            item["inbound_date"],
            item["sku"],
            item["name"],
            item["qty"],
            item["amount"],
            item["barcode"],
            item["po_no"],
            adjusted_qty,
            adjusted_memo,
            item["po_no"],
        ])
        saved_count += 1

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 55)
    wb.save(SALES_LEDGER_PATH)
    write_sales_display_workbook()
    backup_sales_files_to_supabase()
    return saved_count, sum(item["amount"] for item in grouped.values())


def rebuild_monthly_sales_from_uploaded_pos() -> tuple[int, int, int]:
    latest_by_po: dict[str, tuple[Path, list[object]]] = {}
    for po_path in find_uploaded_sales_po_files():
        try:
            lines = read_po_lines(po_path)
        except Exception as exc:
            print(f"[sales] skipped uploaded PO {po_path}: {exc}", flush=True)
            continue
        po_keys = sorted({po for line in lines for po in split_po_numbers(getattr(line, "po_no", ""))})
        for po_no in po_keys:
            latest_by_po[po_no] = (po_path, [line for line in lines if po_no in split_po_numbers(getattr(line, "po_no", ""))])
    if not latest_by_po:
        raise ValueError("재정리할 업로드 PO 원본 파일을 찾지 못했습니다.")
    all_lines = []
    used_files = set()
    for _po_no, (po_path, lines) in sorted(latest_by_po.items()):
        all_lines.extend(lines)
        used_files.add(po_path)
    saved_count, total_amount = update_monthly_sales(all_lines)
    return len(latest_by_po), saved_count, total_amount

def load_monthly_sales_summary(limit_rows: int | None = 300, aggregate_by_sku: bool = False) -> tuple[list[tuple[str, int, int, int]], list[list[object]]]:
    restore_sales_files_from_supabase()
    if not SALES_LEDGER_PATH.exists():
        return [], []
    master_amounts = get_master_amounts_by_sku()
    wb = load_workbook(SALES_LEDGER_PATH, data_only=True)
    ws = wb.active
    summary = defaultdict(lambda: {"qty": 0, "amount": 0, "pos": set()})
    rows = []
    for row in range(2, ws.max_row + 1):
        day = str(ws.cell(row, 1).value or "")
        sku = str(ws.cell(row, 4).value or "")
        name = str(ws.cell(row, 5).value or "")
        original_qty = parse_int(ws.cell(row, 6).value)
        original_amount = parse_int(ws.cell(row, 7).value)
        remarks = str(ws.cell(row, 9).value or "")
        row_po_no = str(ws.cell(row, 12).value or "").strip() or ", ".join(split_po_numbers(remarks))
        adjusted_qty_raw = str(ws.cell(row, 10).value or "").strip()
        adjusted_memo = str(ws.cell(row, 11).value or "").strip()
        adjusted_qty = parse_int(adjusted_qty_raw, original_qty) if adjusted_qty_raw else original_qty
        master_unit_price = master_amounts.get(str(sku).strip(), 0)
        unit_price = master_unit_price if master_unit_price > 0 else (round(original_amount / original_qty) if original_qty else 0)
        amount = unit_price * adjusted_qty
        clean_po_numbers = []
        if not day:
            continue
        summary[day]["qty"] += adjusted_qty
        summary[day]["amount"] += amount
        for part in split_po_numbers(row_po_no or remarks):
            summary[day]["pos"].add(part)
            clean_po_numbers.append(part)
        rows.append([
            row,
            day,
            sku,
            name,
            original_qty,
            adjusted_qty,
            unit_price,
            amount,
            row_po_no or ", ".join(clean_po_numbers),
            adjusted_memo,
            adjusted_qty != original_qty or bool(adjusted_memo),
        ])
    if aggregate_by_sku:
        grouped_rows = {}
        for row_no, day, sku, name, original_qty, qty, unit_price, amount, remarks, memo, changed in rows:
            key = (day, sku)
            item = grouped_rows.setdefault(key, {
                "row_no": row_no,
                "day": day,
                "sku": sku,
                "name": name,
                "original_qty": 0,
                "qty": 0,
                "unit_price": unit_price,
                "amount": 0,
                "remarks": set(),
                "memo": [],
                "changed": False,
            })
            item["original_qty"] += original_qty
            item["qty"] += qty
            item["amount"] += amount
            item["changed"] = bool(item["changed"] or changed)
            for po_no in split_po_numbers(remarks):
                item["remarks"].add(po_no)
            if memo:
                item["memo"].append(str(memo))
        rows = [
            [
                item["row_no"],
                item["day"],
                item["sku"],
                item["name"],
                item["original_qty"],
                item["qty"],
                item["unit_price"],
                item["amount"],
                ", ".join(sorted(item["remarks"])),
                " / ".join(item["memo"]),
                item["changed"],
            ]
            for item in grouped_rows.values()
        ]
    summary_rows = [
        (month, values["qty"], values["amount"], len(values["pos"]))
        for month, values in sorted(summary.items())
    ]
    if limit_rows is None:
        return summary_rows, rows
    return summary_rows, rows[-limit_rows:]


def render_growth_incentive_tables() -> str:
    config = load_config(GROWTH_INCENTIVE_PATH)

    def period_table(key: str, period: dict, visible: bool) -> str:
        period_no = int(period["period"])
        cards = []
        for band in period["bands"]:
            number = int(band["band"])
            upper = "" if band.get("upper") is None else f'{int(band["upper"]):,}'
            rate = f'{float(band["rate"]) * 100:g}'
            level = "low" if number <= 3 else ("mid" if number <= 6 else "high")
            cards.append(
                f'<div class="growth-tier-card {level}">'
                f'<div class="growth-tier-head"><span class="growth-tier-badge">{number}등급</span>'
                f'<label class="growth-rate"><input name="{key}_{period_no}_rate_{number}" type="number" min="0" max="100" step="0.1" value="{rate}" required><b>%</b></label></div>'
                '<div class="growth-range">'
                f'<label><span>초과</span><input class="growth-money" name="{key}_{period_no}_lower_{number}" inputmode="numeric" value="{int(band["lower"]):,}" required><b>원</b></label>'
                '<div class="growth-range-arrow">↓</div>'
                f'<label><span>이하</span><input class="growth-money" name="{key}_{period_no}_upper_{number}" inputmode="numeric" value="{upper}" placeholder="상한 없음"><b>원</b></label>'
                '</div></div>'
            )
        display = "block" if visible else "none"
        return (
            f'<div class="growth-period" data-growth-group="{key}" data-growth-period="{period_no}" style="display:{display};">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;margin:0 0 10px;gap:12px;">'
            f'<h3 style="margin:0;color:#17365d;">{html.escape(period["label"])} · 9등급</h3>'
            f'<span style="color:#667085;font-size:13px;">산정기간 {html.escape(period["start"])} ~ {html.escape(period["end"])}</span></div>'
            f'<div class="growth-tier-grid">{"".join(cards)}</div></div>'
        )

    def group_html(key: str, title: str, description: str) -> str:
        buttons = "".join(
            f'<button class="growth-period-btn{" active" if index == 0 else ""}" type="button" data-growth-button="{key}-{int(period["period"])}" onclick="showGrowthPeriod(\'{key}\',{int(period["period"])},this)">{html.escape(period["label"])}</button>'
            for index, period in enumerate(config[key])
        )
        tables = "".join(period_table(key, period, index == 0) for index, period in enumerate(config[key]))
        opened = " open" if key == "monthly" else ""
        toggle_text = "숨기기" if key == "monthly" else "펼치기"
        return (
            f'<details class="growth-group"{opened} style="border:1px solid #d9e1ee;border-radius:12px;background:#fff;margin-bottom:16px;overflow:hidden;">'
            f'<summary style="cursor:pointer;padding:14px 18px;background:#eaf2fb;font-size:17px;font-weight:800;color:#17365d;display:flex;align-items:center;justify-content:space-between;list-style:none;">'
            f'<span>{html.escape(title)}</span><span class="growth-toggle-text">{toggle_text}</span></summary>'
            f'<div style="padding:14px 18px 18px;"><div style="color:#667085;font-size:13px;margin-bottom:12px;">{html.escape(description)}</div>'
            f'<div class="growth-period-buttons">{buttons}</div>{tables}</div></details>'
        )

    return (
        '<style>'
        '.growth-group summary::-webkit-details-marker{display:none}.growth-toggle-text{font-size:12px;color:#1f4e79;background:#fff;border:1px solid #b9c9dc;border-radius:7px;padding:6px 10px;min-width:44px;text-align:center}'
        '.growth-period-buttons{display:flex;flex-wrap:wrap;gap:7px;margin-bottom:18px}'
        '[data-growth-group="quarterly"]~*{}'
        '.growth-period-btn{flex:0 0 58px;border:1px solid #c7d3e3;background:#f8fafc;color:#29455f;border-radius:8px;padding:7px 5px;font-size:12px;font-weight:800;cursor:pointer;transition:.15s}'
        '.growth-period-btn:hover{background:#eaf2fb}.growth-period-btn.active{background:#1f4e79;color:#fff;border-color:#1f4e79;box-shadow:0 3px 9px rgba(31,78,121,.18)}'
        '.growth-tier-grid{display:grid;grid-template-columns:repeat(3,minmax(250px,1fr));gap:12px;max-width:1180px}'
        '.growth-tier-card{border:1px solid #d9e1ee;border-radius:12px;padding:13px 14px;background:#fbfcfe;box-shadow:0 2px 7px rgba(24,50,75,.05)}'
        '.growth-tier-card.low{border-top:4px solid #7ba7d1}.growth-tier-card.mid{border-top:4px solid #77b697}.growth-tier-card.high{border-top:4px solid #d8ad5d}'
        '.growth-tier-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:11px}'
        '.growth-tier-badge{font-weight:900;color:#17365d;font-size:15px}.growth-rate{display:flex;align-items:center;background:#edf4fb;color:#315b7d;border:1px solid #c8d9e9;border-radius:8px;padding:3px 8px}'
        '.growth-rate input{width:48px!important;border:0!important;background:transparent!important;color:#315b7d!important;text-align:right!important;font-weight:900;font-size:16px;padding:3px!important}.growth-rate b{font-size:13px;color:#5d7790}'
        '.growth-range{display:grid;gap:5px}.growth-range label{display:grid;grid-template-columns:34px 1fr 18px;align-items:center;gap:5px;color:#667085;font-size:12px}'
        '.growth-range input{width:100%;box-sizing:border-box;border:1px solid #d7e0ea;border-radius:7px;background:#fff;padding:7px 8px;text-align:right;font-weight:700;color:#25364a}'
        '.growth-range label b{font-weight:600;color:#667085}.growth-range-arrow{text-align:center;color:#9aa9b8;height:8px;line-height:8px;font-size:12px}'
        '@media(max-width:1100px){.growth-tier-grid{grid-template-columns:repeat(2,minmax(250px,1fr))}}'
        '@media(max-width:760px){.growth-period-btn{flex-basis:52px}.growth-tier-grid{grid-template-columns:1fr}}'
        '</style>'
        '<script>function showGrowthPeriod(group,period,button){document.querySelectorAll(\'[data-growth-group="\'+group+\'"]\').forEach(function(el){el.style.display=el.dataset.growthPeriod==String(period)?"block":"none"});document.querySelectorAll(\'[data-growth-button^="\'+group+\'-"]\').forEach(function(el){el.classList.remove("active")});button.classList.add("active")}function formatGrowthMoney(input){var digits=String(input.value||"").replace(/[^0-9]/g,"");input.value=digits?Number(digits).toLocaleString("ko-KR"):""}window.addEventListener("load",function(){document.querySelectorAll(".growth-money").forEach(function(input){input.addEventListener("input",function(){formatGrowthMoney(input)})});document.querySelectorAll(".growth-group").forEach(function(group){group.addEventListener("toggle",function(){var text=group.querySelector(".growth-toggle-text");if(text)text.textContent=group.open?"숨기기":"펼치기"})})});</script>'
        '<div style="padding:4px 18px 8px;">'
        + group_html("monthly", "표준 성장장려금 · 월별", "1월부터 12월까지 각 월별로 9등급을 관리합니다.")
        + group_html("quarterly", "타입B 성장장려금 · 분기별", "1분기부터 4분기까지 각 분기별로 9등급을 관리합니다.")
        + '</div><div style="padding:0 18px 18px;color:#667085;font-size:12px;">금액은 VAT 제외 기준입니다. 마지막 9등급의 종료 금액은 비워두세요.</div>'
    )


def save_growth_incentive_form(form: cgi.FieldStorage) -> None:
    current = load_config(GROWTH_INCENTIVE_PATH)
    groups = {"monthly": [], "quarterly": []}
    for key, count in (("monthly", 12), ("quarterly", 4)):
        for period_no in range(1, count + 1):
            bands = []
            for number in range(1, 10):
                prefix = f"{key}_{period_no}"
                lower_text = form[f"{prefix}_lower_{number}"].value.strip().replace(",", "")
                upper_text = form[f"{prefix}_upper_{number}"].value.strip().replace(",", "")
                rate_text = form[f"{prefix}_rate_{number}"].value.strip()
                bands.append({"band": number, "lower": int(lower_text), "upper": int(upper_text) if upper_text else None, "rate": float(rate_text) / 100})
            period = dict(current[key][period_no - 1])
            period["bands"] = bands
            groups[key].append(period)
    save_config(GROWTH_INCENTIVE_PATH, groups["monthly"], groups["quarterly"])


def sales_extra_amounts(amount: int) -> tuple[int, int]:
    vat_excluded = round(amount / 1.1)
    ad_budget = round(vat_excluded * 0.035)
    return vat_excluded, ad_budget


def load_tester_files_meta() -> list[dict[str, object]]:
    restore_file_from_supabase("tester_files_metadata", TESTER_FILES_META_PATH)
    if not TESTER_FILES_META_PATH.exists():
        return []
    try:
        value = json.loads(TESTER_FILES_META_PATH.read_text(encoding="utf-8"))
        return value if isinstance(value, list) else []
    except (OSError, json.JSONDecodeError):
        return []


def save_tester_files_meta(items: list[dict[str, object]]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    TESTER_FILES_META_PATH.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    backup_file_to_supabase("tester_files_metadata", TESTER_FILES_META_PATH)


def parse_tester_month(filename: str) -> str:
    match = re.search(r"(20\d{2})[^0-9]{0,2}(\d{1,2})월", filename)
    if not match:
        raise ValueError(f"파일명에서 연·월을 찾지 못했습니다: {filename}")
    year, month = int(match.group(1)), int(match.group(2))
    if not 1 <= month <= 12:
        raise ValueError(f"파일명의 월을 확인해주세요: {filename}")
    return f"{year}-{month:02d}"


def parse_tester_workbook(file_bytes: bytes, filename: str) -> tuple[str, int]:
    month = parse_tester_month(filename)
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for index, cell in enumerate(row):
                    if norm_header(cell.value) != "총진행금액":
                        continue
                    for next_cell in row[index + 1:index + 4]:
                        amount = parse_int(next_cell.value)
                        if amount > 0:
                            return month, amount
    finally:
        wb.close()
    raise ValueError(f"'{filename}'에서 총 진행 금액을 찾지 못했습니다.")


def save_uploaded_tester_files(items: list[tuple[str, bytes]]) -> tuple[int, list[str]]:
    metadata = load_tester_files_meta()
    affected_months: set[str] = set()
    TESTER_FILES_DIR.mkdir(parents=True, exist_ok=True)
    for filename, file_bytes in items:
        month, amount = parse_tester_workbook(file_bytes, filename)
        file_id = uuid.uuid5(uuid.NAMESPACE_URL, f"{month}:{filename}").hex[:16]
        stored_name = f"{file_id}_{safe_name(filename)}"
        stored_path = TESTER_FILES_DIR / stored_name
        stored_path.write_bytes(file_bytes)
        file_key = f"tester_file_{file_id}"
        backup_file_to_supabase(file_key, stored_path)
        metadata = [item for item in metadata if str(item.get("id")) != file_id]
        metadata.append({
            "id": file_id,
            "name": safe_name(filename),
            "stored_name": stored_name,
            "file_key": file_key,
            "month": month,
            "amount": amount,
            "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        })
        affected_months.add(month)
    metadata.sort(key=lambda item: (str(item.get("month", "")), str(item.get("name", ""))))
    save_tester_files_meta(metadata)
    manual = load_year_manual()
    for month in affected_months:
        manual.setdefault(month, {})["tester"] = str(sum(int(item.get("amount", 0)) for item in metadata if item.get("month") == month))
    save_year_manual_data(manual)
    return len(items), sorted(affected_months)


def render_tester_files() -> str:
    items = load_tester_files_meta()
    if not items:
        return '<div style="color:#667085;font-size:13px;">저장된 체험단 원본 파일이 없습니다.</div>'
    by_month: dict[str, list[dict[str, object]]] = {}
    for item in items:
        by_month.setdefault(str(item.get("month", "")), []).append(item)
    folders = []
    for month in sorted(by_month, reverse=True):
        month_items = sorted(by_month[month], key=lambda value: str(value.get("name", "")))
        month_total = sum(int(item.get("amount", 0)) for item in month_items)
        rows = []
        for item in month_items:
            file_id = urllib.parse.quote(str(item.get("id", "")))
            rows.append(
                '<div style="display:grid;grid-template-columns:minmax(0,1fr) 150px 90px;gap:10px;align-items:center;padding:9px 10px;border-top:1px solid #e5eaf1;font-size:13px;">'
                f'<span>{html.escape(str(item.get("name", "")))}</span>'
                f'<span style="text-align:right;font-weight:800;">{int(item.get("amount", 0)):,}원</span>'
                f'<a class="btn secondary" style="padding:7px 9px;text-align:center;" href="/sales/tester/download?id={file_id}">원본 받기</a></div>'
            )
        folders.append(
            f'<details data-month="{html.escape(month, quote=True)}" style="border:1px solid #dbe4ef;border-radius:8px;background:#fff;overflow:hidden;">'
            '<summary style="display:flex;justify-content:space-between;align-items:center;gap:12px;padding:11px 13px;background:#eef5fb;color:#173f68;cursor:pointer;font-size:14px;font-weight:900;">'
            f'<span>{html.escape(month)} 체험단</span><span>{month_total:,}원 · {len(month_items)}개</span></summary>'
            + "".join(rows) + "</details>"
        )
    return '<div style="display:grid;gap:8px;">' + "".join(folders) + "</div>"


def load_year_manual() -> dict:
    restore_file_from_supabase("year_manual", YEAR_MANUAL_PATH)
    if not YEAR_MANUAL_PATH.exists():
        return {}
    try:
        return json.loads(YEAR_MANUAL_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def save_year_manual_data(manual: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    YEAR_MANUAL_PATH.write_text(json.dumps(manual, ensure_ascii=False, indent=2), encoding="utf-8")
    backup_file_to_supabase("year_manual", YEAR_MANUAL_PATH)


def save_year_manual(form: cgi.FieldStorage) -> int:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    manual = load_year_manual()
    changed = 0
    for month in range(1, 13):
        key = f"{datetime.now().year}-{month:02d}"
        if key not in manual:
            manual[key] = {}
        for field in ("tester", "ad", "discount", "partner_discount", "extra_ad"):
            form_key = f"{field}_{month}"
            new_value = form[form_key].value.strip() if form_key in form else ""
            old_value = str(manual[key].get(field, ""))
            if new_value != old_value:
                manual[key][field] = new_value
                changed += 1
    save_year_manual_data(manual)
    return changed


def render_year_rows(summary_rows: list[tuple[str, int, int, int]]) -> str:
    manual = load_year_manual()
    year = datetime.now().year
    monthly_sales = defaultdict(int)
    for day, _qty, amount, _po_count in summary_rows:
        month_key = str(day)[:7]
        monthly_sales[month_key] += amount

    vat_excluded_amounts = [sales_extra_amounts(monthly_sales[f"{year}-{month:02d}"])[0] for month in range(1, 13)]
    incentive_results = calculate_year(vat_excluded_amounts, load_config(GROWTH_INCENTIVE_PATH))

    rows = []
    totals = defaultdict(int)
    for month in range(1, 13):
        key = f"{year}-{month:02d}"
        amount = monthly_sales[key]
        vat_excluded, ad_budget = sales_extra_amounts(amount)
        values = manual.get(key, {})
        tester = parse_int(values.get("tester", ""))
        ad = parse_int(values.get("ad", ""))
        incentive = incentive_results[month - 1]
        support = int(incentive["total"])
        if int(incentive["band"]) == 0:
            support_detail = "기준 미달"
        else:
            support_detail = f'{int(incentive["band"])}등급 · {float(incentive["rate"]) * 100:g}%'
        if int(incentive["quarter_extra"]) > 0:
            support_detail += f' · 분기 추가 {int(incentive["quarter_extra"]):,}원'
        discount = parse_int(values.get("discount", ""))
        partner_discount = parse_int(values.get("partner_discount", ""))
        extra_ad = parse_int(values.get("extra_ad", ""))
        over = ad_budget - tester - ad
        net = vat_excluded - tester - ad - support - discount - partner_discount - extra_ad
        totals["sales"] += vat_excluded
        totals["budget"] += ad_budget
        totals["tester"] += tester
        totals["ad"] += ad
        totals["over"] += over
        totals["support"] += support
        totals["discount"] += discount
        totals["partner_discount"] += partner_discount
        totals["extra_ad"] += extra_ad
        totals["net"] += net
        rows.append(
            f'<tr class="year-row" data-month="{month:02d}" data-sales="{vat_excluded}" data-budget="{ad_budget}" data-support="{support}">'
            f"<td>{month}월</td>"
            f'<td class="auto-cell year-sales">{vat_excluded:,}원</td>'
            f'<td class="auto-cell year-budget">{ad_budget:,}원</td>'
            f'<td><input class="money-input tester-input" name="tester_{month}" value="{html.escape(str(values.get("tester", "")), quote=True)}"></td>'
            f'<td><input class="money-input ad-input" name="ad_{month}" value="{html.escape(str(values.get("ad", "")), quote=True)}"></td>'
            f'<td class="over-cell{" negative" if over < 0 else ""}">{over:,}원</td>'
            f'<td class="auto-cell support-display" title="월 기본 {int(incentive["base"]):,}원 / 분기 추가 {int(incentive["quarter_extra"]):,}원"><strong>{support:,}원</strong><small style="display:block;margin-top:4px;color:#667085;font-size:11px;white-space:nowrap;">{html.escape(support_detail)}</small></td>'
            f'<td><input class="money-input discount-input" name="discount_{month}" value="{html.escape(str(values.get("discount", "")), quote=True)}"><span class="discount-display" style="display:none">{discount:,}원</span></td>'
            f'<td><input class="money-input partner-discount-input" name="partner_discount_{month}" value="{html.escape(str(values.get("partner_discount", "")), quote=True)}"><span class="partner-discount-display" style="display:none">{partner_discount:,}원</span></td>'
            f'<td><input class="money-input extra-ad-input" name="extra_ad_{month}" value="{html.escape(str(values.get("extra_ad", "")), quote=True)}"><span class="extra-ad-display" style="display:none">{extra_ad:,}원</span></td>'
            f'<td class="auto-cell net-cell{" negative" if net < 0 else ""}" style="font-weight:900;">{net:,}원</td>'
            f"</tr>"
        )
    rows.append(
        '<tr class="year-total-row" style="background:#fff2cc;font-weight:800;">'
        '<td>연도합계</td>'
        f'<td class="year-total-sales">{totals["sales"]:,}원</td>'
        f'<td class="year-total-budget">{totals["budget"]:,}원</td>'
        f'<td class="year-total-tester">{totals["tester"]:,}원</td>'
        f'<td class="year-total-ad">{totals["ad"]:,}원</td>'
        f'<td class="year-total-over">{totals["over"]:,}원</td>'
        f'<td class="year-total-support">{totals["support"]:,}원</td>'
        f'<td class="year-total-discount">{totals["discount"]:,}원</td>'
        f'<td class="year-total-partner-discount">{totals["partner_discount"]:,}원</td>'
        f'<td class="year-total-extra-ad">{totals["extra_ad"]:,}원</td>'
        f'<td class="year-total-net">{totals["net"]:,}원</td>'
        '</tr>'
    )
    return "\n".join(rows)


def write_sales_display_workbook() -> Path:
    summary_rows, detail_rows = load_monthly_sales_summary(limit_rows=None, aggregate_by_sku=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "매출확인용"

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="E7E6E6")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    row_no = 1
    if summary_rows:
        ws.cell(row_no, 1, "일자별 납품 상품 합계")
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=7)
        ws.cell(row_no, 1).font = Font(bold=True, size=14)
        ws.cell(row_no, 1).fill = title_fill
        ws.cell(row_no, 1).alignment = center
        row_no += 2

        by_day = defaultdict(list)
        for _row_no, day, sku, name, _original_qty, qty, unit_price, amount, remarks, memo, _changed in detail_rows:
            by_day[day].append((sku, name, qty, unit_price, amount, remarks, memo))

        for day in sorted(by_day.keys()):
            day_qty = sum(row[2] for row in by_day[day])
            day_amount = sum(row[4] for row in by_day[day])
            ws.cell(row_no, 1, f"{day} 합계: 수량 {day_qty:,}개 / 금액 {day_amount:,}원")
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=7)
            ws.cell(row_no, 1).fill = title_fill
            ws.cell(row_no, 1).font = bold
            ws.cell(row_no, 1).alignment = center
            row_no += 1

            for col, header in enumerate(["SKU ID", "상품명", "납품수량", "단가", "금액", "비고", "메모"], start=1):
                cell = ws.cell(row_no, col, header)
                cell.fill = header_fill
                cell.font = bold
                cell.alignment = center
            row_no += 1

            for sku, name, qty, unit_price, amount, remarks, memo in sorted(by_day[day], key=lambda item: str(item[0])):
                ws.cell(row_no, 1, sku)
                ws.cell(row_no, 2, name)
                ws.cell(row_no, 3, qty)
                ws.cell(row_no, 4, unit_price)
                ws.cell(row_no, 5, amount)
                ws.cell(row_no, 6, remarks)
                ws.cell(row_no, 7, memo)
                ws.cell(row_no, 3).number_format = '#,##0'
                ws.cell(row_no, 4).number_format = '#,##0'
                ws.cell(row_no, 5).number_format = '#,##0'
                row_no += 1

            ws.cell(row_no, 1, "합계")
            ws.cell(row_no, 3, day_qty)
            ws.cell(row_no, 5, day_amount)
            for col in range(1, 8):
                ws.cell(row_no, col).fill = total_fill
                ws.cell(row_no, col).font = bold
            ws.cell(row_no, 3).number_format = '#,##0'
            ws.cell(row_no, 5).number_format = '#,##0'
            row_no += 2
    else:
        ws.append(["아직 누적된 월매출 자료가 없습니다."])

    widths = {"A": 14, "B": 60, "C": 12, "D": 12, "E": 14, "F": 28, "G": 32}
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width
    for cells in ws.iter_rows():
        for cell in cells:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A3"
    wb.save(MONTHLY_SALES_PATH)
    return MONTHLY_SALES_PATH


def render_confirmation_history(record: dict[str, object]) -> str:
    history = record.get("history", [])
    if not isinstance(history, list) or not history:
        return '<div class="history-empty">변경 이력이 없습니다.</div>'
    rows = []
    for item in reversed(history):
        if not isinstance(item, dict):
            continue
        changes = item.get("changes", [])
        if isinstance(changes, list):
            change_text = "<br>".join(
                f'{html.escape(str(change.get("field", "")))}: '
                f'{html.escape(str(change.get("before", "")))} → {html.escape(str(change.get("after", "")))}'
                for change in changes if isinstance(change, dict)
            )
        else:
            change_text = html.escape(str(changes or ""))
        rows.append(
            "<tr>"
            f'<td>{html.escape(str(item.get("at", "")))}</td>'
            f'<td>{html.escape(str(item.get("user", "")))}</td>'
            f'<td>{html.escape(str(item.get("action", "")))}</td>'
            f'<td>{change_text or "-"}</td>'
            f'<td>{html.escape(str(item.get("reason", "")))}</td>'
            "</tr>"
        )
    return (
        '<div class="history-scroll"><table class="history-table"><thead><tr>'
        '<th>일시</th><th>작업자</th><th>구분</th><th>변경 내용</th><th>사유</th>'
        f'</tr></thead><tbody>{"".join(rows)}</tbody></table></div>'
    )


def render_confirmation_panel(month: str, sales_amount: int) -> str:
    record = sales_confirmation(month)
    invoice_amount = parse_int(record.get("invoice_amount", 0))
    confirmed = bool(record.get("confirmed"))
    needs_recheck = bool(record.get("needs_recheck"))
    complete_after_edit = bool(record.get("edited_after_confirm"))
    if not invoice_amount:
        match_label, match_class = "확인 필요", "status-warn"
    elif invoice_amount == sales_amount:
        match_label, match_class = "금액 매칭 완료", "status-ok"
    else:
        difference = invoice_amount - sales_amount
        match_label, match_class = f"금액 불일치 (차액 {difference:+,}원)", "status-bad"
    state_labels = []
    if confirmed:
        state_labels.append('<span class="status-badge status-ok">계산서 발행 확인 완료</span>')
    if complete_after_edit:
        state_labels.append('<span class="status-badge status-bad">확인 완료 후 수정됨</span>')
    if needs_recheck:
        state_labels.append('<span class="status-badge status-warn">재확인 필요</span>')
    checked = " checked" if confirmed else ""
    confirm_info = ""
    if confirmed:
        confirm_info = (
            '<div class="confirm-meta">'
            f'확인자: {html.escape(str(record.get("confirmed_by", "")))} · '
            f'확인일시: {html.escape(str(record.get("confirmed_at", "")))} · '
            f'확인 당시 금액: {parse_int(record.get("confirmed_invoice_amount", invoice_amount)):,}원 / '
            f'{parse_int(record.get("confirmed_sales_amount", sales_amount)):,}원</div>'
        )
    recheck_button = ""
    if confirmed and needs_recheck and invoice_amount == sales_amount:
        recheck_button = '<button class="confirm-action" type="submit" name="action" value="reconfirm">재확인 완료</button>'
    return (
        f'<div class="invoice-check" data-confirmed="{str(confirmed).lower()}" data-needs-recheck="{str(needs_recheck).lower()}">'
        f'<form method="post" action="/sales/confirm" class="confirm-form">'
        f'<input type="hidden" name="month" value="{html.escape(month, quote=True)}">'
        f'<input type="hidden" name="sales_amount" value="{sales_amount}">'
        f'<strong>{html.escape(month)} 계산서 확인</strong>'
        f'<label>매출확인용 금액<input class="confirm-money" value="{sales_amount:,}" readonly></label>'
        f'<label>세금계산서 발행 금액<input class="confirm-money" name="invoice_amount" inputmode="numeric" value="{invoice_amount or ""}" placeholder="금액 입력"></label>'
        f'<span class="status-badge {match_class}">{match_label}</span>'
        f'<label class="confirm-check"><input type="checkbox" name="confirmed" value="1"{checked} onchange="if(!this.checked&&!confirm(\'계산서 발행 확인 완료 상태를 해제하시겠습니까?\'))this.checked=true;"> 계산서 발행 확인 완료</label>'
        f'<button class="confirm-action" type="submit" name="action" value="save">저장</button>{recheck_button}'
        f'</form><div class="confirm-states">{"".join(state_labels)}</div>{confirm_info}'
        f'<details class="history-details"><summary>변경 이력 보기</summary>{render_confirmation_history(record)}</details>'
        '</div>'
    )


def render_sales_page(message: str = "", folder_mode: bool = False) -> str:
    if folder_mode:
        summary_rows, detail_rows = load_monthly_sales_summary(limit_rows=None, aggregate_by_sku=True)
        _po_summary_rows, po_detail_rows = load_monthly_sales_summary(limit_rows=None, aggregate_by_sku=False)
    else:
        summary_rows, detail_rows = load_monthly_sales_summary(aggregate_by_sku=True)
        po_detail_rows = []
    current_month = datetime.now().strftime("%Y-%m")
    visible_detail_rows = detail_rows if folder_mode else [
        row for row in detail_rows if str(row[1]).startswith(current_month)
    ]
    visible_po_detail_rows = po_detail_rows if folder_mode else []
    if summary_rows:
        def sales_extra_amounts(amount: int) -> tuple[int, int]:
            vat_excluded = round(amount / 1.1)
            ad_budget = round(vat_excluded * 0.035)
            return vat_excluded, ad_budget

        monthly_summary: dict[str, list[tuple[str, int, int, int]]] = defaultdict(list)
        for day, qty, amount, po_count in summary_rows:
            monthly_summary[str(day)[:7]].append((str(day), qty, amount, po_count))
        latest_summary_month = max(monthly_summary.keys())
        summary_parts = []
        for month in sorted(monthly_summary.keys(), reverse=True):
            rows_for_month = sorted(monthly_summary[month], key=lambda row: row[0])
            month_qty = sum(qty for _day, qty, _amount, _po_count in rows_for_month)
            month_amount = sum(amount for _day, _qty, amount, _po_count in rows_for_month)
            month_po_count = sum(po_count for _day, _qty, _amount, po_count in rows_for_month)
            month_vat, month_budget = sales_extra_amounts(month_amount)
            confirmation_panel = render_confirmation_panel(month, month_amount)
            year_text, month_text = month.split("-", 1) if "-" in month else ("", month)
            collapsed_class = "" if month == latest_summary_month else " is-collapsed"
            toggle_label = "숨기기" if month == latest_summary_month else "펼치기"
            row_html = "\n".join(
                f'<tr class="summary-row" data-day="{html.escape(day)}" data-po-count="{po_count}">'
                f"<td>{html.escape(day)}</td><td>{po_count:,}</td>"
                f'<td class="summary-qty">{qty:,}</td><td class="summary-amount">{amount:,}원</td>'
                f'<td class="summary-vat">{sales_extra_amounts(amount)[0]:,}원</td><td class="summary-budget">{sales_extra_amounts(amount)[1]:,}원</td></tr>'
                for day, qty, amount, po_count in rows_for_month
            )
            summary_parts.append(
                f'<section class="summary-month-section{collapsed_class}" data-month="{html.escape(month, quote=True)}">'
                f'<div class="summary-month-head">'
                f'<div class="summary-month-title">{html.escape(year_text)}년 {html.escape(month_text)}월 매출</div>'
                f'<button class="summary-month-toggle" type="button" onclick="toggleSummaryMonth(this)">{toggle_label}</button></div>'
                f'<div class="summary-month-body"><table class="summary-table resizable-table">'
                f'<colgroup><col style="width:15%;"><col style="width:10%;"><col style="width:13%;"><col style="width:22%;"><col style="width:20%;"><col style="width:20%;"></colgroup>'
                f'<thead><tr><th>일자</th><th>PO 수</th><th>납품수량</th><th>납품상품 합계금액</th><th>VAT 별도</th><th>광고비예산</th></tr></thead>'
                f'<tbody>{row_html}'
                f'<tr class="summary-month-total"><td>월 합계</td><td class="summary-month-po">{month_po_count:,}</td><td class="summary-month-qty">{month_qty:,}</td><td class="summary-month-amount">{month_amount:,}원</td><td class="summary-month-vat">{month_vat:,}원</td><td class="summary-month-budget">{month_budget:,}원</td></tr>'
                f'</tbody></table>{confirmation_panel}</div></section>'
            )
        summary_html = "\n".join(summary_parts)
    else:
        summary_html = '<div class="summary-month-section"><div class="summary-month-body"><table class="summary-table"><tbody><tr><td colspan="6">아직 누적된 월매출 자료가 없습니다.</td></tr></tbody></table></div></div>'

    if visible_detail_rows:
        detail_sources = [("sku", visible_detail_rows)]
        if folder_mode:
            po_rows = []
            for row in visible_po_detail_rows:
                if str(row[1])[:7] < "2026-07":
                    continue
                row = list(row)
                row[8] = str(row[8]).replace("PO:", "").strip()
                po_rows.append(row)
            detail_sources.append(("po", po_rows))
        parts = []
        for view_mode, source_rows in detail_sources:
            by_day = defaultdict(list)
            for row_no, day, sku, name, original_qty, qty, unit_price, amount, remarks, memo, changed in source_rows:
                by_day[day].append((row_no, sku, name, original_qty, qty, unit_price, amount, remarks, memo, changed))
            by_month = defaultdict(list)
            for day in sorted(by_day.keys(), reverse=True):
                by_month[str(day)[:7]].append(day)
            for month in sorted(by_month.keys(), reverse=True):
                hidden_class = " month-hidden" if folder_mode else ""
                mode_hidden = " view-hidden" if view_mode != "sku" else ""
                mode_label = "SKU 합계" if view_mode == "sku" else "PO별 상세"
                if folder_mode:
                    parts.append(
                        f'<tr class="month-folder is-closed{mode_hidden}" data-view-mode="{view_mode}" data-month="{html.escape(month, quote=True)}">'
                        f'<th colspan="8">{html.escape(month)} {mode_label} '
                        f'<button class="toggle-btn" type="button" onclick="toggleMonthFolder(\'{html.escape(month, quote=True)}\', this)">펼치기</button> '
                        f'{"" if view_mode == "sku" else f"<button class=\"delete-btn\" type=\"submit\" formaction=\"/sales/delete\" name=\"delete_month\" value=\"{html.escape(month, quote=True)}\" onclick=\"return confirmDeleteMonth(\'{html.escape(month, quote=True)}\')\">월 삭제</button>"}</th></tr>'
                    )
                for day in by_month[month]:
                    day_qty = sum(row[4] for row in by_day[day])
                    day_amount = sum(row[6] for row in by_day[day])
                    parts.append(
                        f'<tr class="detail-day-row{hidden_class}{mode_hidden}" data-view-mode="{view_mode}" data-month="{html.escape(month, quote=True)}" data-day="{html.escape(str(day), quote=True)}"><th colspan="8" data-day-total="{html.escape(str(day), quote=True)}" data-view-mode="{view_mode}" style="background:#e8f1fb;color:#1f4e79;">'
                        f'{html.escape(str(day))} 합계: 수량 {day_qty:,}개 / 금액 {day_amount:,}원 '
                        f'{"" if view_mode == "sku" else f"<button class=\"delete-btn\" type=\"submit\" formaction=\"/sales/delete\" name=\"delete_day\" value=\"{html.escape(str(day), quote=True)}\" onclick=\"return confirmDeleteDay(\'{html.escape(str(day), quote=True)}\')\">일자 삭제</button>"}</th></tr>'
                    )
                    for row_no, sku, name, original_qty, qty, unit_price, amount, remarks, memo, changed in sorted(by_day[day], key=lambda row: str(row[1])):
                        changed_class = " changed" if changed else ""
                        month_record = sales_confirmation(str(day)[:7])
                        month_confirmed = str(bool(month_record.get("confirmed"))).lower()
                        if folder_mode and view_mode == "po":
                            parts.append(
                                f'<tr class="detail-row {changed_class.strip()}{hidden_class}{mode_hidden}" data-view-mode="po" data-day="{html.escape(str(day), quote=True)}" data-month="{html.escape(str(day)[:7], quote=True)}" data-confirmed="{month_confirmed}" data-original-qty="{original_qty}" data-saved-qty="{qty}" data-saved-memo="{html.escape(str(memo), quote=True)}" data-unit-price="{unit_price}" data-po-list="{html.escape(str(remarks), quote=True)}" data-search="{html.escape((str(sku) + " " + str(name) + " " + str(remarks) + " " + str(memo)), quote=True)}"><td>{html.escape(str(sku))}'
                                f'<input type="hidden" name="row" value="{row_no}"></td>'
                                f"<td>{html.escape(str(name))}</td>"
                                f'<td class="{changed_class.strip()}"><input class="qty-input" type="number" min="0" step="1" '
                                f'name="qty_{row_no}" value="{qty}"></td>'
                                f'<td>{unit_price:,}원</td><td class="row-amount">{amount:,}원</td><td class="po-cell" data-original-po="{html.escape(str(remarks), quote=True)}">{html.escape(str(remarks))}</td>'
                                f'<td><input class="memo-input" type="text" name="memo_{row_no}" value="{html.escape(str(memo), quote=True)}" '
                                f'placeholder="수정 사유"></td>'
                                f'<td><button class="delete-btn" type="submit" formaction="/sales/delete" name="delete_row" value="{row_no}" onclick="return confirm(\'이 상품 줄을 삭제할까요?\')">삭제</button></td></tr>'
                            )
                        else:
                            parts.append(
                                f'<tr class="detail-row {changed_class.strip()}{hidden_class}{mode_hidden}" data-view-mode="sku" data-day="{html.escape(str(day), quote=True)}" data-month="{html.escape(str(day)[:7], quote=True)}" data-original-qty="{original_qty}" data-unit-price="{unit_price}" data-search="{html.escape((str(sku) + " " + str(name) + " " + str(remarks) + " " + str(memo)), quote=True)}">'
                                f"<td>{html.escape(str(sku))}</td><td>{html.escape(str(name))}</td>"
                                f'<td class="{changed_class.strip()}">{qty:,}</td>'
                                f'<td>{unit_price:,}원</td><td class="row-amount">{amount:,}원</td><td class="po-cell" data-original-po="{html.escape(str(remarks), quote=True)}">{html.escape(str(remarks))}</td>'
                                f"<td>{html.escape(str(memo))}</td><td>조회용</td></tr>"
                            )
        detail_html = "\n".join(parts)
    else:
        detail_html = '<tr><td colspan="8">해당 월 납품 상품 내역이 없습니다.</td></tr>'
    year_rows = render_year_rows(summary_rows)
    page_title = "월별납품관리" if folder_mode else "매출확인용"
    page_sub = (
        "누적된 납품 상품 내역을 월별 폴더로 펼쳐서 확인합니다."
        if folder_mode
        else "업로드, 연도총매출, 일자별 합계, 해당월 납품 상품을 확인합니다."
    )
    detail_title = "월별 납품 상품 폴더" if folder_mode else f"{current_month} 납품 상품 내역"
    return (
        SALES_PAGE
        .replace("{message}", message)
        .replace("{year_rows}", year_rows)
        .replace("{tester_files}", render_tester_files())
        .replace("{summary_rows}", summary_html)
        .replace("{detail_rows}", detail_html)
        .replace("{page_title}", page_title)
        .replace("{page_sub}", page_sub)
        .replace("{detail_title}", detail_title)
        .replace("{current_month}", current_month)
        .replace("{sales_active}", "active" if not folder_mode else "")
        .replace("{folders_active}", "active" if folder_mode else "")
        .replace("{folder_filter_class}", "" if folder_mode else "not-shown")
        .replace("{year_section_class}", "not-shown" if folder_mode else "")
        .replace("{upload_section_class}", "" if folder_mode else "not-shown")
        .replace("{summary_section_class}", "not-shown" if folder_mode else "")
    )


def normalize_simple_no(value: object) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\D", "", text)


def extract_simple_no(values: list[object]) -> str:
    for value in values:
        text = str(value or "").strip()
        match = re.search(r"(\d{3,6})\s*[:：]", text)
        if match:
            return match.group(1)
    for value in values:
        no = normalize_simple_no(value)
        if 3 <= len(no) <= 6:
            return no
    return ""


def get_master_simpleworks_maps() -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    master_path = get_saved_master_path()
    if master_path is None:
        raise ValueError("저장된 기초자료가 없습니다. 먼저 기초자료에 심플웍스 No를 입력해주세요.")
    wb = load_workbook(master_path)
    ws = wb.active
    cols = find_master_columns(ws)
    wb.save(master_path)
    sku_to_simple: dict[str, str] = {}
    simple_to_info: dict[str, dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row, cols["sku"]).value or "").strip()
        simple_no = normalize_simple_no(ws.cell(row, cols["simple_no"]).value)
        name = str(ws.cell(row, cols["name"]).value or "").strip()
        if not sku or not simple_no:
            continue
        sku_to_simple[sku] = simple_no
        simple_to_info.setdefault(simple_no, {"sku": sku, "name": name})
    return sku_to_simple, simple_to_info


def parse_simpleworks_excel(path: Path) -> dict[str, int]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = None
    product_col = None
    qty_col = None
    for row in range(1, min(ws.max_row, 30) + 1):
        for col in range(1, ws.max_column + 1):
            text = norm_header(ws.cell(row, col).value)
            if text in ["상품명", "상품"]:
                header_row = row
                product_col = col
            if text == "수량":
                header_row = row if header_row is None else header_row
                qty_col = col
        if product_col and qty_col:
            break

    quantities: dict[str, int] = defaultdict(int)
    start_row = (header_row + 1) if header_row else 1
    for row in range(start_row, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if all(value in [None, ""] for value in values):
            continue
        simple_no = extract_simple_no(values)
        if not simple_no:
            continue
        if qty_col:
            qty = parse_int(ws.cell(row, qty_col).value)
        else:
            qty = 0
            for value in reversed(values):
                text = str(value or "").strip()
                if re.search(r"\d+\s*개?$", text) or isinstance(value, (int, float)):
                    qty = parse_int(text)
                    break
        if qty > 0:
            quantities[simple_no] += qty
    return dict(quantities)


def extract_simpleworks_quantities_from_text(text: str) -> dict[str, int]:
    quantities: dict[str, int] = defaultdict(int)
    pending_simple_no = ""
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line:
            continue
        simple_match = re.search(r"([0-9$SIl|Tt]{3,7})\s*[:：]", line)
        if simple_match:
            pending_simple_no = (
                simple_match.group(1)
                .replace("$", "9")
                .replace("S", "5")
                .replace("I", "1")
                .replace("l", "1")
                .replace("|", "1")
                .replace("T", "7")
                .replace("t", "7")
            )
            pending_simple_no = re.sub(r"\D", "", pending_simple_no)
        elif not pending_simple_no:
            start_match = re.match(r"^\D*(\d{3,6})\b", line)
            if start_match:
                pending_simple_no = start_match.group(1)
        qty_matches = re.findall(r"(\d{1,4})\s*(?:개|m|M)", line)
        if not qty_matches and pending_simple_no:
            trailing_numbers = re.findall(r"\b(\d+)\b", line)
            trailing_numbers = [number for number in trailing_numbers if number != pending_simple_no]
            if trailing_numbers:
                qty_matches = [trailing_numbers[-1]]
        if pending_simple_no and qty_matches:
            quantities[pending_simple_no] += parse_int(qty_matches[-1])
            pending_simple_no = ""
    return dict(quantities)


def merge_quantity_maps(*quantity_maps: dict[str, int]) -> dict[str, int]:
    merged: dict[str, int] = {}
    for quantity_map in quantity_maps:
        for simple_no, qty in quantity_map.items():
            if qty <= 0:
                continue
            if simple_no not in merged:
                merged[simple_no] = qty
    return merged


def filter_known_simpleworks_quantities(quantities: dict[str, int]) -> dict[str, int]:
    try:
        _sku_to_simple, simple_to_info = get_master_simpleworks_maps()
    except Exception:
        return quantities
    known_simple_nos = set(simple_to_info)
    if not known_simple_nos:
        return quantities
    filtered = {simple_no: qty for simple_no, qty in quantities.items() if simple_no in known_simple_nos}
    return filtered or quantities


def normalize_ocr_qty_against_expected(actual: int, expected: int) -> int:
    if expected < 10 or actual <= expected:
        return actual
    actual_text = str(actual)
    expected_text = str(expected)
    if actual >= expected * 10 and actual_text.startswith(expected_text):
        return expected
    return actual


def get_tesseract_candidates() -> list[str]:
    return [
        os.environ.get("TESSERACT_CMD", ""),
        shutil.which("tesseract") or "",
        "/usr/bin/tesseract",
        "/usr/local/bin/tesseract",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    ]


def get_tesseract_path() -> str | None:
    tesseract_path = shutil.which("tesseract")
    if tesseract_path:
        return tesseract_path
    for candidate in get_tesseract_candidates():
        if candidate and Path(candidate).exists():
            return candidate
    return None


def get_tesseract_language(tesseract_path: str) -> str:
    completed = subprocess.run(
        [tesseract_path, "--list-langs"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="ignore",
        timeout=20,
    )
    langs = set(re.findall(r"^[a-zA-Z_]+$", completed.stdout, flags=re.MULTILINE))
    if "kor" in langs and "eng" in langs:
        return "kor+eng"
    if "kor" in langs:
        return "kor"
    return "eng"


def clean_ocr_simple_no(text: str) -> str:
    cleaned = (
        str(text or "")
        .replace("$", "9")
        .replace("S", "5")
        .replace("I", "1")
        .replace("l", "1")
        .replace("|", "1")
        .replace("T", "7")
        .replace("t", "7")
    )
    digits = re.sub(r"\D", "", cleaned)
    if len(digits) == 5 and digits.endswith("2"):
        return digits[:4]
    if len(digits) > 4:
        return digits[:4]
    return digits


def clean_ocr_qty(text: str) -> int:
    digits = re.sub(r"\D", "", str(text or ""))
    if len(digits) >= 3 and int(digits[-2:]) >= 50:
        digits = digits[:-2]
    return parse_int(digits)


def run_tesseract_text(tesseract_path: str, image_path: Path, mode: str = "6", whitelist: str = "") -> str:
    command = [tesseract_path, str(image_path), "stdout", "-l", get_tesseract_language(tesseract_path), "--psm", mode]
    if whitelist:
        command.extend(["-c", f"tessedit_char_whitelist={whitelist}"])
    completed = subprocess.run(command, capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=60)
    if completed.returncode != 0:
        raise ValueError(f"캡쳐 이미지 글자읽기에 실패했습니다: {completed.stderr.strip() or 'OCR 오류'}")
    return completed.stdout


def parse_simpleworks_table_image(path: Path, tesseract_path: str) -> dict[str, int]:
    try:
        from PIL import Image, ImageEnhance
    except Exception:
        return {}

    image = Image.open(path).convert("L")
    width, height = image.size
    pixels = image.load()

    horizontal_lines: list[int] = []
    for y in range(height):
        dark_count = sum(1 for x in range(width) if pixels[x, y] < 235)
        if dark_count > width * 0.45:
            if not horizontal_lines or y - horizontal_lines[-1] > 4:
                horizontal_lines.append(y)

    if len(horizontal_lines) < 3:
        return {}

    vertical_candidates: list[tuple[int, int]] = []
    start_y = horizontal_lines[1]
    for x in range(int(width * 0.55), width):
        dark_count = sum(1 for y in range(start_y, height) if pixels[x, y] < 235)
        if dark_count > (height - start_y) * 0.25:
            vertical_candidates.append((x, dark_count))
    divider_x = min(vertical_candidates, key=lambda item: item[0])[0] if vertical_candidates else int(width * 0.82)

    quantities: dict[str, int] = defaultdict(int)
    temp_dir = RUNS_DIR / "ocr_crops"
    temp_dir.mkdir(parents=True, exist_ok=True)
    row_pairs = list(zip(horizontal_lines[1:-1], horizontal_lines[2:]))
    for index, (y1, y2) in enumerate(row_pairs, start=1):
        if y2 - y1 < 18:
            continue
        top = y1 + 4
        bottom = y2 - 4
        if bottom <= top:
            continue

        no_crop = image.crop((8, top, min(120, max(80, divider_x // 4)), bottom))
        no_crop = no_crop.resize((no_crop.width * 6, no_crop.height * 6))
        no_crop = ImageEnhance.Contrast(no_crop).enhance(2.2)
        no_path = temp_dir / f"{path.stem}_no_{index}.png"
        no_crop.save(no_path)
        no_text = run_tesseract_text(tesseract_path, no_path, "7", "0123456789$SIl|Tt")
        simple_no = clean_ocr_simple_no(no_text)
        if not simple_no:
            row_crop = image.crop((0, top, divider_x, bottom))
            row_crop = row_crop.resize((row_crop.width * 3, row_crop.height * 3))
            row_crop = ImageEnhance.Contrast(row_crop).enhance(2.0)
            row_path = temp_dir / f"{path.stem}_row_{index}.png"
            row_crop.save(row_path)
            row_text = run_tesseract_text(tesseract_path, row_path, "6")
            simple_no = extract_simple_no([row_text])

        qty_crop = image.crop((divider_x + 4, top, min(width, divider_x + 56), bottom))
        qty_crop = qty_crop.resize((qty_crop.width * 6, qty_crop.height * 6))
        qty_crop = ImageEnhance.Contrast(qty_crop).enhance(2.2)
        qty_path = temp_dir / f"{path.stem}_qty_{index}.png"
        qty_crop.save(qty_path)
        qty_text = run_tesseract_text(tesseract_path, qty_path, "7", "0123456789")
        qty = clean_ocr_qty(qty_text)

        if simple_no and qty > 0:
            quantities[simple_no] += qty

    return dict(quantities)


def parse_simpleworks_image(path: Path) -> dict[str, int]:
    tesseract_path = get_tesseract_path()
    if not tesseract_path:
        checked = ", ".join(candidate for candidate in get_tesseract_candidates() if candidate) or "없음"
        raise ValueError(
            "캡쳐 이미지를 읽으려면 OCR 프로그램인 Tesseract 설치가 필요합니다. "
            f"지금은 심플웍스 엑셀 업로드를 사용해주세요. 확인한 경로: {checked} / PATH: {os.environ.get('PATH', '')}"
        )
    table_quantities = parse_simpleworks_table_image(path, tesseract_path)
    full_text_quantities: dict[str, int] = {}
    try:
        full_text = run_tesseract_text(tesseract_path, path, "6")
        lower_full_text = full_text.lower()
        if any(keyword in full_text for keyword in ["검수 결과", "심플웍스만 있음", "쿠팡만 있음", "수량차이"]) or "sku id" in lower_full_text:
            raise ValueError("검수 결과 화면 캡쳐가 아니라 심플웍스 상품/수량 화면 캡쳐를 올려주세요.")
        full_text_quantities = extract_simpleworks_quantities_from_text(full_text)
    except Exception as exc:
        if "검수 결과 화면 캡쳐" in str(exc):
            raise
        if not table_quantities:
            raise
    merged = merge_quantity_maps(table_quantities, full_text_quantities)
    return filter_known_simpleworks_quantities(merged)


def get_coupang_quantities_by_simple_no(date_from: str, date_to: str) -> tuple[dict[str, int], list[tuple[str, str, int]]]:
    restore_sales_files_from_supabase()
    sku_to_simple, _simple_to_info = get_master_simpleworks_maps()
    expected: dict[str, int] = defaultdict(int)
    unmapped: list[tuple[str, str, int]] = []
    if not SALES_LEDGER_PATH.exists():
        rows = iter_sales_rows_from_display_workbook()
    else:
        rows = iter_sales_rows_from_ledger()
    for day, sku, name, qty in rows:
        if date_from and day < date_from:
            continue
        if date_to and day > date_to:
            continue
        simple_no = sku_to_simple.get(sku, "")
        if simple_no:
            expected[simple_no] += qty
        elif qty:
            unmapped.append((sku, name, qty))
    return dict(expected), unmapped


def iter_sales_rows_from_ledger() -> list[tuple[str, str, str, int]]:
    wb = load_workbook(SALES_LEDGER_PATH, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str, str, int]] = []
    for row in range(2, ws.max_row + 1):
        day = str(ws.cell(row, 1).value or "").strip()
        sku = str(ws.cell(row, 4).value or "").strip()
        name = str(ws.cell(row, 5).value or "").strip()
        original_qty = parse_int(ws.cell(row, 6).value)
        adjusted_qty_raw = str(ws.cell(row, 10).value or "").strip()
        qty = parse_int(adjusted_qty_raw, original_qty) if adjusted_qty_raw else original_qty
        rows.append((day, sku, name, qty))
    return rows


def iter_sales_rows_from_display_workbook() -> list[tuple[str, str, str, int]]:
    if not MONTHLY_SALES_PATH.exists():
        return []
    wb = load_workbook(MONTHLY_SALES_PATH, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str, str, int]] = []
    current_day = ""
    for row in range(1, ws.max_row + 1):
        first = str(ws.cell(row, 1).value or "").strip()
        day_match = re.match(r"(\d{4}-\d{2}-\d{2})\s+합계", first)
        if day_match:
            current_day = day_match.group(1)
            continue
        if not current_day or first in ["SKU ID", "합계"] or not first:
            continue
        sku = first
        name = str(ws.cell(row, 2).value or "").strip()
        qty = parse_int(ws.cell(row, 3).value)
        if sku and qty:
            rows.append((current_day, sku, name, qty))
    return rows


def get_sales_ledger_date_summary() -> str:
    restore_sales_files_from_supabase()
    totals: dict[str, int] = defaultdict(int)
    if SALES_LEDGER_PATH.exists():
        rows = iter_sales_rows_from_ledger()
    else:
        rows = iter_sales_rows_from_display_workbook()
    for day, _sku, _name, qty in rows:
        if day:
            totals[day] += qty
    if not totals:
        return "서버의 월별납품 자료가 비어 있습니다."
    parts = [f"{day}: {qty:,}개" for day, qty in sorted(totals.items())]
    return "서버에 있는 쿠팡 납품자료 날짜: " + ", ".join(parts)


def render_check_result(
    simpleworks_qty: dict[str, int],
    date_from: str,
    date_to: str,
) -> str:
    _sku_to_simple, simple_to_info = get_master_simpleworks_maps()
    coupang_qty, unmapped = get_coupang_quantities_by_simple_no(date_from, date_to)
    if not coupang_qty and not unmapped:
        date_summary = html.escape(get_sales_ledger_date_summary())
        return (
            '<section class="panel"><div class="panel-head">검수 결과</div><div class="panel-body">'
            '<div class="msg err">선택한 날짜 범위에 쿠팡 납품자료가 없습니다. '
            '월별납품관리에서 해당 날짜 PO를 먼저 반영했는지, 검수 날짜가 맞는지 확인해주세요.'
            f'<br>{date_summary}</div>'
            '</div></section>'
        )
    all_simple_nos = sorted(set(coupang_qty) | set(simpleworks_qty), key=lambda value: int(value) if value.isdigit() else value)
    rows = []
    ok_count = diff_count = coupang_only = simple_only = 0
    for simple_no in all_simple_nos:
        expected = coupang_qty.get(simple_no, 0)
        actual = simpleworks_qty.get(simple_no, 0)
        actual = normalize_ocr_qty_against_expected(actual, expected)
        diff = actual - expected
        info = simple_to_info.get(simple_no, {"sku": "", "name": ""})
        if expected == actual:
            status = '<span class="ok-text">일치</span>'
            ok_count += 1
        elif expected and actual:
            status = '<span class="bad-text">수량차이</span>'
            diff_count += 1
        elif expected:
            status = '<span class="warn-text">쿠팡만 있음</span>'
            coupang_only += 1
        else:
            status = '<span class="warn-text">심플웍스만 있음</span>'
            simple_only += 1
        rows.append(
            "<tr>"
            f"<td>{status}</td><td>{html.escape(simple_no)}</td><td>{html.escape(info.get('sku', ''))}</td>"
            f"<td>{html.escape(info.get('name', ''))}</td><td>{expected:,}</td><td>{actual:,}</td><td>{diff:,}</td>"
            "</tr>"
        )
    for sku, name, qty in unmapped:
        rows.append(
            '<tr><td><span class="bad-text">기초자료 매칭 없음</span></td>'
            f"<td></td><td>{html.escape(sku)}</td><td>{html.escape(name)}</td><td>{qty:,}</td><td>0</td><td>{-qty:,}</td></tr>"
        )
    if not rows:
        rows.append('<tr><td colspan="7">검수할 자료가 없습니다.</td></tr>')
    return (
        '<section class="panel"><div class="panel-head">검수 결과</div><div class="panel-body">'
        f'<div class="summary"><div>일치<b>{ok_count:,}</b></div><div>수량차이<b>{diff_count:,}</b></div>'
        f'<div>쿠팡만 있음<b>{coupang_only + len(unmapped):,}</b></div><div>심플웍스만 있음<b>{simple_only:,}</b></div></div>'
        '<div class="scroll" style="margin-top:14px;"><table><colgroup><col style="width:120px;"><col style="width:110px;"><col style="width:120px;"><col style="width:380px;"><col style="width:110px;"><col style="width:120px;"><col style="width:100px;"></colgroup>'
        '<thead><tr><th>상태</th><th>심플웍스 No</th><th>SKU ID</th><th>상품명</th><th>쿠팡 수량</th><th>심플웍스 수량</th><th>차이</th></tr></thead>'
        f"<tbody>{''.join(rows)}</tbody></table></div></div></section>"
    )


def render_check_page(message: str = "", result: str = "", date_from: str = "", date_to: str = "") -> str:
    today = datetime.now()
    default_from = date_from or today.strftime("%Y-%m-01")
    default_to = date_to or today.strftime("%Y-%m-%d")
    return (
        CHECK_PAGE
        .replace("{message}", message)
        .replace("{result}", result)
        .replace("{date_from}", html.escape(default_from, quote=True))
        .replace("{date_to}", html.escape(default_to, quote=True))
    )


def handle_sales_upload(form: cgi.FieldStorage) -> tuple[str, str | None]:
    po_items = form["sales_po_files"] if "sales_po_files" in form else None
    if po_items is None:
        raise ValueError("매출확인용 PO 파일을 올려주세요.")
    if not isinstance(po_items, list):
        po_items = [po_items]

    run_id = uuid.uuid4().hex[:12]
    work_dir = RUNS_DIR / f"sales_{run_id}"
    input_dir = work_dir / "uploaded_po"
    input_dir.mkdir(parents=True, exist_ok=True)

    all_lines = []
    file_count = 0
    for item in po_items:
        if not getattr(item, "filename", ""):
            continue
        name = safe_name(item.filename)
        if name.startswith("~$") or not name.lower().endswith(".xlsx"):
            continue
        po_path = input_dir / name
        po_path.write_bytes(item.file.read())
        all_lines.extend(read_po_lines(po_path))
        file_count += 1

    if file_count == 0:
        raise ValueError("처리할 PO 엑셀 파일이 없습니다.")
    if not all_lines:
        raise ValueError("PO 안에서 상품 내역을 찾지 못했습니다.")

    filled_amounts = update_master_amounts_from_lines(all_lines)
    line_count, total_amount = update_monthly_sales(all_lines)
    prefer_inbound = has_inbound_sales_data(all_lines)
    total_qty = sum(get_sales_qty(line, prefer_inbound) for line in all_lines)
    amount_message = f" 기초자료 금액 {filled_amounts}건도 자동으로 채웠습니다." if filled_amounts else ""
    date_message = describe_sales_date_breakdown(all_lines)
    return (
        f"매출확인용으로 저장했습니다. PO {file_count}개, 상품 줄 {line_count}개, "
        f"총 수량 {total_qty:,}개, 납품상품 합계금액 {total_amount:,}원입니다.{amount_message}{date_message}"
    ), None


def save_sales_confirmation_form(form: cgi.FieldStorage, username: str) -> str:
    month = form["month"].value.strip() if "month" in form else ""
    sales_amount = parse_int(form["sales_amount"].value) if "sales_amount" in form else 0
    invoice_amount = parse_int(form["invoice_amount"].value) if "invoice_amount" in form else 0
    action = form["action"].value.strip() if "action" in form else "save"
    wants_confirmed = "confirmed" in form or action == "reconfirm"
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("확인할 월이 올바르지 않습니다.")
    if wants_confirmed and (not invoice_amount or invoice_amount != sales_amount):
        raise ValueError("두 금액이 정확히 일치할 때만 계산서 발행 확인을 완료할 수 있습니다.")

    data = load_sales_confirmations()
    months = data.setdefault("months", {})
    record = months.setdefault(month, {})
    old_confirmed = bool(record.get("confirmed"))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    history = record.setdefault("history", [])
    if old_confirmed and not wants_confirmed:
        history.append({"at": now, "user": username, "action": "확인 완료 해제", "reason": "사용자 확인 후 해제", "changes": []})
    elif action == "reconfirm":
        history.append({"at": now, "user": username, "action": "재확인 완료", "reason": "기존 변경 이력 확인", "changes": []})
    elif wants_confirmed and not old_confirmed:
        history.append({"at": now, "user": username, "action": "계산서 발행 확인 완료", "reason": "금액 일치 확인", "changes": []})

    record["invoice_amount"] = invoice_amount
    record["sales_amount"] = sales_amount
    record["confirmed"] = wants_confirmed
    if wants_confirmed:
        record["confirmed_by"] = username
        record["confirmed_at"] = now
        record["confirmed_invoice_amount"] = invoice_amount
        record["confirmed_sales_amount"] = sales_amount
    if action == "reconfirm":
        record["needs_recheck"] = False
        record["edited_after_confirm"] = False
        record["reconfirmed_by"] = username
        record["reconfirmed_at"] = now
    save_sales_confirmations(data)
    if action == "reconfirm":
        return f"{month} 변경 이력을 확인하고 재확인 완료했습니다."
    return f"{month} 계산서 발행 확인 정보를 저장했습니다."


def save_sales_detail_form(form: cgi.FieldStorage, username: str) -> int:
    wb, ws = ensure_monthly_sales_book()
    row_values = form["row"] if "row" in form else []
    if not isinstance(row_values, list):
        row_values = [row_values]

    changed = 0
    override_reason = form["override_reason"].value.strip() if "override_reason" in form else ""
    confirmation_data = load_sales_confirmations()
    confirmation_months = confirmation_data.setdefault("months", {})
    audit_by_month: dict[str, list[dict[str, object]]] = defaultdict(list)
    for item in row_values:
        row = parse_int(item.value)
        if row < 2 or row > ws.max_row:
            continue
        qty_key = f"qty_{row}"
        memo_key = f"memo_{row}"
        original_qty = parse_int(ws.cell(row, 6).value)
        new_qty_text = form[qty_key].value.strip() if qty_key in form else str(original_qty)
        new_memo = form[memo_key].value.strip() if memo_key in form else ""
        new_qty = parse_int(new_qty_text, original_qty)
        adjusted_value = "" if new_qty == original_qty else new_qty
        old_adjusted = str(ws.cell(row, 10).value or "").strip()
        old_memo = str(ws.cell(row, 11).value or "").strip()
        row_changes = []
        if str(adjusted_value) != old_adjusted:
            row_changes.append({"field": f"{ws.cell(row, 4).value} 수량", "before": parse_int(old_adjusted, original_qty), "after": new_qty})
            ws.cell(row, 10).value = adjusted_value
            changed += 1
        if new_memo != old_memo:
            row_changes.append({"field": f"{ws.cell(row, 4).value} 메모", "before": old_memo, "after": new_memo})
            ws.cell(row, 11).value = new_memo
            changed += 1
        month = str(ws.cell(row, 1).value or "")[:7]
        month_record = confirmation_months.get(month, {}) if isinstance(confirmation_months, dict) else {}
        if row_changes and isinstance(month_record, dict) and month_record.get("confirmed"):
            if not override_reason:
                raise ValueError("계산서 발행 확인 완료 건을 수정하려면 수정 사유가 필요합니다.")
            audit_by_month[month].extend(row_changes)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for month, changes in audit_by_month.items():
        record = confirmation_months.setdefault(month, {})
        record["needs_recheck"] = True
        record["edited_after_confirm"] = True
        record.setdefault("history", []).append({
            "at": now,
            "user": username,
            "action": "확인 완료 후 수정",
            "reason": override_reason,
            "changes": changes,
        })
    wb.save(SALES_LEDGER_PATH)
    if audit_by_month:
        save_sales_confirmations(confirmation_data)
    write_sales_display_workbook()
    backup_sales_files_to_supabase()
    return changed


def delete_sales_detail_form(form: cgi.FieldStorage) -> tuple[int, str]:
    wb, ws = ensure_monthly_sales_book()
    target_day = form["delete_day"].value.strip() if "delete_day" in form else ""
    target_month = form["delete_month"].value.strip() if "delete_month" in form else ""
    target_row = parse_int(form["delete_row"].value) if "delete_row" in form else 0

    rows_to_delete = []
    if target_month:
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, 1).value or "").strip().startswith(target_month):
                rows_to_delete.append(row)
        label = f"{target_month} 월 전체"
    elif target_day:
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, 1).value or "").strip() == target_day:
                rows_to_delete.append(row)
        label = f"{target_day} 일자"
    elif target_row:
        if 2 <= target_row <= ws.max_row:
            rows_to_delete.append(target_row)
        sku = str(ws.cell(target_row, 4).value or "").strip() if rows_to_delete else ""
        day = str(ws.cell(target_row, 1).value or "").strip() if rows_to_delete else ""
        label = f"{day} / {sku}"
    else:
        raise ValueError("삭제할 일자 또는 상품 줄을 찾지 못했습니다.")

    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row, 1)
    wb.save(SALES_LEDGER_PATH)
    write_sales_display_workbook()
    backup_sales_files_to_supabase()
    return len(rows_to_delete), label


def zip_dir(source_dir: Path, zip_path: Path) -> None:
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in source_dir.rglob("*"):
            if file_path.is_file():
                zf.write(file_path, file_path.relative_to(source_dir))


def build_message(kind: str, text: str, link: str | None = None) -> str:
    cls = "ok" if kind == "ok" else "err"
    link_html = f'<br><a class="download" href="{html.escape(link)}">결과 ZIP 다운로드</a>' if link else ""
    return f'<div style="height:18px"></div><div class="status {cls}">{html.escape(text)}{link_html}</div>'


class BonnieHandler(BaseHTTPRequestHandler):
    SESSION_COOKIE = "coupang_session"
    COOKIE_SECURE = os.environ.get("COOKIE_SECURE", "").strip() == "1"

    def get_cookie(self, name: str) -> str | None:
        for part in self.headers.get("Cookie", "").split(";"):
            key, _, value = part.strip().partition("=")
            if key == name:
                return value
        return None

    def current_user(self) -> dict[str, object] | None:
        return AUTH.get_user(self.get_cookie(self.SESSION_COOKIE))

    def read_urlencoded_form(self) -> dict[str, str]:
        length = int(self.headers.get("Content-Length", "0") or "0")
        body = self.rfile.read(length).decode("utf-8")
        parsed = parse_qs(body, keep_blank_values=True)
        return {key: values[-1] if values else "" for key, values in parsed.items()}

    def send_redirect(self, location: str, cookies: list[str] | None = None) -> None:
        self.send_response(302)
        self.send_header("Location", location)
        for cookie in cookies or []:
            self.send_header("Set-Cookie", cookie)
        self.end_headers()

    def session_cookie(self, token: str) -> str:
        secure = "; Secure" if self.COOKIE_SECURE else ""
        return f"{self.SESSION_COOKIE}={token}; Path=/; HttpOnly; SameSite=Lax{secure}; Max-Age=28800"

    def clear_session_cookie(self) -> str:
        secure = "; Secure" if self.COOKIE_SECURE else ""
        return f"{self.SESSION_COOKIE}=; Path=/; HttpOnly; SameSite=Lax{secure}; Max-Age=0"

    def require_user(self) -> dict[str, object] | None:
        user = self.current_user()
        if user is None:
            self.send_redirect("/login")
            return None
        return user

    def require_permission(self, permission: str) -> dict[str, object] | None:
        user = self.require_user()
        if user is None:
            return None
        if not AUTH.has_permission(user, permission):
            self.send_html(login_page("접근 권한이 없습니다."), status=403)
            return None
        return user

    def decorate_page(self, text: str) -> str:
        user = self.current_user()
        if user is None:
            return text
        admin_link = ""
        if AUTH.has_permission(user, "admin"):
            admin_link = '<div class="nav-item" onclick="location.href=\'/admin\'">관리자모드</div>'
        user_html = (
            '<div style="margin-top:24px;padding:12px;border-top:1px solid rgba(255,255,255,.18);'
            'color:#d9e7f3;font-size:13px;line-height:1.6;">'
            f'{html.escape(str(user["username"]))}<br>'
            '<a href="/logout" style="color:#fff;font-weight:800;text-decoration:none;">로그아웃</a>'
            '</div>'
        )
        return text.replace("</aside>", f"{admin_link}{user_html}</aside>", 1)

    def page(self, message: str = "") -> str:
        saved_master = get_saved_master_path()
        if saved_master:
            master_status = f"현재 저장된 기초자료: {html.escape(saved_master.name)}"
        else:
            master_status = "저장된 기초자료가 없습니다. 처음 1회는 기초자료 엑셀을 선택해주세요."
        return self.decorate_page(HTML_PAGE.replace("{message}", message).replace("{master_status}", master_status))

    def master_page(self, message: str = "") -> str:
        master_path = get_saved_master_path()
        if master_path is None:
            rows = '<tr><td colspan="10">저장된 기초자료가 없습니다. 먼저 PO 변환 화면에서 기초자료를 올려주세요.</td></tr>'
        else:
            rows = render_master_rows(master_path)
        return self.decorate_page(
            MASTER_PAGE
            .replace("{message}", message)
            .replace("{incentive_tables}", render_growth_incentive_tables())
            .replace("{rows}", rows)
        )

    def pallet_page(self, message: str = "") -> str:
        saved_master = get_saved_master_path()
        if saved_master:
            master_status = f"현재 저장된 기초자료: {html.escape(saved_master.name)}"
        else:
            master_status = "저장된 기초자료가 없습니다. 먼저 쿠팡 PO 변환 화면에서 기초자료를 1회 올려주세요."
        return self.decorate_page(PALLET_PAGE.replace("{message}", message).replace("{master_status}", master_status))

    def check_page(self, message: str = "", result: str = "", date_from: str = "", date_to: str = "") -> str:
        return self.decorate_page(render_check_page(message, result, date_from, date_to))

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/login":
            self.send_html(login_page())
            return

        if parsed.path == "/logout":
            AUTH.logout(self.get_cookie(self.SESSION_COOKIE))
            self.send_redirect("/login", [self.clear_session_cookie()])
            return

        if parsed.path == "/admin":
            if self.require_permission("admin") is None:
                return
            self.send_html(admin_page(AUTH.list_users(), AUTH.admin_email))
            return

        if parsed.path == "/":
            if self.require_permission("po_convert") is None:
                return
            self.send_html(self.page())
            return

        if parsed.path == "/master":
            if self.require_permission("master") is None:
                return
            self.send_html(self.master_page())
            return

        if parsed.path == "/pallet":
            if self.require_permission("pallet") is None:
                return
            self.send_html(self.pallet_page())
            return

        if parsed.path == "/check":
            if self.require_permission("check") is None:
                return
            self.send_html(self.check_page())
            return

        if parsed.path == "/sales":
            if self.require_permission("sales") is None:
                return
            self.send_html(self.decorate_page(render_sales_page()))
            return

        if parsed.path == "/sales/folders":
            if self.require_permission("sales") is None:
                return
            self.send_html(self.decorate_page(render_sales_page(folder_mode=True)))
            return

        if parsed.path == "/sales/download":
            if self.require_permission("sales") is None:
                return
            if SALES_LEDGER_PATH.exists():
                write_sales_display_workbook()
            if not MONTHLY_SALES_PATH.exists():
                self.send_html(render_sales_page(build_message("err", "아직 월매출 엑셀 파일이 없습니다.")), status=404)
                return
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote('월매출_납품상품.xlsx')}")
            self.send_header("Content-Length", str(MONTHLY_SALES_PATH.stat().st_size))
            self.end_headers()
            with MONTHLY_SALES_PATH.open("rb") as f:
                shutil.copyfileobj(f, self.wfile)
            return

        if parsed.path == "/sales/tester/download":
            if self.require_permission("sales") is None:
                return
            file_id = parse_qs(parsed.query).get("id", [""])[-1]
            item = next((value for value in load_tester_files_meta() if str(value.get("id")) == file_id), None)
            if item is None:
                self.send_html(render_sales_page(build_message("err", "체험단 원본 파일을 찾지 못했습니다.")), status=404)
                return
            stored_path = TESTER_FILES_DIR / Path(str(item.get("stored_name", ""))).name
            restore_file_from_supabase(str(item.get("file_key", "")), stored_path)
            if not stored_path.exists():
                self.send_html(render_sales_page(build_message("err", "저장된 체험단 원본 파일을 복원하지 못했습니다.")), status=404)
                return
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(str(item.get('name', '체험단.xlsx')))}")
            self.send_header("Content-Length", str(stored_path.stat().st_size))
            self.end_headers()
            with stored_path.open("rb") as file:
                shutil.copyfileobj(file, self.wfile)
            return

        if parsed.path.startswith("/download/"):
            if self.require_user() is None:
                return
            run_id = unquote(parsed.path.removeprefix("/download/"))
            zip_path = RUNS_DIR / run_id / "result.zip"
            if not zip_path.exists():
                self.send_html(self.page(build_message("err", "결과 파일을 찾지 못했습니다.")), status=404)
                return
            self.send_response(200)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote('쿠팡_PO_변환결과.zip')}")
            self.send_header("Content-Length", str(zip_path.stat().st_size))
            self.end_headers()
            with zip_path.open("rb") as f:
                shutil.copyfileobj(f, self.wfile)
            return

        self.send_html(self.page(build_message("err", "없는 화면입니다.")), status=404)

    def do_POST(self) -> None:
        path = urlparse(self.path).path
        if path == "/login":
            form = self.read_urlencoded_form()
            token = AUTH.login(form.get("username", "").strip(), form.get("password", ""))
            if token is None:
                self.send_html(login_page("아이디 또는 비밀번호가 맞지 않습니다."), status=401)
                return
            self.send_redirect("/", [self.session_cookie(token)])
            return

        if path != "/convert":
            if path == "/admin/users/save":
                if self.require_permission("admin") is None:
                    return
                form = parse_qs(self.rfile.read(int(self.headers.get("Content-Length", "0") or "0")).decode("utf-8"))
                username = form.get("username", [""])[-1].strip()
                password = form.get("password", [""])[-1]
                permissions = form.get("permissions", [])
                try:
                    AUTH.upsert_user(username, password, permissions)
                    self.send_html(admin_page(AUTH.list_users(), AUTH.admin_email, "저장되었습니다."))
                except Exception as exc:
                    self.send_html(admin_page(AUTH.list_users(), AUTH.admin_email, f"저장 중 오류가 났습니다: {exc}"), status=400)
                return
            if path == "/admin/users/delete":
                if self.require_permission("admin") is None:
                    return
                form = self.read_urlencoded_form()
                try:
                    AUTH.delete_user(form.get("username", "").strip())
                    self.send_html(admin_page(AUTH.list_users(), AUTH.admin_email, "삭제되었습니다."))
                except Exception as exc:
                    self.send_html(admin_page(AUTH.list_users(), AUTH.admin_email, f"삭제 중 오류가 났습니다: {exc}"), status=400)
                return
            if path == "/master/save":
                if self.require_permission("master") is None:
                    return
                self.handle_master_save()
                return
            if path == "/master/growth-incentive/save":
                if self.require_permission("master") is None:
                    return
                self.handle_growth_incentive_save()
                return
            if path == "/master/upload":
                if self.require_permission("po_convert") is None:
                    return
                self.handle_master_upload_request()
                return
            if path == "/sales/upload":
                if self.require_permission("sales") is None:
                    return
                self.handle_sales_upload_request()
                return
            if path == "/sales/tester/upload":
                if self.require_permission("sales") is None:
                    return
                self.handle_tester_upload_request()
                return
            if path == "/sales/save":
                if self.require_permission("sales") is None:
                    return
                self.handle_sales_save_request()
                return
            if path == "/sales/confirm":
                if self.require_permission("sales") is None:
                    return
                self.handle_sales_confirm_request()
                return
            if path == "/sales/delete":
                if self.require_permission("sales") is None:
                    return
                self.handle_sales_delete_request()
                return
            if path == "/sales/year/save":
                if self.require_permission("sales") is None:
                    return
                self.handle_sales_year_save_request()
                return
            if path == "/pallet/create":
                if self.require_permission("pallet") is None:
                    return
                self.handle_pallet_create_request()
                return
            if path == "/check/run":
                if self.require_permission("check") is None:
                    return
                self.handle_check_request()
                return
            self.send_html(self.page(build_message("err", "없는 요청입니다.")), status=404)
            return

        if self.require_permission("po_convert") is None:
            return
        try:
            message, link = self.handle_convert()
            self.send_html(self.page(build_message("ok", message, link)))
        except Exception as exc:
            self.send_html(self.page(build_message("err", f"처리 중 오류가 났습니다: {exc}")), status=500)

    def handle_tester_upload_request(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        try:
            upload_items = form["tester_files"] if "tester_files" in form else []
            if not isinstance(upload_items, list):
                upload_items = [upload_items]
            files = []
            for item in upload_items:
                filename = safe_name(getattr(item, "filename", ""))
                if not filename:
                    continue
                if filename.startswith("~$") or not filename.lower().endswith(".xlsx"):
                    raise ValueError("체험단 자료는 .xlsx 파일만 올려주세요.")
                files.append((filename, item.file.read()))
            if not files:
                raise ValueError("체험단 엑셀 파일을 선택해주세요.")
            count, months = save_uploaded_tester_files(files)
            message = f'체험단 원본 {count}개를 저장하고 {", ".join(months)} 체험단 금액에 자동 반영했습니다.'
            self.send_html(self.decorate_page(render_sales_page(build_message("ok", message))))
        except Exception as exc:
            self.send_html(self.decorate_page(render_sales_page(build_message("err", f"체험단 자료 저장 중 오류가 났습니다: {exc}"))), status=500)

    def handle_sales_upload_request(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        try:
            message, _ = handle_sales_upload(form)
            self.send_html(render_sales_page(build_message("ok", message), folder_mode=True))
        except Exception as exc:
            self.send_html(render_sales_page(build_message("err", f"월별납품 저장 중 오류가 났습니다: {exc}"), folder_mode=True), status=500)

    def handle_check_request(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        date_from = form["date_from"].value.strip() if "date_from" in form else ""
        date_to = form["date_to"].value.strip() if "date_to" in form else ""
        try:
            if date_from and date_to and date_from > date_to:
                raise ValueError("검수 시작일이 종료일보다 뒤입니다. 날짜를 다시 선택해주세요.")
            excel_items = form["simpleworks_file"] if "simpleworks_file" in form else []
            image_items = form["simpleworks_image"] if "simpleworks_image" in form else []
            if not isinstance(excel_items, list):
                excel_items = [excel_items]
            if not isinstance(image_items, list):
                image_items = [image_items]
            excel_items = [item for item in excel_items if getattr(item, "filename", "")]
            image_items = [item for item in image_items if getattr(item, "filename", "")]
            if not excel_items and not image_items:
                raise ValueError("심플웍스 엑셀 또는 캡쳐 이미지 파일을 올려주세요.")
            run_id = uuid.uuid4().hex[:12]
            work_dir = RUNS_DIR / f"check_{run_id}"
            work_dir.mkdir(parents=True, exist_ok=True)
            simpleworks_qty: dict[str, int] = defaultdict(int)
            excel_count = 0
            image_count = 0
            for item in excel_items:
                name = safe_name(item.filename)
                if not name.lower().endswith(".xlsx"):
                    raise ValueError("심플웍스 엑셀은 .xlsx 파일로 올려주세요.")
                file_path = work_dir / name
                file_path.write_bytes(item.file.read())
                parsed_qty = parse_simpleworks_excel(file_path)
                for simple_no, qty in parsed_qty.items():
                    simpleworks_qty[simple_no] += qty
                excel_count += 1
            for item in image_items:
                name = safe_name(item.filename)
                if not name.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".webp")):
                    raise ValueError("캡쳐 이미지는 png, jpg, jpeg, bmp, webp 파일로 올려주세요.")
                file_path = work_dir / name
                file_path.write_bytes(item.file.read())
                parsed_qty = parse_simpleworks_image(file_path)
                for simple_no, qty in parsed_qty.items():
                    simpleworks_qty[simple_no] += qty
                image_count += 1
            simpleworks_qty = dict(simpleworks_qty)
            if not simpleworks_qty:
                raise ValueError("올린 파일에서 심플웍스 No와 수량을 찾지 못했습니다.")
            result = render_check_result(simpleworks_qty, date_from, date_to)
            msg = build_message("ok", f"검수 완료: 심플웍스 엑셀 {excel_count:,}개, 캡쳐 이미지 {image_count:,}개에서 심플웍스 No {len(simpleworks_qty):,}개를 합산했습니다.")
            self.send_html(self.check_page(msg, result, date_from, date_to))
        except Exception as exc:
            self.send_html(self.check_page(build_message("err", f"수량검수 중 오류가 났습니다: {exc}"), "", date_from, date_to), status=500)

    def handle_sales_save_request(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        try:
            user = self.current_user() or {}
            changed = save_sales_detail_form(form, str(user.get("username", "알 수 없음")))
            self.send_html(self.decorate_page(render_sales_page(build_message("ok", f"수량/메모 수정사항 {changed}건을 저장했습니다."), folder_mode=True)))
        except Exception as exc:
            self.send_html(self.decorate_page(render_sales_page(build_message("err", f"수량/메모 저장 중 오류가 났습니다: {exc}"), folder_mode=True)), status=400)

    def handle_sales_confirm_request(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        try:
            user = self.current_user() or {}
            message = save_sales_confirmation_form(form, str(user.get("username", "알 수 없음")))
            self.send_html(self.decorate_page(render_sales_page(build_message("ok", message))))
        except Exception as exc:
            self.send_html(self.decorate_page(render_sales_page(build_message("err", f"계산서 확인 저장 중 오류가 났습니다: {exc}"))), status=400)

    def handle_sales_delete_request(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        try:
            deleted, label = delete_sales_detail_form(form)
            self.send_html(render_sales_page(build_message("ok", f"{label} 삭제 완료: {deleted}줄을 삭제했습니다.")))
        except Exception as exc:
            self.send_html(render_sales_page(build_message("err", f"삭제 중 오류가 났습니다: {exc}")), status=500)

    def handle_sales_year_save_request(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        try:
            changed = save_year_manual(form)
            self.send_html(render_sales_page(build_message("ok", f"연도총매출 수기 입력값 {changed}건을 저장했습니다.")))
        except Exception as exc:
            self.send_html(render_sales_page(build_message("err", f"연도총매출 저장 중 오류가 났습니다: {exc}")), status=500)

    def handle_master_save(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        try:
            changed = save_master_form(form)
            self.send_html(self.master_page(build_message("ok", f"저장되었습니다. 변경 {changed}건을 반영했습니다.")))
        except Exception as exc:
            self.send_html(self.master_page(build_message("err", f"저장 중 오류가 났습니다: {exc}")), status=500)

    def handle_growth_incentive_save(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        try:
            save_growth_incentive_form(form)
            self.send_html(self.master_page(build_message("ok", "성장장려금 기초자료를 저장했습니다. 연도총매출에 자동 반영됩니다.")))
        except Exception as exc:
            self.send_html(self.master_page(build_message("err", f"성장장려금 기초자료 저장 중 오류가 났습니다: {exc}")), status=500)

    def handle_master_upload_request(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )
        next_page = form["next"].value.strip() if "next" in form else ""
        try:
            item = form["master"] if "master" in form else None
            if item is None or not getattr(item, "filename", ""):
                raise ValueError("기초자료 엑셀 파일을 선택해주세요.")
            name = safe_name(item.filename)
            if name.startswith("~$") or not name.lower().endswith(".xlsx"):
                raise ValueError("기초자료는 .xlsx 파일로 올려주세요.")

            master_bytes = item.file.read()
            fd, check_name = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            check_path = Path(check_name)
            try:
                check_path.write_bytes(master_bytes)
                read_master(check_path)
            finally:
                check_path.unlink(missing_ok=True)

            replace_saved_master_file(name, master_bytes)
            message = build_message("ok", f"기초자료를 교체 저장했습니다: {name}")
            if next_page == "master":
                self.send_html(self.master_page(message))
            else:
                self.send_html(self.page(message))
        except Exception as exc:
            message = build_message("err", f"기초자료 저장 중 오류가 났습니다: {exc}")
            if next_page == "master":
                self.send_html(self.master_page(message), status=500)
            else:
                self.send_html(self.page(message), status=500)

    def handle_pallet_create_request(self) -> None:
        try:
            message, link = self.handle_pallet_create()
            self.send_html(self.pallet_page(build_message("ok", message, link)))
        except Exception as exc:
            self.send_html(self.pallet_page(build_message("err", f"파렛트/쉽먼트 초안 생성 중 오류가 났습니다: {exc}")), status=500)

    def handle_pallet_create(self) -> tuple[str, str]:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )

        saved_master = get_saved_master_path()
        if saved_master is None:
            raise ValueError("저장된 기초자료가 없습니다. 먼저 쿠팡 PO 변환 화면에서 기초자료를 1회 올려주세요.")

        po_items = form["po_files"] if "po_files" in form else None
        if po_items is None:
            raise ValueError("PO 엑셀을 올려주세요.")
        if not isinstance(po_items, list):
            po_items = [po_items]

        run_id = uuid.uuid4().hex[:12]
        work_dir = RUNS_DIR / run_id
        input_dir = work_dir / "input_po"
        output_dir = work_dir / "result"
        processed_dir = output_dir / "심플웍스_업로드용_PO복사본"
        input_dir.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)

        po_paths: list[Path] = []
        for item in po_items:
            if not getattr(item, "filename", ""):
                continue
            name = safe_name(item.filename)
            if name.startswith("~$") or not name.lower().endswith(".xlsx"):
                continue
            po_path = input_dir / name
            po_path.write_bytes(item.file.read())
            po_paths.append(po_path)
        if not po_paths:
            raise ValueError("처리할 PO 엑셀 파일이 없습니다.")

        master = read_master(saved_master)
        all_lines = []
        for po_path in sorted(po_paths):
            output_file = processed_dir / po_path.name.replace(".xlsx", " - 복사본.xlsx")
            all_lines.extend(create_processed_po(po_path, output_file, master))

        filled_amounts = update_master_amounts_from_lines(all_lines)
        write_summary_workbook(all_lines, output_dir / "파렛트_쉽먼트_초안.xlsx")
        zip_path = work_dir / "result.zip"
        zip_dir(output_dir, zip_path)

        total_qty = sum(line.available_qty for line in all_lines)
        pallet_ready = sum(1 for line in all_lines if line.pallet_qty)
        amount_message = f" 기초자료 금액 {filled_amounts}건도 자동으로 채웠습니다." if filled_amounts else ""
        return (
            f"파렛트/쉽먼트 초안을 만들었습니다. PO {len(po_paths)}개, 상품 줄 {len(all_lines)}개, "
            f"총 수량 {total_qty:,}개입니다. 파렛트 기준이 있는 줄은 {pallet_ready}개입니다.{amount_message}"
        ), f"/download/{run_id}"

    def handle_convert(self) -> tuple[str, str]:
        content_type = self.headers.get("Content-Type", "")
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type},
        )

        master_item = form["master"] if "master" in form else None
        po_items = form["po_files"] if "po_files" in form else None
        if po_items is None:
            raise ValueError("PO 엑셀을 올려주세요.")
        if not isinstance(po_items, list):
            po_items = [po_items]

        run_id = uuid.uuid4().hex[:12]
        work_dir = RUNS_DIR / run_id
        input_dir = work_dir / "input_po"
        output_dir = work_dir / "result"
        processed_dir = output_dir / "심플웍스_업로드용_PO복사본"
        input_dir.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)

        master_path: Path | None = None
        if master_item is not None and getattr(master_item, "filename", ""):
            master_name = safe_name(master_item.filename)
            master_path = work_dir / master_name
            master_bytes = master_item.file.read()
            master_path.write_bytes(master_bytes)
            replace_saved_master_file(master_name, master_bytes)
        else:
            saved_master = get_saved_master_path()
            if saved_master is None:
                raise ValueError("저장된 기초자료가 없습니다. 처음 1회는 기초자료 엑셀을 올려주세요.")
            master_path = saved_master

        po_paths: list[Path] = []
        for item in po_items:
            if not getattr(item, "filename", ""):
                continue
            name = safe_name(item.filename)
            if name.startswith("~$") or not name.lower().endswith(".xlsx"):
                continue
            po_path = input_dir / name
            po_path.write_bytes(item.file.read())
            po_paths.append(po_path)
        if not po_paths:
            raise ValueError("처리할 PO 엑셀 파일이 없습니다.")

        master = read_master(master_path)

        all_lines = []
        for po_path in sorted(po_paths):
            output_file = processed_dir / po_path.name.replace(".xlsx", " - 복사본.xlsx")
            all_lines.extend(create_processed_po(po_path, output_file, master))

        filled_amounts = update_master_amounts_from_lines(all_lines)

        zip_path = work_dir / "result.zip"
        zip_dir(output_dir, zip_path)
        total_qty = sum(line.available_qty for line in all_lines)
        total_amount = sum(line.order_amount for line in all_lines)
        amount_message = f" 기초자료 금액 {filled_amounts}건도 자동으로 채웠습니다." if filled_amounts else ""
        return (
            f"완료되었습니다. PO {len(po_paths)}개, 상품 줄 {len(all_lines)}개를 처리했습니다. "
            f"총 수량 {total_qty:,}개, 합계 금액 {total_amount:,}원입니다.{amount_message}"
        ), f"/download/{run_id}"

    def send_html(self, text: str, status: int = 200) -> None:
        data = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format: str, *args) -> None:
        return


def main() -> None:
    RUNS_DIR.mkdir(parents=True, exist_ok=True)
    port = int(os.environ.get("PORT", "8765"))
    server = ThreadingHTTPServer(("0.0.0.0", port), BonnieHandler)
    print("Bonnie Coupang PO web app")
    print(f"Open: http://127.0.0.1:{port}")
    print(f"LAN:  http://<this-computer-ip>:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()



















