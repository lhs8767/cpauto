from __future__ import annotations

import argparse
import csv
import re
import warnings
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

SKU_HEADER_CANDIDATES = ["상품코드", "sku id", "skuid", "sku", "스큐", "스큐아이디"]
UNAVAILABLE_HEADER_CANDIDATES = ["납품불가", "납품불가여부", "불가", "제외"]
QTY_HEADER_CANDIDATES = ["요청", "납품가능수량", "가능수량", "납품수량", "수량"]
SHIPMENT_HEADER_CANDIDATES = ["쉽먼트", "shipment", "쉽먼트그룹", "차수", "트럭", "파렛트"]


@dataclass
class MasterItem:
    sku_id: str
    available: bool
    available_qty: int | None
    shipment_group: str
    note: str
    product_name: str = ""
    barcode: str = ""
    length_mm: int = 0
    width_mm: int = 0
    height_mm: int = 0
    weight_g: int = 0
    inner_qty: int = 0
    box_qty: int = 0
    pallet_layer_qty: int = 0
    pallet_qty: int = 0


@dataclass
class PoLine:
    po_no: str
    source_file: str
    center: str
    inbound_date: str
    row_number: int
    barcode_row_number: int | None
    original_no: str
    sku_id: str
    product_name: str
    barcode: str
    tax_type: str
    order_type: str
    order_qty: int
    available_qty: int
    inbound_qty: int
    purchase_price: int
    supply_price: int
    vat_unit: int
    order_amount: int
    supply_amount: int
    vat_amount: int
    inbound_amount: int = 0
    inbound_supply_amount: int = 0
    inbound_vat_amount: int = 0
    shipment_group: str = ""
    note: str = ""
    length_mm: int = 0
    width_mm: int = 0
    height_mm: int = 0
    weight_g: int = 0
    inner_qty: int = 0
    box_qty: int = 0
    pallet_layer_qty: int = 0
    pallet_qty: int = 0


def normalize(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def to_int(value: Any) -> int:
    text = str(value or "").replace(",", "").strip()
    if text in ["", "-", "None"]:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def yes_no_to_bool(value: Any) -> bool:
    text = normalize(value)
    if text in ["", "0", "n", "no", "x", "불가", "불가능", "납품불가", "제외"]:
        return False
    if any(word in text for word in ["불가", "제외", "중단", "취소"]):
        return False
    if text in ["정상", "가능", "발주가능", "y", "yes", "o", "ok"]:
        return True
    return True


PALLET_BASE_SIDE_MM = 1800
PALLET_OVERHANG_LIMIT_MM = 300
PALLET_OVERHANG_MAX_SIDE_MM = PALLET_BASE_SIDE_MM + PALLET_OVERHANG_LIMIT_MM


def pallet_overhang_note(length_mm: int, width_mm: int) -> str:
    side_mm = max(length_mm or 0, width_mm or 0)
    if side_mm <= 0:
        return ""
    if side_mm <= PALLET_BASE_SIDE_MM:
        return "오버행 기준내"
    if side_mm <= PALLET_OVERHANG_MAX_SIDE_MM:
        return "오버행 확인 필요: E18/E19/AU 18P 1파렛트 한정, 한 변당 150mm/총 300mm까지"
    return "납품불가 확인: 오버행 허용범위 초과"


def is_unavailable_flag(value: Any) -> bool:
    text = normalize(value)
    if text in ["", "0", "n", "no", "x", "false", "아니오", "아님", "가능"]:
        return False
    return True


def find_header_row(ws: Worksheet, candidates: list[str]) -> tuple[int, dict[str, int]]:
    candidate_set = {normalize(c) for c in candidates}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        found = {}
        for cell in row:
            key = normalize(cell.value)
            if key in candidate_set:
                found[key] = cell.column
        if found:
            headers = {}
            for cell in row:
                if cell.value is not None:
                    headers[normalize(cell.value)] = cell.column
            return row[0].row, headers
    raise ValueError("기초자료에서 상품코드 열을 찾지 못했습니다.")


def find_first_header(headers: dict[str, int], candidates: list[str]) -> int | None:
    normalized_candidates = [normalize(c) for c in candidates]
    for candidate in normalized_candidates:
        if candidate in headers:
            return headers[candidate]
    for header, col in headers.items():
        if any(candidate in header for candidate in normalized_candidates):
            return col
    return None


def find_unavailable_value_column(ws: Worksheet, header_row: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        hit_count = 0
        for row in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
            if "납품불가" in str(ws.cell(row, col).value or "").strip():
                hit_count += 1
        if hit_count:
            return col
    return None


def read_master(master_path: Path) -> dict[str, MasterItem]:
    wb = load_workbook(master_path, data_only=True)
    ws = wb.active
    header_row, headers = find_header_row(ws, SKU_HEADER_CANDIDATES)
    sku_col = find_first_header(headers, SKU_HEADER_CANDIDATES)
    unavailable_col = find_first_header(headers, UNAVAILABLE_HEADER_CANDIDATES)
    if unavailable_col is None:
        unavailable_col = find_unavailable_value_column(ws, header_row)
    qty_col = find_first_header(headers, QTY_HEADER_CANDIDATES)
    shipment_col = find_first_header(headers, SHIPMENT_HEADER_CANDIDATES)
    note_col = find_first_header(headers, ["비고", "메모", "note"])
    name_col = find_first_header(headers, ["상품명"])
    barcode_col = find_first_header(headers, ["바코드"])
    length_col = find_first_header(headers, ["길이(mm)", "길이"])
    width_col = find_first_header(headers, ["넓이(mm)", "너비(mm)", "넓이", "너비"])
    height_col = find_first_header(headers, ["높이(mm)", "높이"])
    weight_col = find_first_header(headers, ["중량(g)", "중량"])
    inner_col = find_first_header(headers, ["innercase내sku수량", "innercasesku수량", "inner"])
    box_col = find_first_header(headers, ["box내총수량", "box총수량"])
    pallet_layer_col = find_first_header(headers, ["pallet1단의sku수량", "pallet1단sku수량", "1단"])
    pallet_col = find_first_header(headers, ["pallet내sku수량", "palletsku수량"])

    if sku_col is None:
        raise ValueError("기초자료에 상품코드 또는 SKU ID 열이 필요합니다.")

    master: dict[str, MasterItem] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        sku_id = str(ws.cell(row, sku_col).value or "").strip()
        if not sku_id:
            continue
        unavailable_value = ws.cell(row, unavailable_col).value if unavailable_col else ""
        qty_value = ws.cell(row, qty_col).value if qty_col else None
        shipment_value = ws.cell(row, shipment_col).value if shipment_col else ""
        note_value = ws.cell(row, note_col).value if note_col else ""
        master[sku_id] = MasterItem(
            sku_id=sku_id,
            available=not is_unavailable_flag(unavailable_value),
            available_qty=to_int(qty_value) if qty_value not in [None, ""] else None,
            shipment_group=str(shipment_value or "").strip(),
            note=str(note_value or "").strip(),
            product_name=str(ws.cell(row, name_col).value or "").strip() if name_col else "",
            barcode=str(ws.cell(row, barcode_col).value or "").strip() if barcode_col else "",
            length_mm=to_int(ws.cell(row, length_col).value) if length_col else 0,
            width_mm=to_int(ws.cell(row, width_col).value) if width_col else 0,
            height_mm=to_int(ws.cell(row, height_col).value) if height_col else 0,
            weight_g=to_int(ws.cell(row, weight_col).value) if weight_col else 0,
            inner_qty=to_int(ws.cell(row, inner_col).value) if inner_col else 0,
            box_qty=to_int(ws.cell(row, box_col).value) if box_col else 0,
            pallet_layer_qty=to_int(ws.cell(row, pallet_layer_col).value) if pallet_layer_col else 0,
            pallet_qty=to_int(ws.cell(row, pallet_col).value) if pallet_col else 0,
        )
    return master


def find_po_header(ws: Worksheet) -> int:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() == "No.":
            return row
    raise ValueError(f"{ws.title} 시트에서 상품정보 시작행을 찾지 못했습니다.")


def get_po_no(ws: Worksheet, file_name: str) -> str:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        for cell in row:
            text = str(cell.value or "")
            if text.startswith("발주서 No."):
                return text.replace("발주서 No.", "").strip()
    match = re.search(r"PO_(\d+)", file_name)
    return match.group(1) if match else ""


def get_inbound_date(ws: Worksheet, header_row: int) -> str:
    for row in range(1, header_row):
        for col in range(1, ws.max_column + 1):
            label = normalize(ws.cell(row, col).value)
            if "입고예정" not in label:
                continue
            for check_row in range(row, min(row + 4, header_row) + 1):
                for check_col in range(col, min(col + 3, ws.max_column) + 1):
                    value = ws.cell(check_row, check_col).value
                    text = str(value or "").strip()
                    if not text or "입고예정" in text:
                        continue
                    if re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", text):
                        return text
    return str(ws.cell(header_row - 7, 6).value or "")


def read_po_lines(po_path: Path, master: dict[str, MasterItem] | None = None) -> list[PoLine]:
    wb = load_workbook(po_path, data_only=True)
    ws = wb.active
    po_no = get_po_no(ws, po_path.name)
    header_row = find_po_header(ws)
    center = str(ws.cell(header_row - 7, 3).value or "")
    inbound_date = get_inbound_date(ws, header_row)
    lines: list[PoLine] = []

    row = header_row + 2
    while row <= ws.max_row:
        first_value = str(ws.cell(row, 1).value or "").strip()
        if not first_value:
            row += 1
            continue
        if "합계" in first_value:
            break

        sku_id = str(ws.cell(row, 2).value or "").strip()
        product_name = str(ws.cell(row, 3).value or "").strip()
        if not sku_id or not product_name:
            row += 1
            continue

        barcode_row = row + 1 if row + 1 <= ws.max_row else None
        item = master.get(sku_id) if master else None
        line = PoLine(
            po_no=po_no,
            source_file=po_path.name,
            center=center,
            inbound_date=inbound_date,
            row_number=row,
            barcode_row_number=barcode_row,
            original_no=first_value,
            sku_id=sku_id,
            product_name=product_name,
            barcode=str(ws.cell(barcode_row, 3).value or "") if barcode_row else "",
            tax_type=str(ws.cell(barcode_row, 4).value or "") if barcode_row else "",
            order_type=str(ws.cell(row, 5).value or ""),
            order_qty=to_int(ws.cell(row, 7).value),
            available_qty=to_int(ws.cell(row, 8).value),
            inbound_qty=to_int(ws.cell(row, 9).value),
            purchase_price=to_int(ws.cell(row, 10).value),
            supply_price=to_int(ws.cell(row, 11).value),
            vat_unit=to_int(ws.cell(row, 12).value),
            order_amount=to_int(ws.cell(row, 13).value),
            supply_amount=to_int(ws.cell(row, 14).value),
            vat_amount=to_int(ws.cell(row, 15).value),
            inbound_amount=to_int(ws.cell(row, 17).value),
            inbound_supply_amount=to_int(ws.cell(row, 18).value),
            inbound_vat_amount=to_int(ws.cell(row, 19).value),
            shipment_group=item.shipment_group if item else "",
            note=item.note if item else "",
            length_mm=item.length_mm if item else 0,
            width_mm=item.width_mm if item else 0,
            height_mm=item.height_mm if item else 0,
            weight_g=item.weight_g if item else 0,
            inner_qty=item.inner_qty if item else 0,
            box_qty=item.box_qty if item else 0,
            pallet_layer_qty=item.pallet_layer_qty if item else 0,
            pallet_qty=item.pallet_qty if item else 0,
        )
        lines.append(line)
        row += 2
    return lines


def line_is_available(line: PoLine, master: dict[str, MasterItem]) -> bool:
    item = master.get(line.sku_id)
    if item is None:
        return True
    return item.available


def create_processed_po(
    po_path: Path,
    output_path: Path,
    master: dict[str, MasterItem],
) -> list[PoLine]:
    wb = load_workbook(po_path)
    ws = wb.active
    lines = read_po_lines(po_path, master)
    delete_rows: list[int] = []

    for line in lines:
        if line_is_available(line, master):
            continue
        delete_rows.extend([line.row_number, line.row_number + 1])

    for row in sorted(delete_rows, reverse=True):
        ws.delete_rows(row)

    refresh_total_row(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return read_po_lines(output_path, master)


def refresh_total_row(ws: Worksheet) -> None:
    header_row = find_po_header(ws)
    total_row = None
    for row in range(header_row + 2, ws.max_row + 1):
        if "합계" in str(ws.cell(row, 1).value or ""):
            total_row = row
            break
    if total_row is None:
        return

    product_rows = []
    row = header_row + 2
    while row < total_row:
        if ws.cell(row, 2).value and ws.cell(row, 3).value:
            product_rows.append(row)
        row += 2

    for col in [7, 8, 9, 13, 14, 15]:
        ws.cell(total_row, col).value = sum(to_int(ws.cell(row, col).value) for row in product_rows)


def write_summary_workbook(lines: list[PoLine], output_path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "스큐별 합산"
    detail = wb.create_sheet("PO별 상세")
    shipment = wb.create_sheet("쉽먼트 초안")
    guide = wb.create_sheet("사용 안내")

    grouped: dict[str, dict[str, Any]] = defaultdict(lambda: {
        "product_name": "",
        "barcode": "",
        "qty": 0,
        "amount": 0,
        "supply": 0,
        "vat": 0,
        "po_nos": set(),
        "shipment_group": "",
        "note": "",
        "length_mm": 0,
        "width_mm": 0,
        "height_mm": 0,
        "weight_g": 0,
        "inner_qty": 0,
        "box_qty": 0,
        "pallet_layer_qty": 0,
        "pallet_qty": 0,
    })
    for line in lines:
        row = grouped[line.sku_id]
        row["product_name"] = row["product_name"] or line.product_name
        row["barcode"] = row["barcode"] or line.barcode
        row["qty"] += line.available_qty
        row["amount"] += line.order_amount
        row["supply"] += line.supply_amount
        row["vat"] += line.vat_amount
        row["po_nos"].add(line.po_no)
        row["shipment_group"] = row["shipment_group"] or line.shipment_group
        row["note"] = row["note"] or line.note
        row["length_mm"] = row["length_mm"] or line.length_mm
        row["width_mm"] = row["width_mm"] or line.width_mm
        row["height_mm"] = row["height_mm"] or line.height_mm
        row["weight_g"] = row["weight_g"] or line.weight_g
        row["inner_qty"] = row["inner_qty"] or line.inner_qty
        row["box_qty"] = row["box_qty"] or line.box_qty
        row["pallet_layer_qty"] = row["pallet_layer_qty"] or line.pallet_layer_qty
        row["pallet_qty"] = row["pallet_qty"] or line.pallet_qty

    summary.append(["SKU ID(스큐 아이디)", "상품명", "바코드", "납품가능수량 합계", "발주금액 합계", "공급가액 합계", "부가세 합계", "포함 PO 개수", "포함 PO 번호", "길이(mm)", "넓이(mm)", "높이(mm)", "중량(g)", "쉽먼트", "비고"])
    for sku_id, row in sorted(grouped.items()):
        summary.append([sku_id, row["product_name"], row["barcode"], row["qty"], row["amount"], row["supply"], row["vat"], len(row["po_nos"]), ", ".join(sorted(row["po_nos"])), row["length_mm"], row["width_mm"], row["height_mm"], row["weight_g"], row["shipment_group"], row["note"]])
    summary.append(["합계", "", "", sum(line.available_qty for line in lines), sum(line.order_amount for line in lines), sum(line.supply_amount for line in lines), sum(line.vat_amount for line in lines), "", "", "", "", "", "", "", ""])

    detail.append(["PO 번호", "파일명", "물류센터", "입고예정일", "원본 No.", "SKU ID(스큐 아이디)", "상품명", "바코드", "발주수량", "납품가능수량", "매입가", "발주금액", "길이(mm)", "넓이(mm)", "높이(mm)", "중량(g)", "쉽먼트", "비고"])
    for line in lines:
        detail.append([line.po_no, line.source_file, line.center, line.inbound_date, line.original_no, line.sku_id, line.product_name, line.barcode, line.order_qty, line.available_qty, line.purchase_price, line.order_amount, line.length_mm, line.width_mm, line.height_mm, line.weight_g, line.shipment_group, line.note])

    shipment.append(["쉽먼트", "SKU ID(스큐 아이디)", "상품명", "납품가능수량", "길이(mm)", "넓이(mm)", "높이(mm)", "중량(g)", "Inner 수량", "Box 총수량", "Pallet 1단 수량", "Pallet 총수량", "예상 파렛트 수", "오버행 기준", "포함 PO 번호", "비고"])
    for sku_id, row in sorted(grouped.items(), key=lambda item: (str(item[1]["shipment_group"]), item[0])):
        pallet_qty = row["pallet_qty"]
        estimated_pallets = round(row["qty"] / pallet_qty, 2) if pallet_qty else ""
        overhang_note = pallet_overhang_note(row["length_mm"], row["width_mm"])
        shipment.append([row["shipment_group"], sku_id, row["product_name"], row["qty"], row["length_mm"], row["width_mm"], row["height_mm"], row["weight_g"], row["inner_qty"], row["box_qty"], row["pallet_layer_qty"], row["pallet_qty"], estimated_pallets, overhang_note, ", ".join(sorted(row["po_nos"])), row["note"]])

    guide.append(["용어", "읽는 법", "뜻", "언제 쓰는지"])
    guide.append(["PO", "피오", "쿠팡 발주서", "쿠팡이 납품 요청한 파일을 말할 때 씁니다."])
    guide.append(["SKU ID", "스큐 아이디", "같은 상품인지 구분하는 상품 고유번호", "여러 PO에 같은 상품이 나뉘었을 때 수량을 합칠 때 씁니다."])
    guide.append(["Upload", "업로드", "내 컴퓨터 파일을 사이트에 올리는 일", "심플웍스에 엑셀 파일을 등록할 때 씁니다."])
    guide.append(["Shipment", "쉽먼트", "납품 묶음 또는 배송 계획", "쿠팡에 어떤 납품 묶음으로 보낼지 정리할 때 씁니다."])
    guide.append([])
    guide.append(["처리일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    guide.append(["PO별 상품 줄 수", len(lines)])
    guide.append(["스큐 아이디 수", len(grouped)])
    guide.append(["총 납품가능수량", sum(line.available_qty for line in lines)])
    guide.append(["총 발주금액", sum(line.order_amount for line in lines)])

    for ws in [summary, detail, shipment, guide]:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            new_font = copy(cell.font)
            new_font.bold = True
            cell.font = new_font
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_google_csv(lines: list[PoLine], output_path: Path) -> None:
    grouped: dict[str, dict[str, Any]] = defaultdict(lambda: {"product_name": "", "qty": 0, "amount": 0, "po_nos": set()})
    for line in lines:
        row = grouped[line.sku_id]
        row["product_name"] = row["product_name"] or line.product_name
        row["qty"] += line.available_qty
        row["amount"] += line.order_amount
        row["po_nos"].add(line.po_no)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["SKU ID(스큐 아이디)", "상품명", "납품가능수량 합계", "발주금액 합계", "포함 PO 번호"])
        for sku_id, row in sorted(grouped.items()):
            writer.writerow([sku_id, row["product_name"], row["qty"], row["amount"], ", ".join(sorted(row["po_nos"]))])


def main() -> None:
    parser = argparse.ArgumentParser(description="쿠팡 PO 반자동 처리")
    parser.add_argument("--po-dir", required=True, help="원본 PO 엑셀 파일이 들어 있는 폴더")
    parser.add_argument("--master", required=True, help="납품 가능/불가능을 체크한 기초자료 엑셀")
    parser.add_argument("--out-dir", required=True, help="결과 파일을 저장할 폴더")
    args = parser.parse_args()

    po_dir = Path(args.po_dir)
    master_path = Path(args.master)
    out_dir = Path(args.out_dir)
    processed_dir = out_dir / "심플웍스_업로드용_PO복사본"

    master = read_master(master_path)
    all_processed_lines: list[PoLine] = []

    po_files = sorted(path for path in po_dir.glob("*.xlsx") if not path.name.startswith("~$") and "복사본" not in path.name)
    if not po_files:
        raise ValueError("원본 PO 폴더에 처리할 .xlsx 파일이 없습니다.")

    for po_file in po_files:
        output_file = processed_dir / po_file.name.replace(".xlsx", " - 복사본.xlsx")
        processed_lines = create_processed_po(po_file, output_file, master)
        all_processed_lines.extend(processed_lines)

    print("Done")
    print(f"PO files: {len(po_files)}")
    print(f"Processed lines: {len(all_processed_lines)}")
    print(f"Output folder: {processed_dir}")


if __name__ == "__main__":
    main()
